#users/views.py
import hashlib, openpyxl, pandas as pd, calendar
from io import BytesIO
from decimal import Decimal,InvalidOperation
from multiprocessing import context
from urllib import request
from decimal import Decimal, InvalidOperation
from .decorators import permiso_finanzas_required
from colegios.forms import PublicacionForm, GaleriaForm
from django.forms import inlineformset_factory
from .models import Persona, RelacionFamiliar, Usuario, Rol, BilleteraCantina, RubroCantina, VentaCantina, DetalleVentaCantina, RecargaBilletera
from academico.models import DocenteDetalle, RepresentanteDetalle, Seccion, Inscripcion, EstudianteDetalle, AnioEscolar, AdministrativoDetalle, Pago, TasaCambio, Asignatura, CargaAcademica, PlanEvaluacion, NotaCuantitativa, InformeCualitativo, NotaCualitativa
from colegios.models import Colegio, ImagenGaleria, Publicacion
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, date, timedelta
from academico.services import obtener_tasa_vigente
from django.shortcuts import redirect, render, get_object_or_404
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.utils.text import slugify
from django.views.decorators.http import require_POST
from django.views.decorators.cache import never_cache
from django.contrib.auth import logout
from django.contrib.auth.views import LoginView
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import make_password
from django.contrib import messages
from django.shortcuts import redirect
from django.conf import settings
from django.db import transaction
from django.db.models import Q, Prefetch, Count, Sum
from django.core.exceptions import PermissionDenied, ValidationError
from reportlab.pdfbase import pdfdoc
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string, get_template
from xhtml2pdf import pisa
from django.urls import reverse
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage

def link_callback(uri, rel):
    """
    Convierte las rutas de los archivos en rutas locales para pisa.
    """
    # Si la URL es de Cloudinary (empieza con http), pisa la puede manejar
    # Si es una URL interna, intenta resolverla
    if uri.startswith(settings.MEDIA_URL):
        path = os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ""))
    elif uri.startswith(settings.STATIC_URL):
        path = os.path.join(settings.STATIC_ROOT, uri.replace(settings.STATIC_URL, ""))
    else:
        return uri  # Devuelve la URL original si es de Cloudinary
    return path

@method_decorator(never_cache, name='dispatch')
class ColegioLoginView(LoginView):
    template_name = 'users/login.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Obtenemos el colegio por el slug de la URL
        slug = self.kwargs.get('colegio_slug')        
        # Pasamos el objeto colegio al template para usar sus colores y logo
        context['colegio'] = Colegio.objects.get(slug=slug)
        return context
    
    def get_success_url(self):
        # Redirige al dashboard del colegio que acaba de loguearse
        slug = self.kwargs.get('colegio_slug')
        return f'/{slug}/dashboard/'
    
@never_cache
def logout_colegio(request, colegio_slug):
    logout(request)  # Limpia la sesión del usuario de forma segura
    return redirect('login_colegio', colegio_slug=colegio_slug)

@login_required
def dashboard_colegio(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)

    # 🔒 Seguridad Básica de entrada al Colegio
    if not request.user.is_superuser and request.user.colegio != colegio:
        raise PermissionDenied
    
    # 🔀 TRÁFICO DE ROLES: Redirecciones inmediatas a sus propios módulos
    if request.user.rol:
        if request.user.rol.nombre == 'Cantinero':
            return redirect('dashboard_cantina', colegio_slug=colegio.slug)
        
        if request.user.rol.nombre == 'Docente':
            return redirect('dashboard_docente', colegio_slug=colegio.slug)

    # === DATOS COMUNES (Solo para Representantes y Administradores) ===
    ultima_tasa = TasaCambio.objects.filter(moneda='USD').order_by('-fecha').first()
    tasa_precio = ultima_tasa.precio if ultima_tasa else None
    ultimas_publicaciones = colegio.publicaciones.all().order_by('-fecha_creacion')[:5]

    context = {
        'colegio': colegio,
        'tasa_actual': tasa_precio,
        'ultimas_publicaciones': ultimas_publicaciones,
    }

    anio_activo = AnioEscolar.objects.filter(colegio=colegio, activo=True).order_by('-id').first()

    # === BIFURCACIÓN REPRESENTANTE VS ADMINISTRADOR ===
    if request.user.rol and request.user.rol.nombre == 'Representante':
        # --- LÓGICA EXCLUSIVA PARA EL REPRESENTANTE ---
        representante_persona = getattr(request.user, 'perfil', None)
        
        if representante_persona:
            # 1. Obtenemos los IDs de todos sus representados históricos
            estudiantes_ids = RelacionFamiliar.objects.filter(
                representante=representante_persona
            ).values_list('estudiante_id', flat=True)
            
            if anio_activo:
                # 2. 🔥 FILTRADO CRUCIAL: Solo estudiantes con inscripción ACTIVA este año
                estudiantes = Persona.objects.filter(
                    id__in=estudiantes_ids,
                    inscripciones__anio_escolar=anio_activo,
                    inscripciones__estado='ACTIVO'
                ).distinct()
                
                # 3. Traemos las inscripciones correspondientes para armar el mapa de secciones
                inscripciones = Inscripcion.objects.filter(
                    estudiante_id__in=estudiantes_ids,
                    anio_escolar=anio_activo,
                    estado='ACTIVO'
                ).select_related('seccion')
                
                mapa_secciones = {ins.estudiante_id: ins.seccion for ins in inscripciones}
                
                # 4. Evaluamos solvencia solo para los inscritos activos
                for estudiante in estudiantes:
                    estudiante.seccion = mapa_secciones.get(estudiante.id)
                    tiene_deudas = Pago.objects.filter(
                        estudiante=estudiante,
                        anio_escolar=anio_activo,
                        pagado=False,
                        activo=True
                    ).exists()
                    estudiante.solvente = not tiene_deudas
            else:
                # Si no hay año escolar activo, no mostramos ningún representado activo
                estudiantes = Persona.objects.none()
            
            context['sus_representados'] = estudiantes
        else:
            context['sus_representados'] = Persona.objects.none()

    else:
        # --- LÓGICA EXCLUSIVA PARA ADMIN / SUPER ---
        context['total_publicaciones'] = colegio.publicaciones.count()
        context['total_fotos'] = colegio.imagenes.count()
        context['imagenes'] = colegio.imagenes.all().order_by('-id')[:6] 
        
        if anio_activo:
            context['personas'] = Persona.objects.filter(
                colegio=colegio,
                es_estudiante=True,
                inscripciones__anio_escolar=anio_activo,
                inscripciones__estado='ACTIVO'
            ).distinct()
            
            context['pagos_por_conciliar'] = Pago.objects.filter(
                colegio=colegio,
                anio_escolar=anio_activo,
                pagado=False,
                activo=True,
                num_referencia__isnull=False,
                fecha_pago__isnull=False,
            ).exclude(num_referencia='').count()
            
            context['estudiantes_por_inscribir'] = Persona.objects.filter(
                colegio=colegio,
                es_estudiante=True
            ).exclude(
                inscripciones__anio_escolar=anio_activo,
                inscripciones__estado='ACTIVO'
            ).count()
        else:
            context['personas'] = Persona.objects.none()
            context['pagos_por_conciliar'] = 0
            context['estudiantes_por_inscribir'] = 0        

    return render(request, 'users/dashboard.html', context)

@login_required
def dashboard_docente(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    # Seguridad básica de pertenencia
    if not request.user.is_superuser and request.user.colegio != colegio:
        raise PermissionDenied
        
    if not request.user.rol or request.user.rol.nombre != 'Docente':
        raise PermissionDenied

    anio_activo = AnioEscolar.objects.filter(colegio=colegio, activo=True).first()
    
    # Recuperar el perfil de Persona asociado
    docente_perfil = getattr(request.user, 'perfil', None)
    
    # Traer las asignaciones del periodo activo
    asignaciones = []
    if docente_perfil and anio_activo:
        asignaciones = CargaAcademica.objects.filter(
            docente=docente_perfil,
            seccion__anio_escolar=anio_activo
        ).select_related('asignatura', 'seccion')

    return render(request, 'users/dashboard_docente.html', {
        'colegio': colegio,
        'anio_activo': anio_activo,
        'asignaciones': asignaciones,
        'docente': docente_perfil
    })

@login_required
def ver_curso_docente(request, colegio_slug, asignacion_id):
    # 1. Recuperamos la carga académica con toda su jerarquía
    carga = get_object_or_404(
        CargaAcademica.objects.select_related('seccion', 'asignatura', 'docente'), 
        id=asignacion_id
    )
    colegio = carga.seccion.colegio
    lapso_actual = request.GET.get('lapso', '1')
    seccion = carga.seccion

    # 🔒 Control de seguridad
    if not request.user.is_superuser and carga.docente != request.user.perfil:
        raise PermissionDenied
    
    # =====================================================================
    # FLUJO 1: SECUNDARIA (Media General / Media Técnica)
    # =====================================================================
    if seccion.es_cuantitativo:
        # Al evaluar actividad por actividad, aquí solo mostramos el plan de evaluación del lapso
        plan_actividades = PlanEvaluacion.objects.filter(carga_academica=carga, lapso=lapso_actual)
        porcentaje_total = plan_actividades.aggregate(Sum('ponderacion'))['ponderacion__sum'] or 0

        return render(request, 'users/secundaria_dashboard_curso.html', {
            'colegio': colegio,
            'asignacion': carga,
            'seccion': seccion,
            'lapso_actual': lapso_actual,
            'plan_actividades': plan_actividades,
            'porcentaje_total': porcentaje_total,
        })
    
    # =====================================================================
    # FLUJO 2: INICIAL Y PRIMARIA (Evaluación Cualitativa Continua)
    # =====================================================================
    # A este punto solo llega Primaria/Inicial porque Secundaria ya hizo "return" arriba
    estudiantes = Persona.objects.filter(
        colegio=colegio, 
        es_estudiante=True, 
        inscripciones__seccion=seccion, 
        inscripciones__estado='ACTIVO'
    ).order_by('apellido', 'nombre').distinct()
    
    # 3. PROCESAMIENTO DEL FORMULARIO (POST) - Solo Inicial / Primaria
    if request.method == 'POST':
        for estudiante in estudiantes:
            observacion = request.POST.get(f'observacion_{estudiante.id}', '').strip()
            descriptor = request.POST.get(f'descriptor_{estudiante.id}', None)
            informe_texto = request.POST.get(f'informe_{estudiante.id}', '').strip()

            # Si el docente interactuó con la fila, guardamos o actualizamos
            if descriptor or informe_texto or observacion:
                InformeCualitativo.objects.update_or_create(
                    estudiante=estudiante,
                    carga_academica=carga,
                    lapso=lapso_actual,
                    defaults={
                        'descriptor': descriptor if descriptor != "" else None,
                        'informe_descriptivo': informe_texto if informe_texto else None,
                        'observaciones_recomendaciones': observacion if observacion else None
                    }
                )

        messages.success(request, f"Calificaciones del {lapso_actual}er Lapso guardadas correctamente.")
        return redirect(request.path + f"?lapso={lapso_actual}")

    # 4. CARGA DE DATOS EXISTENTES PARA LA INTERFAZ (GET) - Solo Inicial / Primaria
    notas_existentes = {}
    informes_qs = InformeCualitativo.objects.filter(
        carga_academica=carga,
        lapso=lapso_actual
    )
    for inf in informes_qs:
        notas_existentes[inf.estudiante_id] = {
            'descriptor': inf.descriptor or '',
            'informe': inf.informe_descriptivo or '',
            'observacion': inf.observaciones_recomendaciones or ''
        }

    # Adjuntamos las notas cualitativas previamente salvadas a los estudiantes
    for estudiante in estudiantes:
        datos = notas_existentes.get(estudiante.id, {})
        estudiante.descriptor_guardado = datos.get('descriptor', '')
        estudiante.informe_guardado = datos.get('informe', '')
        estudiante.observacion_guardada = datos.get('observacion', '')

    return render(request, 'users/ver_curso.html', {
        'colegio': colegio,
        'asignacion': carga,
        'estudiantes': estudiantes,
        'seccion': seccion,
        'lapso_actual': lapso_actual,
    })

@login_required
def gestionar_plan_evaluacion(request, colegio_slug, asignacion_id):
    """Permite al docente agregar o eliminar cortes de evaluación para un lapso."""
    carga = get_object_or_404(CargaAcademica, id=asignacion_id)
    lapso_actual = request.GET.get('lapso', '1')
    
    if not request.user.is_superuser and carga.docente != request.user.perfil:
        raise PermissionDenied

    if request.method == 'POST':
        descripcion = request.POST.get('descripcion')
        ponderacion = request.POST.get('ponderacion')
        
        # Validar que no sobrepase el 100% del lapso
        porcentaje_actual = PlanEvaluacion.objects.filter(carga_academica=carga, lapso=lapso_actual).aggregate(Sum('ponderacion'))['ponderacion__sum'] or 0
        if porcentaje_actual + int(ponderacion) > 100:
            messages.error(request, f"Error: La ponderación total no puede superar el 100%. Actualmente tienes {porcentaje_actual}%.")
        else:
            PlanEvaluacion.objects.create(
                carga_academica=carga,
                lapso=lapso_actual,
                descripcion=descripcion,
                ponderacion=int(ponderacion)
            )
            messages.success(request, "Actividad de evaluación añadida al plan.")
        
        return redirect(f"{request.path}?lapso={lapso_actual}")

    plan_actividades = PlanEvaluacion.objects.filter(carga_academica=carga, lapso=lapso_actual)
    porcentaje_total = plan_actividades.aggregate(Sum('ponderacion'))['ponderacion__sum'] or 0

    return render(request, 'users/secundaria_plan.html', {
        'colegio': carga.seccion.colegio,
        'asignacion': carga,
        'lapso_actual': lapso_actual,
        'plan_actividades': plan_actividades,
        'porcentaje_total': porcentaje_total
    })

@login_required
def cargar_notas_actividad(request, colegio_slug, asignacion_id, plan_id):
    """Pantalla de carga masiva de notas (1 al 20) para una actividad específica."""
    carga = get_object_or_404(CargaAcademica, id=asignacion_id)
    plan = get_object_or_404(PlanEvaluacion, id=plan_id, carga_academica=carga)
    colegio = carga.seccion.colegio
    
    if not request.user.is_superuser and carga.docente != request.user.perfil:
        raise PermissionDenied

    estudiantes = Persona.objects.filter(
        colegio=colegio, es_estudiante=True, 
        inscripciones__seccion=carga.seccion, inscripciones__estado='ACTIVO'
    ).order_by('apellido', 'nombre').distinct()

    if request.method == 'POST':
        for est in estudiantes:
            nota_req = request.POST.get(f'nota_{est.id}')
            obs_req = request.POST.get(f'observacion_{est.id}', '')
            
            if nota_req: # Solo guarda si ingresaron un número
                NotaCuantitativa.objects.update_or_create(
                    estudiante=est,
                    plan_evaluacion=plan,
                    defaults={
                        'nota': int(nota_req),
                        'observacion': obs_req.strip()
                    }
                )
        messages.success(request, f"Notas de '{plan.descripcion}' actualizadas con éxito.")
        return redirect('ver_curso_docente', colegio_slug=colegio.slug, asignacion_id=carga.id)

    # Mapeo de notas existentes para mostrarlas en los inputs
    notas_qs = NotaCuantitativa.objects.filter(plan_evaluacion=plan)
    notas_map = {n.estudiante_id: n for n in notas_qs}
    
    for est in estudiantes:
        registro = notas_map.get(est.id)
        est.nota_guardada = registro.nota if registro else ""
        est.obs_guardada = registro.observacion if registro else ""

    return render(request, 'users/secundaria_carga_notas.html', {
        'colegio': colegio,
        'asignacion': carga,
        'plan': plan,
        'estudiantes': estudiantes
    })

@login_required
def exportar_plantilla_excel(request, colegio_slug, asignacion_id):
    """Genera dinámicamente el formato Excel optimizado según el flujo del curso. CALIFICACIONES INICIAL/PRIMARIA: Descriptor + Informe + Observaciones. CALIFICACIONES SECUNDARIA: Nota numérica + Observación."""
    carga = get_object_or_404(CargaAcademica, id=asignacion_id)
    seccion = carga.seccion
    lapso_actual = request.GET.get('lapso', '1')
    plan_id = request.GET.get('plan_id')  # Requerido solo para Cuantitativo

    # Control de seguridad habitual
    if not request.user.is_superuser and carga.docente != request.user.perfil:
        raise PermissionDenied

    estudiantes = Persona.objects.filter(
        colegio=carga.seccion.colegio, es_estudiante=True, 
        inscripciones__seccion=seccion, inscripciones__estado='ACTIVO'
    ).order_by('apellido', 'nombre').distinct()

    # Creación del libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Carga de Notas"

    # Estilos profesionales para cabecera (Color Slate-800)
    header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    # Estructurar columnas según el nivel académico
    if seccion.es_cuantitativo:
        plan = get_object_or_404(PlanEvaluacion, id=plan_id, carga_academica=carga)
        headers = ["ID Estudiante", "Cédula", "Estudiante", f"Nota (1-20) - {plan.descripcion}", "Observación"]
        ws.append(headers)
        
        notas_map = {n.estudiante_id: n for n in NotaCuantitativa.objects.filter(plan_evaluacion=plan)}
        for est in estudiantes:
            reg = notas_map.get(est.id)
            ws.append([est.id, est.cedula, f"{est.apellido}, {est.nombre}", reg.nota if reg else "", reg.observacion if reg else ""])
    else:
        # Cualitativo (Inicial y Primaria)
        if seccion.nivel == 'PRIMARIA':
            headers = ["ID Estudiante", "Cédula", "Estudiante", "Descriptor (I, EP, C)", "Informe Descriptivo", "Observaciones"]
        else: # INICIAL
            headers = ["ID Estudiante", "Cédula", "Estudiante", "Informe Descriptivo", "Observaciones"]
        ws.append(headers)
        
        informes_map = {i.estudiante_id: i for i in InformeCualitativo.objects.filter(carga_academica=carga, lapso=lapso_actual)}
        for est in estudiantes:
            reg = informes_map.get(est.id)
            if seccion.nivel == 'PRIMARIA':
                ws.append([est.id, est.cedula, f"{est.apellido}, {est.nombre}", reg.descriptor if reg else "", reg.informe_descriptivo if reg else "", reg.observaciones_recomendaciones if reg else ""])
            else:
                ws.append([est.id, est.cedula, f"{est.apellido}, {est.nombre}", reg.informe_descriptivo if reg else "", reg.observaciones_recomendaciones if reg else ""])

    # Aplicar estilos estilizados a cabeceras
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Ajustar ancho de columnas automáticamente para mejor visibilidad
    ws.column_dimensions['A'].width = 15  # Mantiene oculto o visible el ID para match seguro
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 25

    nombre_archivo = f"plantilla_{slugify(carga.asignatura.nombre)}_{seccion.grado}_{seccion.nombre}.xlsx"
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response


@login_required
@require_POST
def importar_plantilla_excel(request, colegio_slug, asignacion_id):
    """Procesa el archivo subido, valida datos e impacta en la BD usando transacciones. CALIFICACIONES"""
    carga = get_object_or_404(CargaAcademica, id=asignacion_id)
    seccion = carga.seccion
    lapso_actual = request.POST.get('lapso', '1')
    plan_id = request.POST.get('plan_id')

    archivo = request.FILES.get('archivo_excel')
    if not archivo:
        messages.error(request, "Debe cargar un archivo válido.")
        return redirect(request.META.get('HTTP_REFERER', 'dashboard_docente'))

    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)
        ws = wb.active
        
        with transaction.atomic():
            # Recorrer filas omitiendo cabecera
            for row in ws.iter_rows(min_row=2, values_only=True):
                estudiante_id = row[0]
                if not estudiante_id:
                    continue
                
                estudiante = get_object_or_404(Persona, id=estudiante_id, es_estudiante=True)
                
                if seccion.es_cuantitativo:
                    plan = get_object_or_404(PlanEvaluacion, id=plan_id, carga_academica=carga)
                    nota_val = row[3]
                    obs_val = row[4] if len(row) > 4 else ""
                    
                    if nota_val is not None:
                        nota_int = int(nota_val)
                        if 1 <= nota_int <= 20:
                            NotaCuantitativa.objects.update_or_create(
                                estudiante=estudiante, plan_evaluacion=plan,
                                defaults={'nota': nota_int, 'observacion': str(obs_val).strip() if obs_val else ""}
                            )
                else:
                    if seccion.nivel == 'PRIMARIA':
                        desc_val = str(row[3]).strip().upper() if row[3] else None
                        inf_val = row[4]
                        obs_val = row[5] if len(row) > 5 else ""
                    else: # INICIAL
                        desc_val = None
                        inf_val = row[3]
                        obs_val = row[4] if len(row) > 4 else ""
                        
                    if desc_val or inf_val or obs_val:
                        # Validación del descriptor oficial
                        if desc_val and desc_val not in ['I', 'EP', 'C']:
                            desc_val = None
                            
                        InformeCualitativo.objects.update_or_create(
                            estudiante=estudiante, carga_academica=carga, lapso=lapso_actual,
                            defaults={
                                'descriptor': desc_val,
                                'informe_descriptivo': str(inf_val).strip() if inf_val else None,
                                'observaciones_recomendaciones': str(obs_val).strip() if obs_val else None
                            }
                        )
                        
        messages.success(request, "¡Sincronización de calificaciones completada con éxito!")
    except Exception as e:
        messages.error(request, f"Error de lectura en archivo Excel: {str(e)}")

    return redirect(request.META.get('HTTP_REFERER', 'dashboard_docente'))

@login_required
def reporte_final_curso(request, colegio_slug, asignacion_id):
    """Muestra la sábana consolidada con las definitivas de los 3 Lapsos."""
    carga = get_object_or_404(CargaAcademica, id=asignacion_id)
    seccion = carga.seccion
    
    estudiantes = Persona.objects.filter(
        colegio=carga.seccion.colegio, es_estudiante=True,
        inscripciones__seccion=seccion, inscripciones__estado='ACTIVO'
    ).order_by('apellido', 'nombre').distinct()
    
    reporte_data = []
    
    if seccion.es_cuantitativo:
        # Ponderar actividades por lapso y calcular definitivas reales
        planes = PlanEvaluacion.objects.filter(carga_academica=carga)
        notas = NotaCuantitativa.objects.filter(plan_evaluacion__in=planes).select_related('plan_evaluacion')
        
        mapa_notas = {est.id: {'1': 0.0, '2': 0.0, '3': 0.0, 'has': {'1': False, '2': False, '3': False}} for est in estudiantes}
        
        for n in notas:
            if n.estudiante_id in mapa_notas:
                lapso = n.plan_evaluacion.lapso
                # Sumatoria del peso: (Nota * Ponderación) / 100
                peso = (n.nota * n.plan_evaluacion.ponderacion) / 100.0
                mapa_notas[n.estudiante_id][lapso] += peso
                mapa_notas[n.estudiante_id]['has'][lapso] = True
                
        for est in estudiantes:
            d = mapa_notas[est.id]
            l1 = round(d['1']) if d['has']['1'] else None
            l2 = round(d['2']) if d['has']['2'] else None
            l3 = round(d['3']) if d['has']['3'] else None
            
            valores_validos = [v for v in [l1, l2, l3] if v is not None]
            promedio_anual = round(sum(valores_validos) / len(valores_validos)) if valores_validos else None
            
            reporte_data.append({
                'estudiante': est, 'lapso1': l1 or '--', 'lapso2': l2 or '--', 'lapso3': l3 or '--', 'final': promedio_anual or '--'
            })
    else:
        # Cualitativa
        informes = InformeCualitativo.objects.filter(carga_academica=carga)
        mapa_inf = {est.id: {'1': None, '2': None, '3': None} for est in estudiantes}
        
        for inf in informes:
            if inf.estudiante_id in mapa_inf:
                mapa_inf[inf.estudiante_id][inf.lapso] = inf
                
        for est in estudiantes:
            reporte_data.append({
                'estudiante': est, 'lapso1': mapa_inf[est.id]['1'], 'lapso2': mapa_inf[est.id]['2'], 'lapso3': mapa_inf[est.id]['3']
            })
            
    return render(request, 'users/reporte_final_curso.html', {
        'colegio': seccion.colegio, 'asignacion': carga, 'seccion': seccion, 'reporte_data': reporte_data, 'es_cuantitativo': seccion.es_cuantitativo
    })

@login_required
def ver_boleta_estudiante(request, colegio_slug, asignacion_id, estudiante_id):
    """Genera la boleta informativa oficial e integral de un estudiante."""
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    estudiante = get_object_or_404(Persona, id=estudiante_id, colegio=colegio, es_estudiante=True)
    inscripcion = get_object_or_404(Inscripcion, estudiante=estudiante, estado='ACTIVO', anio_escolar__activo=True)
    seccion = inscripcion.seccion
    
    # Extraer todas las asignaturas dictadas en la sección del alumno
    cargas = CargaAcademica.objects.filter(seccion=seccion).select_related('asignatura')
    boleta_rendimiento = []
    
    if seccion.es_cuantitativo:
        for c in cargas:
            planes = PlanEvaluacion.objects.filter(carga_academica=c)
            notas = NotaCuantitativa.objects.filter(plan_evaluacion__in=planes, estudiante=estudiante)
            
            lapsos = {'1': 0.0, '2': 0.0, '3': 0.0, 'has': {'1': False, '2': False, '3': False}}
            for n in notas:
                lapsos[n.plan_evaluacion.lapso] += (n.nota * n.plan_evaluacion.ponderacion) / 100.0
                lapsos['has'][n.plan_evaluacion.lapso] = True
                
            l1 = round(lapsos['1']) if lapsos['has']['1'] else '--'
            l2 = round(lapsos['2']) if lapsos['has']['2'] else '--'
            l3 = round(lapsos['3']) if lapsos['has']['3'] else '--'
            
            validos = [v for v in [l1, l2, l3] if isinstance(v, int)]
            def_anual = round(sum(validos) / len(validos)) if validos else '--'
            
            boleta_rendimiento.append({
                'asignatura': c.asignatura.nombre, 'l1': l1, 'l2': l2, 'l3': l3, 'final': def_anual
            })
    else:
        for c in cargas:
            informes = InformeCualitativo.objects.filter(carga_academica=c, estudiante=estudiante)
            inf_map = {i.lapso: i for i in informes}
            boleta_rendimiento.append({
                'asignatura': c.asignatura.nombre, 'l1': inf_map.get('1'), 'l2': inf_map.get('2'), 'l3': inf_map.get('3')
            })
            
    return render(request, 'users/boleta_estudiante.html', {
        'colegio': colegio,
        'estudiante': estudiante,
        'seccion': seccion, 
        'rendimiento': boleta_rendimiento, 
        'es_cuantitativo': seccion.es_cuantitativo, 
        'anio': inscripcion.anio_escolar, 
        'asignacion_id': asignacion_id
    })

@login_required
def imprimir_todas_boletas(request, colegio_slug, asignacion_id):
    """Genera una sola vista que agrupa las boletas de todos los estudiantes para impresión masiva."""
    carga_actual = get_object_or_404(CargaAcademica, id=asignacion_id)
    seccion = carga_actual.seccion
    colegio = seccion.colegio
    
    # Conseguir todas las asignaturas dictadas en esta sección para construir boletas completas
    cargas_seccion = CargaAcademica.objects.filter(seccion=seccion).select_related('asignatura')
    
    # Obtener la lista de estudiantes inscritos y activos
    estudiantes = Persona.objects.filter(
        colegio=colegio, es_estudiante=True,
        inscripciones__seccion=seccion, inscripciones__estado='ACTIVO'
    ).order_by('apellido', 'nombre').distinct()
    
    boletas_data = []
    
    if seccion.es_cuantitativo:
        # --- PROCESAMIENTO SECUNDARIA (CUANTITATIVO) ---
        from django.db.models import Sum
        planes = PlanEvaluacion.objects.filter(carga_academica__in=cargas_seccion)
        notas = NotaCuantitativa.objects.filter(plan_evaluacion__in=planes).select_related('plan_evaluacion')
        
        mapa_notas = {
            est.id: {
                c.id: {'1': 0.0, '2': 0.0, '3': 0.0, 'has': {'1': False, '2': False, '3': False}} 
                for c in cargas_seccion
            } for est in estudiantes
        }
        
        for n in notas:
            if n.estudiante_id in mapa_notas and n.plan_evaluacion.carga_academica_id in mapa_notas[n.estudiante_id]:
                lapso = n.plan_evaluacion.lapso
                c_id = n.plan_evaluacion.carga_academica_id
                peso = (n.nota * n.plan_evaluacion.ponderacion) / 100.0
                mapa_notas[n.estudiante_id][c_id][lapso] += peso
                mapa_notas[n.estudiante_id][c_id]['has'][lapso] = True
                
        for est in estudiantes:
            rendimiento = []
            for c in cargas_seccion:
                d = mapa_notas[est.id][c.id]
                l1 = round(d['1']) if d['has']['1'] else None
                l2 = round(d['2']) if d['has']['2'] else None
                l3 = round(d['3']) if d['has']['3'] else None
                
                valores_validos = [v for v in [l1, l2, l3] if v is not None]
                promedio_anual = round(sum(valores_validos) / len(valores_validos)) if valores_validos else None
                
                rendimiento.append({
                    'asignatura': c.asignatura.nombre, 'l1': l1, 'l2': l2, 'l3': l3, 'final': promedio_anual
                })
            boletas_data.append({'estudiante': est, 'rendimiento': rendimiento})
            
    else:
        # --- PROCESAMIENTO PRIMARIA / INICIAL (CUALITATIVO) ---
        informes = InformeCualitativo.objects.filter(carga_academica__in=cargas_seccion)
        
        mapa_inf = {
            est.id: {c.id: {'1': None, '2': None, '3': None} for c in cargas_seccion} 
            for est in estudiantes
        }
        
        for inf in informes:
            if inf.estudiante_id in mapa_inf and inf.carga_academica_id in mapa_inf[inf.estudiante_id]:
                mapa_inf[inf.estudiante_id][inf.carga_academica_id][inf.lapso] = inf
                
        for est in estudiantes:
            rendimiento = []
            for c in cargas_seccion:
                d_c = mapa_inf[est.id][c.id]
                rendimiento.append({
                    'asignatura': c.asignatura.nombre, 'l1': d_c['1'], 'l2': d_c['2'], 'l3': d_c['3']
                })
            boletas_data.append({'estudiante': est, 'rendimiento': rendimiento})

    return render(request, 'users/imprimir_todas_boletas.html', {
        'colegio': colegio,
        'seccion': seccion,
        'boletas_data': boletas_data,
        'es_cuantitativo': seccion.es_cuantitativo,
        'anio': "2025-2026", # Puedes computarlo o pasarlo de forma dinámica
    })

@login_required
def gestionar_anio_escolar(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    # 1. CONTROL DE SEGURIDAD INTERNO
    if request.user.rol.nombre not in ['Admin', 'Super']:
        raise PermissionDenied

    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip().upper()
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')
        
        # Helper interno para limpiar montos y convertirlos correctamente
        def limpiar_monto(valor):
            if not valor or valor.strip() == '':
                return None
            try:
                return float(valor.replace(',', '.')) # Manejo por si usan comas decimales
            except ValueError:
                return 0.0

        matricula1 = limpiar_monto(request.POST.get('matricula1'))
        matricula2 = limpiar_monto(request.POST.get('matricula2'))
        matricula3 = limpiar_monto(request.POST.get('matricula3'))
        mensualidad1 = limpiar_monto(request.POST.get('mensualidad1'))
        mensualidad2 = limpiar_monto(request.POST.get('mensualidad2'))
        mensualidad3 = limpiar_monto(request.POST.get('mensualidad3'))
        
        # Validaciones de consistencia
        if not nombre or not fecha_inicio or not fecha_fin:
            messages.error(request, "LOS CAMPOS DE NOMBRE, FECHA INICIO Y FECHA FIN SON OBLIGATORIOS.")
        elif fecha_inicio > fecha_fin:
            messages.error(request, "LA FECHA DE INICIO NO PUEDE SER POSTERIOR A LA FECHA DE FIN.")
        else:
            try:
                AnioEscolar.objects.create(
                    colegio=colegio,
                    nombre=nombre,
                    fecha_inicio=fecha_inicio,
                    fecha_fin=fecha_fin,
                    matricula1=matricula1,
                    matricula2=matricula2,
                    matricula3=matricula3,
                    mensualidad1=mensualidad1,
                    mensualidad2=mensualidad2,
                    mensualidad3=mensualidad3
                )
                messages.success(request, f"AÑO ESCOLAR '{nombre}' CREADO CORRECTAMENTE.")
                return redirect('gestionar_anio_escolar', colegio_slug=colegio.slug)
            except Exception as e:
                messages.error(request, f"ERROR AL GUARDAR EL PERIODO: {str(e).upper()}")
    
    anios_escolares = AnioEscolar.objects.filter(colegio=colegio).order_by('-fecha_inicio')
    
    return render(request, 'users/gestionar_anio_escolar.html', { 
        'colegio': colegio,
        'anios_escolares': anios_escolares
    })

    login_required
def editar_anio_escolar(request, colegio_slug, anio_id):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    if request.user.rol.nombre not in ['Admin', 'Super']:
        raise PermissionDenied
        
    anio = get_object_or_404(AnioEscolar, id=anio_id, colegio=colegio)
    
    if request.method == 'POST':
        anio.nombre = request.POST.get('nombre', '').strip().upper()
        anio.fecha_inicio = request.POST.get('fecha_inicio')
        anio.fecha_fin = request.POST.get('fecha_fin')
        
        def limpiar_monto(valor):
            if not valor or valor.strip() == '': return None
            try: return float(valor.replace(',', '.'))
            except ValueError: return 0.0

        anio.matricula1 = limpiar_monto(request.POST.get('matricula1'))
        anio.matricula2 = limpiar_monto(request.POST.get('matricula2'))
        anio.matricula3 = limpiar_monto(request.POST.get('matricula3'))
        anio.mensualidad1 = limpiar_monto(request.POST.get('mensualidad1'))
        anio.mensualidad2 = limpiar_monto(request.POST.get('mensualidad2'))
        anio.mensualidad3 = limpiar_monto(request.POST.get('mensualidad3'))
        
        if anio.fecha_inicio > anio.fecha_fin:
            messages.error(request, "LA FECHA DE INICIO NO PUEDE SER POSTERIOR A LA FECHA DE FIN.")
        else:
            anio.save()
            messages.success(request, f"PERIODO '{anio.nombre}' ACTUALIZADO CON ÉXITO.")
            
    return redirect('gestionar_anio_escolar', colegio_slug=colegio.slug)

@login_required
def cambiar_estado_anio(request, colegio_slug, anio_id):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    if request.user.rol.nombre not in ['Admin', 'Super']:
        raise PermissionDenied
        
    anio = get_object_or_404(AnioEscolar, id=anio_id, colegio=colegio)
    
    # Si se va a activar este año, desactivamos todos los demás del colegio primero
    if not anio.activo:
        AnioEscolar.objects.filter(colegio=colegio).update(activo=False)
        anio.activo = True
        messages.success(request, f"EL AÑO ESCOLAR {anio.nombre} HA SIDO ACTIVADO COMO PERIODO VIGENTE.")
    else:
        anio.activo = False
        messages.warning(request, f"EL AÑO ESCOLAR {anio.nombre} HA SIDO DESACTIVADO.")
        
    anio.save()
    return redirect('gestionar_anio_escolar', colegio_slug=colegio.slug)


@login_required
def exportar_anios_escolares_pdf(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    # Control de seguridad: Solo Admin o Super
    if request.user.rol.nombre not in ['Admin', 'Super']:
        raise PermissionDenied
        
    # Obtener el histórico de periodos escolares ordenados
    anios_escolares = AnioEscolar.objects.filter(colegio=colegio).order_by('-fecha_inicio')
    
    # Contexto que se enviará a la plantilla HTML
    context = {
        'colegio': colegio,
        'anios_escolares': anios_escolares,
        'usuario': request.user,
    }
    
    # Buscar y cargar la plantilla HTML específica para el PDF
    template = get_template('users/historico_anios_pdf.html')
    html_content = template.render(context)
    
    # Crear el buffer de memoria para almacenar el PDF resultante
    result = io.BytesIO()
    
    # Convertir el HTML a PDF usando xhtml2pdf
    pdf = pisa.pisaDocument(io.BytesIO(html_content.encode("UTF-8")), result)
    
    # Si no hubo errores en la generación, devolvemos el PDF
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        # 'inline' abre el PDF en el navegador, si prefieres descargarlo directo usa 'attachment'
        response['Content-Disposition'] = f'inline; filename="historico_periodos_{colegio.slug}.pdf"'
        return response
        
    return HttpResponse("Error interno al generar el reporte PDF", status=500)

def gestionar_asignaturas(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    # Opciones estáticas para los selectores del sistema educativo venezolano
    NIVELES_CHOICES = [
        ('INICIAL', 'Educación Inicial'),
        ('PRIMARIA', 'Educación Primaria'),
        ('MEDIA_GENERAL', 'Media General'),
        ('MEDIA_TECNICA', 'Media Técnica'),
    ]
    
    GRADOS_CHOICES = [
        ('Preescolar', 'Preescolar'),
        ('1ro', '1ro'), ('2do', '2do'), ('3ro', '3ro'),
        ('4to', '4to'), ('5to', '5to'), ('6to', '6to'),
        ('1er Año', '1er Año'), ('2do Año', '2do Año'), ('3er Año', '3er Año'),
        ('4to Año', '4to Año'), ('5to Año', '5to Año'), ('6to Año', '6to Año'),
    ]

    # --- PROCESAR POST (ACCIONES) ---
    if request.method == 'POST':
        accion = request.POST.get('accion')
        
        # A) CREAR / EDITAR
        if accion in ['crear', 'editar']:
            asignatura_id = request.POST.get('asignatura_id')
            codigo = request.POST.get('codigo', '').strip().upper()
            nombre = request.POST.get('nombre', '').strip()
            descripcion = request.POST.get('descripcion', '').strip()
            nivel = request.POST.get('nivel')
            grado = request.POST.get('grado')
            horas = int(request.POST.get('horas_semanales', 0))
            
            # Banderas booleanas (Checkbox)
            es_area = request.POST.get('es_area_desarrollo') == 'on'
            es_esp = request.POST.get('es_especialidad') == 'on'
            es_tec = request.POST.get('es_tecnica') == 'on'
            mencion = request.POST.get('mencion', '').strip() or None
            activo = request.POST.get('activo') == 'on'

            defaults = {
                'nombre': nombre,
                'descripcion': descripcion if descripcion else None,
                'nivel': nivel,
                'grado': grado,
                'horas_semanales': horas,
                'es_area_desarrollo': es_area,
                'es_especialidad': es_esp,
                'es_tecnica': es_tec,
                'mencion': mencion,
                'activo': activo,
            }

            if accion == 'crear':
                # Validar código único por colegio
                if Asignatura.objects.filter(colegio=colegio, codigo=codigo).exists():
                    messages.error(request, f"El código '{codigo}' ya está asignado a otra asignatura.")
                else:
                    Asignatura.objects.create(colegio=colegio, codigo=codigo, **defaults)
                    messages.success(request, "¡Asignatura creada de forma exitosa!")
            
            elif accion == 'editar':
                asignatura_obj = get_object_or_404(Asignatura, id=asignatura_id, colegio=colegio)
                # Validar si cambió el código y si el nuevo ya existe
                if asignatura_obj.codigo != codigo and Asignatura.objects.filter(colegio=colegio, codigo=codigo).exists():
                    messages.error(request, f"No se puede actualizar. El código '{codigo}' ya está en uso.")
                else:
                    for key, val in defaults.items():
                        setattr(asignatura_obj, key, val)
                    asignatura_obj.codigo = codigo
                    asignatura_obj.save()
                    messages.success(request, "¡Asignatura modificada correctamente!")

        # B) ELIMINAR (Con protección de relaciones activas)
        elif accion == 'eliminar':
            asignatura_id = request.POST.get('asignatura_id')
            asignatura_obj = get_object_or_404(Asignatura, id=asignatura_id, colegio=colegio)
            
            # Verificación explícita de vinculación a cargas académicas / secciones
            # Reemplaza 'cargaacademica_set' si el related_name en tu modelo es distinto
            tiene_vinculos = (
                getattr(asignatura_obj, 'cargaacademica_set', None) and asignatura_obj.cargaacademica_set.exists()
            ) or (
                getattr(asignatura_obj, 'secciones', None) and asignatura_obj.secciones.exists()
            )
            
            if tiene_vinculos:
                messages.error(request, f"No puedes eliminar '{asignatura_obj.nombre}' porque está vinculada a una o más Secciones activas.")
            else:
                try:
                    asignatura_obj.delete()
                    messages.success(request, "Asignatura removida permanentemente del catálogo.")
                except ProtectedError:
                    messages.error(request, "Error de protección: La asignatura posee integridad referencial activa.")

        return redirect('gestionar_asignaturas', colegio_slug=colegio.slug)

    # --- MANEJO DE GET & FILTROS ---
    nivel_sel = request.GET.get('nivel', 'TODOS')
    grado_sel = request.GET.get('grado', 'TODOS')
    busqueda = request.GET.get('q', '').strip()

    asignaturas = Asignatura.objects.filter(colegio=colegio).order_by('nivel', 'grado', 'nombre')

    if nivel_sel != 'TODOS':
        asignaturas = asignaturas.filter(nivel=nivel_sel)
    if grado_sel != 'TODOS':
        asignaturas = asignaturas.filter(grado=grado_sel)
    if busqueda:
        asignaturas = asignaturas.filter(nombre__icontains=busqueda) | asignaturas.filter(codigo__icontains=busqueda)

    context = {
        'colegio': colegio,
        'asignaturas': asignaturas,
        'niveles_opciones': NIVELES_CHOICES,
        'grados_opciones': GRADOS_CHOICES,
        'nivel_seleccionado': nivel_sel,
        'grado_seleccionado': grado_sel,
        'busqueda': busqueda,
    }
    return render(request, 'users/gestionar_asignaturas.html', context)

@login_required
def exportar_asignaturas_pdf(request, colegio_slug):    
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    # Capturar los mismos filtros de la pantalla de gestión
    nivel_sel = request.GET.get('nivel', 'TODOS')
    grado_sel = request.GET.get('grado', 'TODOS')
    busqueda = request.GET.get('q', '').strip()
    
    # Base del QuerySet
    asignaturas = Asignatura.objects.filter(colegio=colegio).order_by('nivel', 'grado', 'nombre')
    
    # Aplicar los filtros idénticos a la vista de gestión
    if nivel_sel != 'TODOS':
        asignaturas = asignaturas.filter(nivel=nivel_sel)
    if grado_sel != 'TODOS':
        asignaturas = asignaturas.filter(grado=grado_sel)
    if busqueda:
        asignaturas = asignaturas.filter(nombre__icontains=busqueda) | asignaturas.filter(codigo__icontains=busqueda)
    
    # Contexto para empaquetar los datos hacia la plantilla del PDF
    context = {
        'colegio': colegio,
        'asignaturas': asignaturas,
        'nivel_seleccionado': nivel_sel,
        'grado_seleccionado': grado_sel,
        'busqueda': busqueda,
    }
    
    # Cargar y renderizar la plantilla HTML específica para el catálogo
    template = get_template('users/catalogo_asignaturas_pdf.html') 
    html = template.render(context)
    
    # Preparar la respuesta HTTP tipo PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="catalogo_asignaturas_{colegio.slug}.pdf"'
    
    # Crear el PDF usando xhtml2pdf
    pisa_status = pisa.CreatePDF(html, dest=response)
    
    if pisa_status.err:
        return HttpResponse('Ocurrió un error al generar el PDF del catálogo', status=500)
        
    return response

@login_required
def gestionar_secciones(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    # 1. Traemos los años escolares (el último primero)
    anios_escolares = AnioEscolar.objects.filter(colegio=colegio).order_by('-id')
    
    # 2. Capturamos los filtros desde la URL (GET)
    periodo_id = request.GET.get('periodo')
    nivel_seleccionado = request.GET.get('nivel', 'TODOS')
    grado_seleccionado = request.GET.get('grado', 'TODOS')
    seccion_seleccionada = request.GET.get('seccion', 'TODOS')

    # Determinamos el año escolar por defecto (Activo o el último creado)
    if periodo_id:
        anio_seleccionado = anios_escolares.filter(id=periodo_id).first()
    else:
        anio_seleccionado = anios_escolares.filter(activo=True).first()
        if not anio_seleccionado:
            anio_seleccionado = anios_escolares.first()

    # 3. Base del QuerySet: Filtrar por colegio y año escolar seleccionado
    secciones = Seccion.objects.filter(
        colegio=colegio, 
        anio_escolar=anio_seleccionado
    ).select_related('docente_guia').annotate(
        total_inscritos=Count(
            'inscripcion', 
            filter=Q(inscripcion__estado='ACTIVO')
        )
    )
    
    # 4. Aplicamos los filtros dinámicos (Nivel, Grado y Sección)
    if nivel_seleccionado and nivel_seleccionado != 'TODOS':
        secciones = secciones.filter(nivel=nivel_seleccionado)
        
    if grado_seleccionado and grado_seleccionado != 'TODOS':
        secciones = secciones.filter(grado=grado_seleccionado)
        
    if seccion_seleccionada and seccion_seleccionada != 'TODOS':
        secciones = secciones.filter(nombre=seccion_seleccionada)
    
    # Ordenamos de forma ascendente por grado y sección
    secciones = secciones.order_by('grado', 'nombre') if anio_seleccionado else Seccion.objects.none()
    
    # Traemos los docentes y las opciones de niveles del modelo
    docentes = Persona.objects.filter(colegio=colegio, es_docente=True).order_by('apellido', 'nombre')
    niveles_opciones = Seccion.NIVEL_CHOICES

    # Procesar la creación de una sección mediante el Modal
    if request.method == 'POST':
        nivel = request.POST.get('nivel')
        grado = request.POST.get('grado')
        nombre = request.POST.get('nombre')
        capacidad = request.POST.get('capacidad')
        docente_guia_id = request.POST.get('docente_guia')
        es_cuantitativo = 'es_cuantitativo' in request.POST
        
        anio_id_form = request.POST.get('anio_escolar_id')
        anio_form = get_object_or_404(AnioEscolar, id=anio_id_form, colegio=colegio) if anio_id_form else anio_seleccionado

        if not anio_form:
            messages.error(request, 'NO SE PUEDE CREAR UNA SECCIÓN SIN UN PERÍODO ESCOLAR ACTIVO.')
            return redirect('gestionar_secciones', colegio_slug=colegio.slug)

        docente_guia = None
        if docente_guia_id:
            docente_guia = Persona.objects.filter(id=docente_guia_id, es_docente=True).first()

        try:
            nueva_seccion = Seccion.objects.create(
                colegio=colegio,
                anio_escolar=anio_form,
                nivel=nivel,
                grado=grado,
                nombre=nombre,
                capacidad=capacidad if capacidad else 35,
                es_cuantitativo=es_cuantitativo,
                docente_guia=docente_guia
            )
            messages.success(request, f'SECCIÓN {nueva_seccion.grado} "{nueva_seccion.nombre}" APERTURADA CON ÉXITO.')
            # Redirecciona aplicando los filtros exactos de la sección recién creada
            return redirect(f"{request.path}?periodo={anio_form.id}&nivel={nivel}&grado={grado}&seccion={nombre}")
            
        except Exception as e:
            messages.error(request, f'ERROR AL REGISTRAR LA SECCIÓN: {str(e)}')
            return redirect(f"{request.path}?periodo={anio_form.id}&nivel={nivel_seleccionado or 'TODOS'}&grado={grado_seleccionado or 'TODOS'}&seccion={seccion_seleccionada or 'TODOS'}")

    # 5. Enviamos los datos necesarios a la plantilla
    context = {
        'colegio': colegio,
        'secciones': secciones,
        'docentes': docentes,
        'anios_escolares': anios_escolares,
        'anio_seleccionado': anio_seleccionado,
        'niveles_opciones': niveles_opciones,
        'nivel_seleccionado': nivel_seleccionado or 'TODOS',
        'grado_seleccionado': grado_seleccionado or 'TODOS',
        'seccion_seleccionada': seccion_seleccionada or 'TODOS',
    }
    return render(request, 'users/gestionar_secciones.html', context)

@login_required
def editar_seccion(request, colegio_slug, seccion_id):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    seccion = get_object_or_404(Seccion, id=seccion_id, colegio=colegio)
    
    if request.method == 'POST':
        try:
            seccion.nivel = request.POST.get('nivel')
            seccion.grado = request.POST.get('grado')
            seccion.nombre = request.POST.get('nombre')
            seccion.capacidad = request.POST.get('capacidad', 35)
            seccion.es_cuantitativo = 'es_cuantitativo' in request.POST
            
            docente_guia_id = request.POST.get('docente_guia')
            if docente_guia_id:
                seccion.docente_guia = Persona.objects.filter(id=docente_guia_id, es_docente=True).first()
            else:
                seccion.docente_guia = None
                
            seccion.save()
            messages.success(request, f'SECCIÓN {seccion.grado} "{seccion.nombre}" ACTUALIZADA CON ÉXITO.')
        except Exception as e:
            messages.error(request, f'ERROR AL ACTUALIZAR LA SECCIÓN: {str(e)}')
            
    # Redirige manteniendo los filtros que estaban activos en la URL original
    return redirect(f"{request.META.get('HTTP_REFERER', 'gestionar_secciones')}")

@login_required
def eliminar_seccion(request, colegio_slug, seccion_id):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    seccion = get_object_or_404(Seccion, id=seccion_id, colegio=colegio)
    
    if request.method == 'POST':
        # --- VALIDACIÓN CRÍTICA DE ESTUDIANTES ---
        # Nota: Ajusta 'inscripcion_set' u 'persona_set' según el nombre 
        # de la relación inversa en tu modelo de Estudiantes/Inscripciones.
        tiene_estudiantes = seccion.inscripcion_set.exists() if hasattr(seccion, 'inscripcion_set') else False
        
        if tiene_estudiantes:
            messages.error(
                request, 
                f'NO SE PUEDE ELIMINAR LA SECCIÓN {seccion.grado} "{seccion.nombre}". '
                f'MOTIVO: TIENE ESTUDIANTES MATRICULADOS ASIGNADOS A ELLA.'
            )
        else:
            nombre_eliminado = f'{seccion.grado} "{seccion.nombre}"'
            seccion.delete()
            messages.success(request, f'SECCIÓN {nombre_eliminado} ELIMINADA CORRECTAMENTE DEL SISTEMA.')
            
    return redirect(f"{request.META.get('HTTP_REFERER', 'gestionar_secciones')}")

@login_required
def imprimir_listados_pdf(request, colegio_slug):    
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    # 1. Capturar todos los filtros de la URL (GET)
    periodo_id = request.GET.get('periodo')
    nivel_seleccionado = request.GET.get('nivel', 'TODOS')
    grado_seleccionado = request.GET.get('grado', 'TODOS')
    seccion_seleccionada = request.GET.get('seccion', 'TODOS')
    
    # 2. Determinar el Año Escolar correspondiente (Igual que en gestionar_secciones)
    anios_escolares = AnioEscolar.objects.filter(colegio=colegio).order_by('-id')
    if periodo_id:
        anio_seleccionado = anios_escolares.filter(id=periodo_id).first()
    else:
        anio_seleccionado = anios_escolares.filter(activo=True).first()
        if not anio_seleccionado:
            anio_seleccionado = anios_escolares.first()
            
    # 3. QuerySet Base: Filtrar secciones del colegio y del año escolar seleccionado
    secciones = Seccion.objects.filter(colegio=colegio, anio_escolar=anio_seleccionado)
    
    # 4. Aplicar los filtros dinámicos evitando la palabra 'TODOS'
    if nivel_seleccionado and nivel_seleccionado != 'TODOS':
        secciones = secciones.filter(nivel=nivel_seleccionado)
        
    if grado_seleccionado and grado_seleccionado != 'TODOS':
        secciones = secciones.filter(grado=grado_seleccionado)
        
    if seccion_seleccionada and seccion_seleccionada != 'TODOS':
        secciones = secciones.filter(nombre=seccion_seleccionada)
    
    # Ordenar las secciones para una impresión prolija
    secciones = secciones.order_by('grado', 'nombre')
    
    # 5. Iterar sobre las secciones filtradas para extraer ÚNICAMENTE estudiantes ACTIVOS
    for seccion in secciones:
        inscripciones_activas = seccion.inscripcion_set.filter(
            estado='ACTIVO',
            anio_escolar=anio_seleccionado
        ).select_related('estudiante').order_by('estudiante__apellido', 'estudiante__nombre')
        
        # Inyectar la lista limpia a la propiedad dinámica de la sección
        seccion.listado_estudiantes = [ins.estudiante for ins in inscripciones_activas]
    
    # 6. Renderizar y retornar el PDF
    context = {
        'colegio': colegio,
        'secciones': secciones,
        'anio_seleccionado': anio_seleccionado,
    }
    
    template = get_template('users/listado_estudiantes_pdf.html') 
    html = template.render(context)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="listados_{colegio.slug}.pdf"'
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    
    if pisa_status.err:
        return HttpResponse('Ocurrió un error al generar el PDF', status=500)
        
    return response

@login_required
def asignar_docentes_seccion(request, colegio_slug, seccion_id):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    seccion = get_object_or_404(Seccion, id=seccion_id, colegio=colegio)
    
    # 1. Filtramos solo los usuarios que pertenecen a este colegio y son docentes
    docentes_colegio = Persona.objects.filter(colegio=colegio, es_docente=True).order_by('apellido', 'nombre')
    
    # 2. Creamos el Inline FormSet relacionando Sección con CargaAcademica
    # extra=0 evita que aparezcan filas vacías para crear nuevas materias (solo editamos las existentes)
    CargaAcademicaFormSet = inlineformset_factory(
        Seccion,
        CargaAcademica,
        fields=('docente', 'docente_auxiliar'),
        extra=0,
        can_delete=False
    )
    
    if request.method == 'POST':
        formset = CargaAcademicaFormSet(request.POST, instance=seccion)
        
        # Personalizamos el queryset de los selectores en el POST para la validación interna de Django
        for form in formset.forms:
            form.fields['docente'].queryset = docentes_colegio
            form.fields['docente_auxiliar'].queryset = docentes_colegio
            
        if formset.is_valid():
            formset.save()
            messages.success(
                request, 
                f"CARGA ACADÉMICA Y DOCENTES DE {seccion.grado} SECCIÓN \"{seccion.nombre}\" ACTUALIZADOS CON ÉXITO."
            )
            # Redireccionamos de vuelta a la gestión general aplicando los filtros de la sección modificada
            url_destino = reverse('gestionar_secciones', kwargs={'colegio_slug': colegio.slug})
            return redirect(f"{url_destino}?periodo={seccion.anio_escolar.id}&nivel={seccion.nivel}&grado={seccion.grado}&seccion={seccion.nombre}")
        else:
            messages.error(request, "POR FAVOR, CORRIJA LOS ERRORES EN EL FORMULARIO.")
    else:
        # GET: Inicializamos el formset con las materias de la sección
        formset = CargaAcademicaFormSet(instance=seccion)
        
        # 3. Aplicamos estilos Tailwind y filtramos los dropdowns para cada fila individual
        for form in formset.forms:
            # Forzar a que solo muestre profesores de esta institución
            form.fields['docente'].queryset = docentes_colegio
            form.fields['docente_auxiliar'].queryset = docentes_colegio
            
            # Modificamos los labels por defecto para los selectores (opcional)
            form.fields['docente'].empty_label = "-- Seleccionar Docente Principal --"
            form.fields['docente_auxiliar'].empty_label = "-- Sin Auxiliar (Opcional) --"
            
            # Inyectamos clases estéticas de Tailwind a los widgets
            clase_select = "w-full p-2.5 bg-slate-50 border-none rounded-xl focus:ring-2 focus:ring-blue-500 font-bold text-slate-700 text-xs uppercase cursor-pointer"
            form.fields['docente'].widget.attrs.update({'class': clase_select})
            form.fields['docente_auxiliar'].widget.attrs.update({'class': clase_select})

    context = {
        'colegio': colegio,
        'seccion': seccion,
        'formset': formset,
    }
    return render(request, 'users/asignar_docentes_seccion.html', context)

@login_required
def historico_tasa_cambio(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    # Al entrar aquí, el Lazy Check verifica si requiere actualizarse
    tasa_actual = obtener_tasa_vigente()
    
    # Traemos todo el registro histórico ordenado por fecha descendente
    historico_tasas = TasaCambio.objects.filter(moneda="USD").order_by('-fecha')
    
    context = {
        'tasa_actual': tasa_actual,
        'historico_tasas': historico_tasas,
        'colegio': colegio,
    }
    return render(request, 'users/tasa_cambio_historico.html', context)

@login_required
def gestionar_pagos(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    user = request.user
    
    # 1. Parámetros de control y filtros
    query = request.GET.get('q', '').strip()
    grado_actual = request.GET.get('grado', '')
    seccion_actual = request.GET.get('seccion', '')
    nivel_actual = request.GET.get('nivel', '')
    anio_actual = request.GET.get('anio_escolar', '')
    
    order_by = request.GET.get('order', 'apellido')
    per_page = request.GET.get('per_page', '10')
    page = request.GET.get('page', 1)
    
    # Determinar año escolar por defecto (el activo)
    anio_activo = AnioEscolar.objects.filter(colegio=colegio, activo=True).order_by('-id').first()
    if not anio_actual and anio_activo:
        anio_actual = str(anio_activo.id)

    # 2. QuerySet Base enfocado en estudiantes inscritos
    estudiantes_list = Persona.objects.filter(
        colegio=colegio,
        es_estudiante=True
    )

    # Filtrar por el año escolar seleccionado a través de su Inscripción
    if anio_actual:
        estudiantes_list = estudiantes_list.filter(inscripciones__anio_escolar_id=anio_actual)

    # RESTRICCIÓN ROL REPRESENTANTE: Solo ve a sus representados asignados
    if user.rol.nombre == 'Representante':
        estudiantes_list = estudiantes_list.filter(tutores__representante__usuario=user)

    # 3. Aplicación de Filtros de Búsqueda
    if query:
        estudiantes_list = estudiantes_list.filter(
            Q(cedula__icontains=query) |
            Q(nombre__icontains=query) |
            Q(apellido__icontains=query) |
            Q(tutores__representante__nombre__icontains=query) |
            Q(tutores__representante__apellido__icontains=query)
        )
        
    if nivel_actual:
        estudiantes_list = estudiantes_list.filter(inscripciones__seccion__nivel=nivel_actual)
    if grado_actual:
        estudiantes_list = estudiantes_list.filter(inscripciones__seccion__grado=grado_actual)
    if seccion_actual:
        estudiantes_list = estudiantes_list.filter(inscripciones__seccion_id=seccion_actual)

    # Limpiar duplicados por Joins de relaciones
    estudiantes_list = estudiantes_list.distinct()

    # 4. Ordenamiento
    if order_by == 'nombre':
        estudiantes_list = estudiantes_list.order_by('nombre', 'apellido')
    elif order_by == 'cedula':
        estudiantes_list = estudiantes_list.order_by('cedula')
    else:
        estudiantes_list = estudiantes_list.order_by('apellido', 'nombre')

    # 5. Carga de los selectores de formularios
    secciones = Seccion.objects.filter(colegio=colegio).order_by('nombre')
    if nivel_actual:
        secciones = secciones.filter(nivel=nivel_actual)
        
    grados = secciones.values_list('grado', flat=True).distinct().order_by('grado')
    anios_escolares = AnioEscolar.objects.filter(colegio=colegio).order_by('-id')
    niveles = Seccion.objects.filter(colegio=colegio).values_list('nivel', flat=True).distinct().order_by('nivel')

    # 6. Paginador
    try: limit = int(per_page)
    except ValueError: limit = 10
        
    paginator = Paginator(estudiantes_list, limit)
    try: estudiantes_paginados = paginator.page(page)
    except PageNotAnInteger: estudiantes_paginados = paginator.page(1)
    except EmptyPage: estudiantes_paginados = paginator.page(paginator.num_pages)

    hoy = date.today()
    for est in estudiantes_paginados:
        # Buscamos solo los cobros/pagos que estén activos (no anulados)
        pagos_est = Pago.objects.filter(
            estudiante=est, 
            colegio=colegio, 
            anio_escolar_id=anio_actual, 
            activo=True
        )
        est.cuotas_pagadas = pagos_est.filter(pagado=True).count()
        est.cuotas_pendientes = pagos_est.filter(pagado=False).count()
        # Vencidos: cuotas por pagar cuya fecha límite de vencimiento es menor a hoy
        est.cuotas_vencidas = pagos_est.filter(pagado=False, fecha_vencimiento__lt=hoy).count()


    # =================================================================
    # 8. SOLUCIÓN: 'pagos_filtrados' SE ADAPTA A LOS FILTROS DE LA PANTALLA
    # =================================================================
    # Base de pagos activos para el colegio y año actual
    pagos_filtrados = Pago.objects.filter(
        colegio=colegio, 
        anio_escolar_id=anio_actual, 
        activo=True
    )

    # Si el administrador aplicó CUALQUIER filtro (Nivel, Grado, Sección o Buscador 'q'),
    # reducimos los pagos para que pertenezcan ÚNICAMENTE a los estudiantes resultantes.
    if nivel_actual or grado_actual or seccion_actual or query:
        pagos_filtrados = pagos_filtrados.filter(estudiante__in=estudiantes_list)


    # =================================================================
    # 9. CÁLCULOS FINANCIEROS GLOBALES (Actualizados y Dinámicos)
    # =================================================================
    # Total Recaudado (Pagos marcados como pagados)
    total_recaudado = pagos_filtrados.filter(pagado=True).aggregate(total=Sum('monto'))['total'] or 0
    
    # Total Pendiente (No pagados y que vencen hoy o en el futuro)
    total_pendiente = pagos_filtrados.filter(pagado=False, fecha_vencimiento__gte=hoy).aggregate(total=Sum('monto'))['total'] or 0
    
    # Total Vencido (Morosidad: no pagados y cuya fecha de vencimiento ya pasó)
    total_vencido = pagos_filtrados.filter(pagado=False, fecha_vencimiento__lt=hoy).aggregate(total=Sum('monto'))['total'] or 0
    
    # Total General del Periodo (Todo el dinero que debería entrar en este filtro)
    total_periodo = total_recaudado + total_pendiente + total_vencido
    
    # Calcular Porcentaje de Morosidad
    porcentaje_morosidad = 0
    if total_periodo > 0:
        porcentaje_morosidad = (total_vencido / total_periodo) * 100

    context = {
        'colegio': colegio,
        'query': query,
        'grado_actual': grado_actual,
        'seccion_actual': seccion_actual,
        'nivel_actual': nivel_actual,
        'anio_actual': anio_actual,
        'order_by': order_by,
        'per_page': str(limit),
        'grados': grados,
        'secciones': secciones,
        'anios_escolares': anios_escolares,
        'niveles': niveles,
        'estudiantes': estudiantes_paginados,
        'total_recaudado': total_recaudado,
        'total_pendiente': total_pendiente,
        'total_vencido': total_vencido,
        'total_periodo': total_periodo,      
        'porcentaje_morosidad': porcentaje_morosidad,
    }
    return render(request, 'users/gestionar_pagos.html', context)

@login_required
def estado_cuenta_estudiante(request, colegio_slug, estudiante_id):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    user = request.user
    estudiante = get_object_or_404(Persona, id=estudiante_id, colegio=colegio, es_estudiante=True)
    
    # Restricción de seguridad: si es representante, verificar que sea dueño del alumno
    if user.rol.nombre == 'Representante':
        if not RelacionFamiliar.objects.filter(estudiante=estudiante, representante__usuario=user).exists():
            messages.error(request, "No tiene permisos para ver el estado de cuenta solicitado.")
            return redirect('gestionar_pagos', colegio_slug=colegio.slug)
            
    anio_actual = request.GET.get('anio_escolar', '')
    anio_activo = AnioEscolar.objects.filter(colegio=colegio, activo=True).order_by('-id').first()
    if not anio_actual and anio_activo:
        anio_actual = str(anio_activo.id)
        
    anio_obj = get_object_or_404(AnioEscolar, id=anio_actual, colegio=colegio)
    
    # Consultar los pagos del estudiante en el periodo seleccionado
    pagos_base = Pago.objects.filter(estudiante=estudiante, colegio=colegio, anio_escolar=anio_obj, activo=True)
    
    hoy = date.today()
    pagos_pagados = pagos_base.filter(pagado=True).order_by('fecha_vencimiento')
    pagos_vencidos = pagos_base.filter(pagado=False, fecha_vencimiento__lt=hoy).order_by('fecha_vencimiento')
    pagos_pendientes = pagos_base.filter(pagado=False, fecha_vencimiento__gte=hoy).order_by('fecha_vencimiento')
    
    # Totales monetarios de apoyo
    # Al ser pagado=True, la cuota exonerada se resta automáticamente de aquí
    total_por_pagar = pagos_base.filter(pagado=False).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
    
    # Excluimos los exonerados del dinero real ingresado
    total_solventado = pagos_base.filter(pagado=True).exclude(num_referencia='EXONERADO').aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
    
    pagos_anulados = Pago.objects.filter(
        estudiante=estudiante, 
        colegio=colegio, 
        anio_escolar=anio_obj, 
        activo=False
    ).order_by('-fecha_actualizacion') # Ordenados del más reciente al más viejo
    
    # Conseguir el representante para mostrar su información básica
    relacion = RelacionFamiliar.objects.filter(estudiante=estudiante).first()
    representante_persona = relacion.representante if relacion else None

    # Agregamos la variable al contexto para que el HTML la pueda renderizar
    context = {
        'colegio': colegio,
        'estudiante': estudiante,
        'anio_obj': anio_obj,
        'pagos_pagados': pagos_pagados,
        'pagos_vencidos': pagos_vencidos,
        'pagos_pendientes': pagos_pendientes,
        'pagos_anulados': pagos_anulados,  
        'total_por_pagar': total_por_pagar,
        'total_solventado': total_solventado,
        'representante_persona': representante_persona,
        'anio_actual': anio_actual
    }
    return render(request, 'users/estado_cuenta_estudiante.html', context)

@login_required
def obtener_tasa_ajax(request, colegio_slug):
    """Endpoint para buscar la tasa de cambio por fecha vía AJAX"""
    fecha_str = request.GET.get('fecha')
    try:
        fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        tasa = TasaCambio.objects.filter(moneda='USD', fecha=fecha_obj).first()
        if tasa:
            return JsonResponse({'success': True, 'precio': float(tasa.precio)})
        else:
            # Si no hay tasa ese día, podemos buscar la última disponible como fallback
            ultima_tasa = TasaCambio.objects.filter(moneda='USD').order_by('-fecha').first()
            return JsonResponse({
                'success': False, 
                'precio': float(ultima_tasa.precio) if ultima_tasa else None,
                'msg': 'No hay tasa registrada para este día específico. Usando última conocida.'
            })
    except Exception as e:
        return JsonResponse({'success': False, 'precio': None, 'error': str(e)})

@login_required
def realizar_pago(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    user = request.user
    
    anio_activo = AnioEscolar.objects.filter(colegio=colegio, activo=True).order_by('-id').first()
    if not anio_activo:
        messages.error(request, "No se encontró un período escolar activo para este colegio.")
        return redirect('gestionar_pagos', colegio_slug=colegio.slug)

    estudiante_id_get = request.GET.get('estudiante_id')
    pago_id_get = request.GET.get('pago_id')
    
    cuota_obj = None
    estudiante_obj = None
    
    if pago_id_get:
        # Validación de seguridad: Si es Representante, el pago debe pertenecer a uno de sus tutorados
        if user.rol.nombre == 'Representante':
            cuota_obj = get_object_or_404(
                Pago, 
                id=pago_id_get, 
                colegio=colegio, 
                estudiante__tutores__representante__usuario=user
            )
        else:
            cuota_obj = get_object_or_404(Pago, id=pago_id_get, colegio=colegio)
            
        estudiante_obj = cuota_obj.estudiante
    elif estudiante_id_get:
        if user.rol.nombre == 'Representante':
            estudiante_obj = get_object_or_404(
                Persona, 
                id=estudiante_id_get, 
                es_estudiante=True, 
                tutores__representante__usuario=user
            )
        else:
            estudiante_obj = get_object_or_404(Persona, id=estudiante_id_get, es_estudiante=True)

    if user.rol.nombre == 'Representante':
        estudiantes_disponibles = Persona.objects.filter(
            es_estudiante=True, 
            tutores__representante__usuario=user,  
            inscripciones__anio_escolar=anio_activo,
            inscripciones__estado='ACTIVO'
        ).distinct().order_by('apellido', 'nombre')
    else:
        estudiantes_disponibles = Persona.objects.filter(
            es_estudiante=True,
            inscripciones__anio_escolar=anio_activo,
            inscripciones__estado='ACTIVO'
        ).distinct().order_by('apellido', 'nombre')

    if request.method == 'POST':
        pago_id_post = request.POST.get('pago_id') or pago_id_get
        estudiante_id = request.POST.get('estudiante') or (estudiante_obj.id if estudiante_obj else None)
        
        monto_usd_str = request.POST.get('monto', '0')
        fecha_pago = request.POST.get('fecha_pago')
        detalle = request.POST.get('detalle')

        try:
            monto_usd = Decimal(monto_usd_str)
        except InvalidOperation:
            monto_usd = Decimal('0')

        # CASO A: ACTUALIZAR REGISTRO EXISTENTE (SOLVENTAR COMPROMISO)
        if pago_id_post:
            if user.rol.nombre == 'Representante':
                pago_existente = get_object_or_404(
                    Pago, 
                    id=pago_id_post, 
                    colegio=colegio, 
                    estudiante__tutores__representante__usuario=user
                )
            else:
                pago_existente = get_object_or_404(Pago, id=pago_id_post, colegio=colegio)
                
            estudiante = pago_existente.estudiante
            monto_bs_str = request.POST.get('monto_bs', '0')
            
            try:
                pago_existente.monto_bs = Decimal(monto_bs_str)
                pago_existente.monto = monto_usd
                pago_existente.banco = request.POST.get('banco')
                pago_existente.titular_cuenta = request.POST.get('titular_cuenta')
                pago_existente.metodo_pago = request.POST.get('metodo_pago')
                pago_existente.num_referencia = request.POST.get('num_referencia')
                pago_existente.fecha_pago = fecha_pago
                if detalle:
                    pago_existente.detalle = detalle
                if request.FILES.get('comprobante'):
                    pago_existente.comprobante = request.FILES.get('comprobante')
                
                if user.rol.nombre == 'Representante':
                    pago_existente.pagado = False
                    mensaje = f"¡Pago de la cuota para {estudiante.nombre} reportado con éxito! Pendiente por conciliación."
                else:
                    pago_existente.pagado = True  
                    mensaje = f"¡Pago de la cuota para {estudiante.nombre} {estudiante.apellido} procesado con éxito!"

                pago_existente.usuario = user  # Guarda el usuario de la sesión actual

                pago_existente.save()
                
                messages.success(request, mensaje)
                if user.rol.nombre == 'Representante':
                    return redirect('estado_cuenta_estudiante', colegio_slug=colegio.slug, estudiante_id=estudiante.id) 
                else:
                    return redirect('gestionar_pagos', colegio_slug=colegio.slug)
            except Exception as e:
                messages.error(request, f"Ocurrió un error al actualizar el pago: {str(e)}")
        
        # CASO B: CREAR UN REGISTRO DESDE CERO (PAGO EXTRAORDINARIO/ESPECIAL)
        else:
            if not estudiante_id:
                messages.error(request, "Debe seleccionar un estudiante o la opción de pago masivo.")
            
            elif estudiante_id == 'todos' and user.rol.nombre != 'Representante':
                filtro_nivel = request.POST.get('filtro_nivel')
                filtro_grado = request.POST.get('filtro_grado')
                filtro_seccion = request.POST.get('filtro_seccion')

                inscripciones_destino = Inscripcion.objects.filter(anio_escolar=anio_activo, estado='ACTIVO')
                
                if filtro_seccion:
                    inscripciones_destino = inscripciones_destino.filter(seccion_id=filtro_seccion)
                else:
                    if filtro_nivel:
                        inscripciones_destino = inscripciones_destino.filter(seccion__nivel=filtro_nivel)
                    if filtro_grado:
                        inscripciones_destino = inscripciones_destino.filter(seccion__grado=filtro_grado)
                
                if not inscripciones_destino.exists():
                    messages.warning(request, "No se encontraron estudiantes activos que cumplan con los filtros seleccionados.")
                else:
                    try:
                        contador_pagos = 0
                        for inscripcion in inscripciones_destino:
                            estudiante = inscripcion.estudiante
                            
                            relacion = estudiante.tutores.first()
                            persona_representante = relacion.representante if relacion else None
                            usuario_representante = persona_representante.usuario if persona_representante else None

                            Pago.objects.create(
                                colegio=colegio,
                                anio_escolar=anio_activo,
                                estudiante=estudiante,
                                representante=usuario_representante,
                                usuario=user,
                                tipo_pago='extraordinario',
                                monto=monto_usd,
                                monto_bs=None, 
                                metodo_pago=None,
                                num_referencia=None,
                                banco=None,
                                titular_cuenta=None,
                                num_cuota='UNICA',
                                fecha_vencimiento=fecha_pago,
                                fecha_pago=None,
                                detalle=detalle,
                                comprobante=None,
                                activo=True,
                                pagado=False
                            )
                            contador_pagos += 1
                
                        messages.success(request, f"¡Pago extraordinario masivo registrado para {contador_pagos} estudiantes con éxito!")
                        return redirect('gestionar_pagos', colegio_slug=colegio.slug)
                    
                    except Exception as e:
                        messages.error(request, f"Error al generar los pagos masivos: {str(e)}")
            
            # Creación individual libre
            else:
                try:
                    estudiante = get_object_or_404(Persona, id=estudiante_id, es_estudiante=True)
                    
                    relacion = estudiante.tutores.first()
                    persona_representante = relacion.representante if relacion else None
                    usuario_representante = persona_representante.usuario if persona_representante else None
                    
                    Pago.objects.create(
                        colegio=colegio,
                        anio_escolar=anio_activo,
                        estudiante=estudiante,
                        representante=usuario_representante,
                        usuario=user,
                        tipo_pago='extraordinario',
                        monto=monto_usd,
                        monto_bs=None,
                        metodo_pago=None,
                        num_referencia=None,
                        banco=None,
                        titular_cuenta=None,
                        num_cuota='UNICA',
                        fecha_vencimiento=fecha_pago,
                        fecha_pago=None,
                        detalle=detalle,
                        pagado=False,
                        activo=True
                    )
                    messages.success(request, f"¡Pago extraordinario de {estudiante.nombre} {estudiante.apellido} registrado con éxito!")
                    return redirect('gestionar_pagos', colegio_slug=colegio.slug)
                    
                except Exception as e:
                    messages.error(request, f"Ocurrió un error al registrar el pago: {str(e)}")

    # --- FLUJO GET ---
    ultima_tasa = TasaCambio.objects.filter(moneda='USD').order_by('-fecha').first()
    tasa_precio = ultima_tasa.precio if ultima_tasa else None

    return render(request, 'users/realizar_pago.html', {
        'colegio': colegio,
        'anio_activo': anio_activo,
        'estudiantes': estudiantes_disponibles,
        'cuota_obj': cuota_obj,          
        'estudiante_obj': estudiante_obj,  
        'metodos_pago': Pago.METODOS_PAGO, 
        'bancos': Pago.BANCOS,
        'tasa_actual': tasa_precio,        
    })

@login_required
def editar_pago(request, colegio_slug, estudiante_id, pago_id):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    # Protección de seguridad perimetral
    if request.user.rol.nombre == 'Representante':
        messages.error(request, "No tienes permisos para realizar modificaciones de auditoría.")
        return redirect('estado_cuenta_estudiante', colegio_slug=colegio_slug, estudiante_id=estudiante_id)

    if request.method == 'POST':
        pago = get_object_or_404(Pago, id=pago_id, colegio=colegio, estudiante_id=estudiante_id)
        
        # Recolectamos la información del modal
        nuevo_detalle = request.POST.get('detalle')
        nuevo_monto = request.POST.get('monto')
        nueva_fecha_vencimiento = request.POST.get('fecha_vencimiento')
        nuevo_num_cuota = request.POST.get('num_cuota')
        nuevo_banco = request.POST.get('banco')
        nuevo_titular = request.POST.get('titular_cuenta')
        motivo_modificacion = request.POST.get('motivo') # Campo obligatorio de auditoría

        try:
            # Modificación de campos principales
            pago.detalle = nuevo_detalle
            pago.monto = nuevo_monto
            pago.num_cuota = nuevo_num_cuota            
            pago.banco = nuevo_banco
            pago.titular_cuenta = nuevo_titular
            
            if nueva_fecha_vencimiento:
                pago.fecha_vencimiento = nueva_fecha_vencimiento
                
            # Bloque de auditoría estricta
            pago.motivo = motivo_modificacion
            pago.usuario = request.user
            
            pago.save()
            messages.success(request, f"Cuota #{pago.id} modificada exitosamente. Se ha registrado en la bitácora.")
            
        except Exception as e:
            messages.error(request, f"Error al intentar modificar la cuota: {str(e)}")
            
        # Redirección limpia manteniendo el contexto del Año Escolar
        url_retorno = reverse('estado_cuenta_estudiante', kwargs={'colegio_slug': colegio.slug, 'estudiante_id': estudiante_id})
        return redirect(f"{url_retorno}?anio_escolar={pago.anio_escolar.id}")

    return redirect('estado_cuenta_estudiante', colegio_slug=colegio_slug, estudiante_id=estudiante_id)

@login_required
def auditar_accion_pago(request, colegio_slug, estudiante_id):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    # 1. Blindaje perimetral: si un representante intenta salirse con la suya, rebota aquí
    if request.user.rol.nombre == 'Representante':
        messages.error(request, "Acción denegada. No tienes permisos para realizar auditorías.")
        return redirect('estado_cuenta_estudiante', colegio_slug=colegio_slug, estudiante_id=estudiante_id)
        
    if request.method == 'POST':
        pago_id = request.POST.get('pago_id')
        accion = request.POST.get('accion')
        motivo = request.POST.get('motivo')
        
        # Obtenemos la cuota garantizando que pertenezca al colegio y al estudiante actual
        pago = get_object_or_404(Pago, id=pago_id, colegio=colegio, estudiante_id=estudiante_id)
        
        # 2. Obligatoriedad de Auditoría
        if not motivo or len(motivo.strip()) < 5:
            messages.error(request, "Debe especificar un motivo válido y explícito para la auditoría.")
            url_retorno = reverse('estado_cuenta_estudiante', kwargs={'colegio_slug': colegio.slug, 'estudiante_id': estudiante_id})
            return redirect(f"{url_retorno}?anio_escolar={pago.anio_escolar.id}")
            
        pago.usuario = request.user
        pago.motivo = motivo.strip()
        
        # 3. Lógica de negocio discriminada por la acción del formulario
        if accion == 'anular':
            pago.activo = False # Pasa a la bitácora de anulados automáticamente
            messages.success(request, f"Se ha ANULADO la cuota #{pago.id} con éxito.")
            
        elif accion == 'exonerar':
            pago.pagado = True 
            pago.num_referencia = 'EXONERADO' 
            messages.success(request, f"Se ha EXONERADO la cuota #{pago.id}.")
            
        else:
            messages.error(request, "La acción solicitada no es válida para el sistema de auditoría.")
            
        pago.save()
        
        # 4. Redirección limpia conservando el año escolar activo
        url_retorno = reverse('estado_cuenta_estudiante', kwargs={'colegio_slug': colegio.slug, 'estudiante_id': estudiante_id})
        return redirect(f"{url_retorno}?anio_escolar={pago.anio_escolar.id}")

    return redirect('estado_cuenta_estudiante', colegio_slug=colegio_slug, estudiante_id=estudiante_id)

@login_required
def imprimir_recibo(request, colegio_slug, pago_id):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    pago = get_object_or_404(Pago, id=pago_id, colegio=colegio, pagado=True)
    
    # Restricción de seguridad idéntica a tu gestión de estados de cuenta
    if request.user.rol.nombre == 'Representante':
        if not RelacionFamiliar.objects.filter(estudiante=pago.estudiante, representante__usuario=request.user).exists():
            return HttpResponse("No tiene permisos para ver o imprimir este documento.", status=403)
            
    tipo = request.GET.get('tipo', 'completo') # 'completo' o 'datos'
    
    # Extraer los datos del representante/cliente para la facturación
    relacion = RelacionFamiliar.objects.filter(estudiante=pago.estudiante).first()
    representante = relacion.representante if relacion else None

    # Simulación/Cálculo de valores fiscales (Ajustar según la tasa de tu base de datos si aplica)
    # Tomando como referencia los datos de la captura legal venezolana provista:
    tasa_cambio = Decimal('565.50') # Ejemplo de tasa BCV referencial si tu BD guarda en USD. falta llamarla del BCV o de tu modelo de configuración financiera si la tienes implementada.
    monto_usd = pago.monto
    monto_bs = monto_usd * tasa_cambio
    
    # Estructura de Totales Fiscales
    base_exenta_bs = monto_bs # Tratándose de matrículas/mensualidades escolares exentas de IVA
    total_factura_bs = monto_bs

    context = {
        'colegio': colegio,
        'pago': pago,
        'representante': representante,
        'tipo': tipo,
        'fecha_emision': pago.fecha_pago or timezone.now().date(),
        'tasa_cambio': tasa_cambio,
        'monto_bs': monto_bs,
        'base_exenta_bs': base_exenta_bs,
        'total_factura_bs': total_factura_bs,
    }
    
    # Renderizar el HTML estructurado
    html_string = render_to_string('users/recibo_pdf.html', context)
    
    # 1. Crear la respuesta HTTP con el tipo de contenido para PDF
    response = HttpResponse(content_type='application/pdf')
    filename = f"factura_{pago.id}.pdf" if tipo == 'completo' else f"factura_datos_{pago.id}.pdf"
    
    # 'inline' para que se abra en el navegador, o 'attachment' si prefieres descarga directa
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    
    # 2. Importar pisa localmente y compilar el HTML directo en la respuesta
    pisa_status = pisa.CreatePDF(html_string, dest=response)
    
    # 3. Si ocurre un error durante la compilación, avisar al sistema
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF con xhtml2pdf', status=500)
        
    return response

@permiso_finanzas_required
def panel_verificar_pagos(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    # Trae los pagos activos, no pagados y que ya tengan una referencia registrada por el representante
    pagos_pendientes = Pago.objects.filter(
        colegio=colegio,
        activo=True,
        pagado=False
    ).exclude(num_referencia__isnull=True).exclude(num_referencia="").order_by('fecha_registro')
    
    return render(request, 'pagos/panel_verificacion.html', {
        'colegio': colegio,
        'pagos_pendientes': pagos_pendientes,
    })

@permiso_finanzas_required
def procesar_verificacion_pago(request, colegio_slug, pago_id):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    pago = get_object_or_404(Pago, id=pago_id, colegio=colegio)
    
    if request.method == 'POST':
        accion = request.POST.get('accion')
        
        if accion == 'aprobar':
            pago.pagado = True
            pago.fecha_pago = timezone.now().date()
            pago.usuario = request.user  # Guardamos qué usuario administrativo lo aprobó
            pago.motivo = "Pago verificado de forma manual por administración."
            pago.save()
            messages.success(request, f"¡Pago de cuota {pago.num_cuota} de {pago.estudiante.nombre} aprobado con éxito!")
            
        elif accion == 'rechazar':
            motivo_rechazo = request.POST.get('motivo_rechazo', '').strip()
            if not motivo_rechazo:
                messages.error(request, "Debes ingresar un motivo para poder rechazar el pago.")
                return redirect('panel_verificar_pagos', colegio_slug=colegio.slug)
                
            # Limpiamos los datos del reporte para permitirle al representante volver a reportar
            pago.num_referencia = None
            pago.comprobante = None
            pago.monto_bs = None
            pago.metodo_pago = None
            pago.motivo = motivo_rechazo  # Le dejamos la nota de por qué se rebotó
            pago.usuario = request.user
            pago.save()
            messages.warning(request, f"El pago de {pago.estudiante.nombre} ha sido devuelto al representante.")
            
    return redirect('panel_verificar_pagos', colegio_slug=colegio.slug)

#esta funcion ahra la hacemos desde el gestionar_estudiantes
@login_required
def gestionar_inscripciones(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    # 1. Parámetros de control desde la URL (GET)
    pestana = request.GET.get('tab', 'activos')
    query = request.GET.get('q', '').strip()
    grado_actual = request.GET.get('grado', '')
    seccion_actual = request.GET.get('seccion', '')
    nivel_actual = request.GET.get('nivel', '')
    anio_actual = request.GET.get('anio_escolar', '')
    
    estado_inscripcion = 'ACTIVO' if pestana == 'activos' else 'RETIRADO'
    order_by = request.GET.get('order', 'apellido')
    per_page = request.GET.get('per_page', '10')
    page = request.GET.get('page', 1)
    
    # Determinar año escolar por defecto si no viene en los filtros
    anio_activo = AnioEscolar.objects.filter(colegio=colegio, activo=True).order_by('-id').first()
    if not anio_actual and anio_activo:
        anio_actual = str(anio_activo.id)

    # 2. QuerySet Base apuntando a Inscripcion
    inscripciones_list = Inscripcion.objects.filter(
        seccion__colegio=colegio, 
        estado=estado_inscripcion
    ).select_related('estudiante', 'seccion', 'anio_escolar')

    # 3. Aplicación de Filtros
    if query:
        inscripciones_list = inscripciones_list.filter(
            Q(estudiante__cedula__icontains=query) |
            Q(estudiante__nombre__icontains=query) |
            Q(estudiante__apellido__icontains=query)
        )
        
    if anio_actual:
        inscripciones_list = inscripciones_list.filter(anio_escolar_id=anio_actual)
        
    if nivel_actual:
        inscripciones_list = inscripciones_list.filter(seccion__nivel=nivel_actual)
        
    if grado_actual:
        inscripciones_list = inscripciones_list.filter(seccion__grado=grado_actual)
        
    if seccion_actual:
        inscripciones_list = inscripciones_list.filter(seccion_id=seccion_actual)

    # 4. Ordenamiento dinámico acoplado a la relación del estudiante
    if order_by == 'nombre':
        inscripciones_list = inscripciones_list.order_by('estudiante__nombre', 'estudiante__apellido')
    elif order_by == 'cedula':
        inscripciones_list = inscripciones_list.order_by('estudiante__cedula')
    else:
        inscripciones_list = inscripciones_list.order_by('estudiante__apellido', 'estudiante__nombre')

    # 5. Data para poblar los selectores del buscador
    secciones = Seccion.objects.filter(colegio=colegio).order_by('nombre')
    if nivel_actual:
        secciones = secciones.filter(nivel=nivel_actual)
        
    grados = secciones.values_list('grado', flat=True).distinct().order_by('grado')
    anios_escolares = AnioEscolar.objects.filter(colegio=colegio).order_by('-id')
    niveles = Seccion.objects.filter(colegio=colegio).values_list('nivel', flat=True).distinct().order_by('nivel')

    # 6. Paginador
    try:
        limit = int(per_page)
    except ValueError:
        limit = 10
        
    paginator = Paginator(inscripciones_list, limit)
    try:
        inscripciones = paginator.page(page)
    except PageNotAnInteger:
        inscripciones = paginator.page(1)
    except EmptyPage:
        inscripciones = paginator.page(paginator.num_pages)

    context = {
        'colegio': colegio,
        'pestana': pestana,
        'query': query,
        'grado_actual': grado_actual,
        'seccion_actual': seccion_actual,
        'nivel_actual': nivel_actual,
        'anio_actual': anio_actual,
        'order_by': order_by,
        'per_page': str(limit),
        'grados': grados,
        'secciones': secciones,
        'anios_escolares': anios_escolares,
        'niveles': niveles,
        'inscripciones': inscripciones,
    }
    return render(request, 'users/gestionar_inscripciones.html', context)


@login_required
def desactivar_inscripcion(request, colegio_slug, inscripcion_id):
    """ Cambia el estatus a RETIRADO, congela evaluaciones y suspende mensualidades futuras """
    if request.method == 'POST':
        inscripcion = get_object_or_404(Inscripcion, id=inscripcion_id, seccion__colegio__slug=colegio_slug)
        
        # 1. Cambiar estado de la inscripción
        inscripcion.estado = 'RETIRADO'
        inscripcion.save()
        
        # 2. Gestionar Estado de Cuenta (Suspensión de cuotas por realizar)
        hoy = timezone.now().date()
        
        # Buscamos cuotas que NO estén pagadas, estén activas y que venzan DESPUÉS de hoy
        pagos_a_suspender = Pago.objects.filter(
            estudiante=inscripcion.estudiante,
            anio_escolar=inscripcion.anio_escolar,
            pagado=False,
            activo=True,
            fecha_vencimiento__gt=hoy  # __gt significa "mayor que" (Futuras)
        )
        
        total_suspendidos = pagos_a_suspender.count()
        
        # Apagamos estas cuotas usando el método update masivo
        pagos_a_suspender.update(
            activo=False,
            usuario=request.user,
            motivo=f"Suspendido automáticamente por retiro del alumno el {hoy.strftime('%d/%m/%Y')}."
        )
        
        messages.success(
            request, 
            f"El estudiante {inscripcion.estudiante.apellido} fue marcado como RETIRADO. "
            f"Se suspendieron {total_suspendidos} mensualidades futuras. "
            f"Las cuotas vencidas antes de hoy permanecen activas para cobro."
        )
        
    return redirect('gestionar_inscripciones', colegio_slug=colegio_slug)


@login_required
def activar_inscripcion(request, colegio_slug, inscripcion_id):
    """ Reincorpora al estudiante y reactiva sus cobros pausados """
    inscripcion = get_object_or_404(Inscripcion, id=inscripcion_id, seccion__colegio__slug=colegio_slug)
    
    # 1. Regresar estado a ACTIVO
    inscripcion.estado = 'ACTIVO'
    inscripcion.save()
    
    # 2. Buscar y reactivar las cuotas que nosotros mismos suspendimos
    hoy = timezone.now().date()
    pagos_a_reactivar = Pago.objects.filter(
        estudiante=inscripcion.estudiante,
        anio_escolar=inscripcion.anio_escolar,
        pagado=False,
        activo=False,
        motivo__contains="Suspendido automáticamente por retiro"
    )
    
    total_reactivados = pagos_a_reactivar.count()
    
    pagos_a_reactivar.update(
        activo=True,
        usuario=request.user,
        motivo=f"Reactivado por reincorporación del alumno el {hoy.strftime('%d/%m/%Y')}."
    )
    
    messages.success(
        request, 
        f"Inscripción de {inscripcion.estudiante.nombre} REINCORPORADA. "
        f"Se reactivaron {total_reactivados} cuotas de mensualidad en su estado de cuenta."
    )
    return redirect('gestionar_inscripciones', colegio_slug=colegio_slug)

def agregar_meses(fecha_base, meses):
    """Suma meses a una fecha de forma segura controlando desbordamientos de días"""
    año = fecha_base.year + (fecha_base.month + meses - 1) // 12
    mes = (fecha_base.month + meses - 1) % 12 + 1
    dia = min(fecha_base.day, calendar.monthrange(año, mes)[1])
    return date(año, mes, dia)

@login_required
def crear_inscripcion(request, colegio_slug, estudiante_id=None):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    anio_activo = AnioEscolar.objects.filter(colegio=colegio, activo=True).first()
    
    if not anio_activo:
        messages.error(request, "No hay un Año Escolar activo configurado.")
        return redirect('gestionar_inscripciones', colegio_slug=colegio.slug)
        
    # Estudiante fijo seleccionado desde la pestaña 'sin_inscripcion'
    estudiante = get_object_or_404(Persona, id=estudiante_id, colegio=colegio, es_estudiante=True)
    secciones = Seccion.objects.filter(colegio=colegio, anio_escolar=anio_activo)
    
    # 1. Buscar la Persona que actúa como representante vinculado
    relacion = RelacionFamiliar.objects.filter(estudiante=estudiante).first()
    representante_persona = relacion.representante if relacion else None

    # 2. Obtener la instancia de Usuario (AUTH_USER_MODEL) desde esa Persona
    representante_usuario = representante_persona.usuario if representante_persona else None

    if request.method == 'POST':
        seccion_id = request.POST.get('seccion')
        seccion = get_object_or_404(Seccion, id=seccion_id, colegio=colegio, anio_escolar=anio_activo)
        
        if not representante_persona:
            messages.error(request, f"No se puede inscribir a {estudiante.nombre} sin un Representante asignado.")
            return redirect('gestionar_estudiantes', colegio_slug=colegio.slug)

        if Inscripcion.objects.filter(estudiante=estudiante, anio_escolar=anio_activo).exists():
            messages.error(request, f"El estudiante {estudiante.nombre} ya se encuentra inscrito en este período.")
            return redirect('gestionar_estudiantes', colegio_slug=colegio.slug)

        # 1. Obtener los costos base según el nivel de la sección asignada
        base_mat = Decimal('0.00')
        base_men = Decimal('0.00')
        
        if seccion.nivel == 'INICIAL':
            base_mat = anio_activo.matricula1 or Decimal('0.00')
            base_men = anio_activo.mensualidad1 or Decimal('0.00')
        elif seccion.nivel == 'PRIMARIA':
            base_mat = anio_activo.matricula2 or Decimal('0.00')
            base_men = anio_activo.mensualidad2 or Decimal('0.00')
        else: # MEDIA General o MEDIA Técnica
            base_mat = anio_activo.matricula3 or Decimal('0.00')
            base_men = anio_activo.mensualidad3 or Decimal('0.00')

        # 2. Capturar los porcentajes de descuento enviados por el formulario
        try:
            desc_mat_pct = Decimal(request.POST.get('descuento_matricula', '0') or '0')
            desc_men_pct = Decimal(request.POST.get('descuento_mensualidad', '0') or '0')
        except ValueError:
            desc_mat_pct = Decimal('0')
            desc_men_pct = Decimal('0')

        # 3. Calcular montos finales aplicando los descuentos de forma segura
        costo_mat_final = base_mat * (Decimal('1') - (desc_mat_pct / Decimal('100')))
        costo_men_final = base_men * (Decimal('1') - (desc_men_pct / Decimal('100')))
        
        # Aseguramos que los montos no sean negativos por error
        costo_mat_final = max(Decimal('0.00'), costo_mat_final)
        costo_men_final = max(Decimal('0.00'), costo_men_final)

        try:
            with transaction.atomic():
                # A. Crear la Inscripción guardando los costos personalizados con descuento
                nueva_inscripcion = Inscripcion.objects.create(
                    estudiante=estudiante,
                    seccion=seccion,
                    anio_escolar=anio_activo,
                    costo_matricula=costo_mat_final,
                    costo_mensualidad=costo_men_final,
                    estado='ACTIVO'
                )
                
                # B. Mover al alumno de "flotante" a activo actualizando su EstudianteDetalle
                detalle, created = EstudianteDetalle.objects.get_or_create(estudiante=estudiante)
                detalle.seccion = seccion
                # Guardamos opcionalmente el descuento de mensualidad en su ficha como historial de beca
                if desc_men_pct > 0:
                    detalle.becado = True
                    detalle.porcentaje_beca = int(desc_men_pct)
                detalle.save()
                
                # C. PLAN DE CUENTAS POR COBRAR (PAGOS)
                generar_cronograma_pagos(nueva_inscripcion, representante_usuario)

                messages.success(request, f"¡Inscripción exitosa! Se ha registrado a {estudiante.nombre} con su plan de financiamiento.")
                return redirect('gestionar_estudiantes', colegio_slug=colegio.slug)

        except Exception as e:
            messages.error(request, f"Error interno en la base de datos: {str(e)}")

    return render(request, 'users/crear_inscripcion.html', {
        'colegio': colegio,
        'anio_activo': anio_activo,
        'estudiante': estudiante,
        'secciones': secciones,
        'representante': representante_persona
    })

def generar_cronograma_pagos(inscripcion, representante_usuario):
    """
    Genera automáticamente 1 Matrícula ('UNICA') y 12 mensualidades consecutivas
    basadas en los costos asignados en la inscripción.
    """
    colegio = inscripcion.seccion.colegio
    anio_escolar = inscripcion.anio_escolar
    estudiante = inscripcion.estudiante
    fecha_inicio_clases = anio_escolar.fecha_inicio

    meses_es = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    # 1. CREAR LA MATRÍCULA
    Pago.objects.create(
        colegio=colegio,
        anio_escolar=anio_escolar,
        representante=representante_usuario,
        estudiante=estudiante,
        tipo_pago='matricula',
        monto=inscripcion.costo_matricula,  # Usa el costo real guardado
        num_cuota='UNICA',
        fecha_vencimiento=fecha_inicio_clases,
        detalle=f"Matrícula Única - Período {anio_escolar.nombre}",
        pagado=False
    )

    # 2. CREAR LAS 12 MENSUALIDADES
    for i in range(1, 13):
        vencimiento_cuota = agregar_meses(fecha_inicio_clases, i - 1)
        nombre_mes = meses_es[vencimiento_cuota.month - 1]

        Pago.objects.create(
            colegio=colegio,
            anio_escolar=anio_escolar,
            representante=representante_usuario,
            estudiante=estudiante,
            tipo_pago='mensualidad',
            monto=inscripcion.costo_mensualidad,  # Usa el costo real guardado
            num_cuota=str(i),
            fecha_vencimiento=vencimiento_cuota,
            detalle=f"Mensualidad Cuota {i}/12 - {nombre_mes} {vencimiento_cuota.year}",
            pagado=False
        )

@login_required
def gestor_pagos_representante(request):
    hoy = date.today()
    
    # 1. Calculamos el último día del mes actual para fijar el límite visible
    ultimo_dia_del_mes = calendar.monthrange(hoy.year, hoy.month)[1]
    fecha_limite_visible = date(hoy.year, hoy.month, ultimo_dia_del_mes)

    # 2. Traer las deudas que ya vencieron o vencen este mes
    cuotas_por_pagar = Pago.objects.filter(
        representante=request.user,
        activo=True,
        pagado=False,
        fecha_vencimiento__lte=fecha_limite_visible # Menor o igual al fin del mes actual
    ).order_by('fecha_vencimiento', 'id')

    # 3. Historial de lo que ya pagó exitosamente (opcional para mostrar recibos)
    pagos_realizados = Pago.objects.filter(
        representante=request.user,
        activo=True,
        pagado=True
    ).order_by('-fecha_pago')

    return render(request, 'pagos/gestor_pagos.html', {
        'cuotas_por_pagar': cuotas_por_pagar,
        'pagos_realizados': pagos_realizados,
    })

@login_required
def panel_verificar_pagos(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    # Traemos los pagos reportados que aún no se han verificado
    pagos_pendientes = Pago.objects.filter(
        colegio=colegio,
        activo=True,
        pagado=False
    ).exclude(num_referencia__isnull=True).exclude(num_referencia="").order_by('fecha_registro')
    
    return render(request, 'users/panel_verificacion.html', {
        'colegio': colegio,
        'pagos_pendientes': pagos_pendientes,
    })

@login_required 
def crear_publicacion(request, colegio_slug): 
    colegio = get_object_or_404(Colegio, slug=colegio_slug) # Obtenemos el colegio o mostramos un error 404 si no existe
    
    if request.method == 'POST': # Si el formulario ha sido enviado
        form = PublicacionForm(request.POST, request.FILES) # Creamos una instancia del formulario con los datos enviados por el usuario
        if form.is_valid(): # Si el formulario es válido, guardamos la nueva publicación pero sin confirmarla aún (commit=False)
            nueva_pub = form.save(commit=False) # Creamos la nueva publicación pero no la guardamos en la base de datos todavía
            nueva_pub.colegio = colegio  # Vinculamos la noticia al colegio actual
            nueva_pub.save() # Guardamos la nueva publicación en la base de datos
            return redirect('dashboard_colegio', colegio_slug=colegio.slug) # Redirigimos al dashboard del colegio después de crear la publicación
    else:
        form = PublicacionForm()    # Si el método no es POST, simplemente mostramos un formulario vacío para crear una nueva publicación
    
    return render(request, 'users/crear_publicacion.html', { 
        'form': form,
        'colegio': colegio
    }) # Renderizamos la plantilla para crear una nueva publicación, pasando el formulario y el colegio como contexto

@login_required # Decorador para asegurar que solo usuarios autenticados puedan acceder a esta vista
def cargar_galeria(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    if request.method == 'POST':
        form = GaleriaForm(request.POST, request.FILES)
        files = request.FILES.getlist('fotos') # Obtenemos la lista de archivos
        if form.is_valid():
            for f in files:
                # Creamos cada imagen vinculada al colegio
                ImagenGaleria.objects.create(colegio=colegio, imagen=f)
            return redirect('dashboard_colegio', colegio_slug=colegio.slug)
    else:
        form = GaleriaForm()
        
    return render(request, 'users/cargar_galeria.html', {
        'form': form,
        'colegio': colegio
    })

# Vista para gestionar el contenido del colegio (noticias e imágenes)
@login_required
def gestionar_contenido(request, colegio_slug): 
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    publicaciones = colegio.publicaciones.all().order_by('-fecha_creacion')
    imagenes = colegio.imagenes.all().order_by('-id')

    return render(request, 'users/gestionar_contenido.html', {
        'colegio': colegio,
        'publicaciones': publicaciones,
        'imagenes': imagenes,
    })

# Vista para editar una publicación existente, recibe el slug del colegio y el ID de la publicación a editar
@login_required 
def editar_publicacion(request, colegio_slug, pub_id): 
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    publicacion = get_object_or_404(Publicacion, id=pub_id, colegio=colegio)
    
    if request.method == 'POST':
        # Pasamos instance=publicacion para que Django sepa que estamos editando, no creando
        form = PublicacionForm(request.POST, request.FILES, instance=publicacion)
        if form.is_valid():
            form.save()
            return redirect('gestionar_contenido', colegio_slug=colegio_slug)
    else:
        form = PublicacionForm(instance=publicacion)
    
    return render(request, 'users/crear_publicacion.html', { # Reutilizamos el template de crear
        'form': form,
        'colegio': colegio,
        'editando': True
    })

# Vista rápida para eliminar noticias
@login_required
def eliminar_publicacion(request, colegio_slug, pub_id):
    pub = get_object_or_404(Publicacion, id=pub_id, colegio__slug=colegio_slug)
    pub.delete()
    return redirect('gestionar_contenido', colegio_slug=colegio_slug)

@login_required
def eliminar_imagen(request, colegio_slug, img_id):
    imagen = get_object_or_404(ImagenGaleria, id=img_id, colegio__slug=colegio_slug)
    imagen.delete()
    return redirect('gestionar_contenido', colegio_slug=colegio_slug)

def crear_rol(request, colegio_slug): # Vista para crear un nuevo rol dentro de un colegio específico
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    if request.method == 'POST':
        # Aquí capturamos los datos del formulario manual
        nombre_rol = request.POST.get('nombre')
        
        # Creamos el rol con los permisos marcados
        nuevo_rol = Rol.objects.create(
            nombre=nombre_rol,
            colegio=colegio,
            can_manage_news='can_manage_news' in request.POST,
            can_manage_gallery='can_manage_gallery' in request.POST,
            can_manage_staff='can_manage_staff' in request.POST,
            can_manage_students='can_manage_students' in request.POST,
            can_manage_grades='can_manage_grades' in request.POST,
            can_manage_finances='can_manage_finances' in request.POST,
            can_manage_canteen='can_manage_canteen' in request.POST,
        )
        return redirect('dashboard_colegio', colegio_slug=colegio.slug)
        
    return render(request, 'users/crear_rol.html', {'colegio': colegio})

@login_required
def crear_usuario_colegio(request, colegio_slug): # Vista para crear un nuevo usuario vinculado a un colegio específico
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    roles = Rol.objects.filter(colegio=colegio) # Solo los roles de este colegio
    
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        rol_id = request.POST.get('rol')
        try:
            with transaction.atomic(): 
                # 1. Crear Usuario
                nuevo_usuario = Usuario.objects.create_user(
                    email=email,
                    password=password,
                    colegio=colegio,
                    rol_id=rol_id
                )
        
                # 2. Creamos la Persona vinculada
                Persona.objects.create(
                    colegio=colegio,
                    cedula=request.POST.get('cedula'),
                    nombre=request.POST.get('nombre'),
                    apellido=request.POST.get('apellido'),
                    tipo=request.POST.get('tipo'),
                    usuario=nuevo_usuario
                )
          
            # Mensaje de éxito
            messages.success(request, f"Usuario {email} creado correctamente.")
            return redirect('gestionar_usuarios', colegio_slug=colegio.slug)
            
        except Exception as e:
            #  MENSAJE DE ERROR
            messages.error(request, f"No se pudo crear el usuario. Error: {e}")
            # Al no hacer redirect, el usuario vuelve al formulario con sus datos (si los manejas en el template)

    return render(request, 'users/crear_usuario.html', {
        'colegio': colegio,
        'roles': roles,
        'tipos_persona': Persona.TIPOS
    })

@login_required
def gestionar_usuarios(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    # Solo Admin o Super pueden entrar
    if getattr(request.user.rol, 'nombre', '') not in ['Admin', 'Super']:
        raise PermissionDenied

    # Filtramos usuarios por colegio y traemos su rol de una vez
    usuarios = Usuario.objects.filter(colegio=colegio).select_related('rol').order_by('-date_joined')

    roles = Rol.objects.all()
    
    return render(request, 'users/seguridad/gestionar_usuarios.html', {
        'colegio': colegio,
        'usuarios': usuarios,
        'roles': roles
    })


@login_required
def gestionar_roles(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)

    if getattr(request.user.rol, 'nombre', '') not in ['Admin', 'Super']:
        raise PermissionDenied

    roles = Rol.objects.filter(colegio=colegio)

    return render(request, 'users/seguridad/gestionar_roles.html', {
        'colegio': colegio,
        'roles': roles
    })

@login_required
@require_POST
def cambiar_rol_usuario(request, usuario_id):
    # Solo administradores o superusuarios pueden cambiar roles
    if request.user.rol.nombre not in ['Admin', 'Super']:
        messages.error(request, "No tienes permiso para realizar esta acción.")
        return redirect('gestionar_usuarios')
    
    usuario = get_object_or_404(Usuario, id=usuario_id)
    nuevo_rol_id = request.POST.get('nuevo_rol')
    
    if nuevo_rol_id:
        usuario.rol_id = nuevo_rol_id
        usuario.save()
        messages.success(request, f"Rol de {usuario.email} actualizado correctamente.")
    
    # Redirigimos a la gestión de usuarios del colegio actual
    return redirect('gestionar_usuarios', colegio_slug=usuario.colegio.slug if usuario.colegio else 'admin')

@login_required
def gestionar_personas(request, colegio_slug): 
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    filtro_tipo = request.GET.get('tipo')
    personas = Persona.objects.filter(colegio=colegio)
    
    if filtro_tipo:
        personas = personas.filter(tipo=filtro_tipo)
    
    return render(request, 'users/gestionar_personas.html', {
        'colegio': colegio,
        'personas': personas,
        'tipos': Persona.TIPOS,
        'filtro_actual': filtro_tipo
    })

@login_required
def gestionar_docentes(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)

    nombre_rol = getattr(request.user.rol, 'nombre', '')
    if nombre_rol not in ['Admin', 'Super', 'ADMINISTRADOR']:
        raise PermissionDenied   
    
    # Captura de parámetros de filtrado, orden y paginación
    estado_actual = request.GET.get('estado', 'activo')
    query = request.GET.get('q', '')
    orden = request.GET.get('orden', 'apellido')
    registros_por_pagina = request.GET.get('registros', '10')
    
    base_queryset = Persona.objects.filter(
        colegio=colegio, 
        tipo='DOCENTE'
    ).select_related('usuario', 'detalle_docente')

    # 1. Filtrar según la pestaña (Activo / Inactivo)
    if estado_actual == 'inactivo':
        docentes = base_queryset.filter(activo=False)
    else:
        docentes = base_queryset.filter(activo=True)
   
    # 2. Filtrar por barra de búsqueda
    if query:
        docentes = docentes.filter(
            Q(nombre__icontains=query) |
            Q(apellido__icontains=query) |
            Q(cedula__icontains=query) |
            Q(detalle_docente__especialidad__icontains=query)
        )

    # 3. Aplicar Criterio de Ordenamiento
    if orden == 'nombre':
        docentes = docentes.order_by('nombre', 'apellido')
    elif orden == 'cedula':
        docentes = docentes.order_by('cedula')
    elif orden == 'cargo':
        docentes = docentes.order_by('detalle_docente__especialidad')
    else:  # Por defecto: 'apellido'
        docentes = docentes.order_by('apellido', 'nombre')

    # 4. Procesar la Paginación
    try:
        limite = int(registros_por_pagina)
    except ValueError:
        limite = 10  # Fallback seguro

    paginator = Paginator(docentes, limite)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
 
    return render(request, 'users/personas/gestionar_docentes.html', {
        'docentes': page_obj,  # El objeto de página es iterable directo en el template
        'estado_actual': estado_actual,
        'colegio': colegio,
        'query': query,
        'orden_actual': orden,
        'registros_actual': registros_por_pagina
    })


@login_required
def crear_docente(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)

    if request.user.rol.nombre != 'Admin' and request.user.rol.nombre != 'Super':
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied
    
    if request.method == 'POST':

        email = request.POST.get('email', '').strip().lower()
        cedula = request.POST.get('cedula', '').strip().upper()
        nombre = request.POST.get('nombre', '').strip().upper()
        apellido = request.POST.get('apellido', '').strip().upper()
        telefono = request.POST.get('telefono', '').strip().upper()
        fecha_nacimiento = request.POST.get('fecha_nacimiento', '').strip()
        profesion = request.POST.get('profesion', '').strip().upper()
        direccion = request.POST.get('direccion', '').strip().upper()


        try:
            # Importante: Asegúrate de que el Rol 'Docente' existe en este colegio
            rol_docente = Rol.objects.filter(colegio=colegio, nombre__iexact='Docente').first()
            if not rol_docente:
                print("ERROR: Debes crear primero el Rol 'Docente' en este colegio.")
                # Podrías crear el rol aquí mismo si no existe

            # Usamos una transacción para asegurar que se creen ambos o ninguno
            from django.db import transaction
            with transaction.atomic():
                nuevo_usuario = Usuario.objects.create_user(
                    email=email,
                    password=cedula,
                    colegio=colegio,
                    rol=rol_docente
                )

                persona = Persona.objects.create(
                    usuario=nuevo_usuario,
                    colegio=colegio,
                    cedula=cedula,
                    nombre=nombre,
                    apellido=apellido,
                    telefono=telefono,
                    direccion=direccion,
                    fecha_nacimiento=fecha_nacimiento or None,
                    profesion=profesion,                    
                    es_docente=True, 
                    tipo='DOCENTE'
                )

                DocenteDetalle.objects.create(
                    docente=persona,
                    especialidad=request.POST.get('especialidad', '').strip().upper(),
                    fecha_inicio=request.POST.get('fecha_inicio') or None
                )
           
            print("¡ÉXITO! Usuario y Persona creados.")
            return redirect('gestionar_docentes', colegio_slug=colegio.slug)
            
        except Rol.DoesNotExist:
            messages.error(request, "El rol 'Docente' no existe. Créalo en la sección de Roles antes de continuar.")
        except Exception as e:
            messages.error(request, f"Error al crear el registro: {e}")
            
    return render(request, 'users/personas/crear_docente.html', {'colegio': colegio})

@login_required
def editar_docente(request, colegio_slug, persona_id):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    # Buscamos la persona asegurándonos que sea un DOCENTE de este colegio
    docente = get_object_or_404(Persona, id=persona_id, colegio=colegio, tipo='DOCENTE')

    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()
        cedula = request.POST.get('cedula', '').strip().upper()
        nombre = request.POST.get('nombre', '').strip().upper()
        apellido = request.POST.get('apellido', '').strip().upper()
        telefono = request.POST.get('telefono', '').strip().upper()
        direccion = request.POST.get('direccion', '').strip().upper()
        profesion = request.POST.get('profesion', '').strip().upper()
        fecha_nacimiento = request.POST.get('fecha_nacimiento', '').strip()

        try:
            with transaction.atomic():
                # 1. Actualizar datos del Usuario vinculado
                usuario = docente.usuario
                if usuario:
                    usuario.email = email
                    # Si quieres que la cédula nueva sea la clave si cambió, podrías usar set_password
                    # pero normalmente solo actualizamos el email aquí.
                    usuario.save()

                # 2. Actualizar datos de la Persona
                docente.cedula = cedula
                docente.nombre = nombre
                docente.apellido = apellido
                docente.telefono = telefono
                docente.direccion = direccion
                docente.fecha_nacimiento = fecha_nacimiento or None
                docente.profesion = profesion
                docente.save()

                # 3. NUEVO: Detalle Docente (Usamos update_or_create por si no existía)
                DocenteDetalle.objects.update_or_create(
                    docente=docente,
                    defaults={
                        'especialidad': request.POST.get('especialidad', '').strip().upper(),
                        'fecha_inicio': request.POST.get('fecha_inicio') or None,
                    }
                )

            messages.success(request, f"DOCENTE {nombre} ACTUALIZADO CORRECTAMENTE.")
            return redirect('gestionar_docentes', colegio_slug=colegio.slug)
        except Exception as e:
            messages.error(request, f"ERROR AL ACTUALIZAR: {str(e).upper()}")

    return render(request, 'users/personas/editar_docente.html', {
        'colegio': colegio,
        'docente': docente
    })

@login_required
def eliminar_docente(request, colegio_slug, persona_id):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    docente = get_object_or_404(Persona, id=persona_id, colegio=colegio, tipo='DOCENTE')

    try:
        with transaction.atomic():
            docente.activo = False # Marcamos como inactivo
            docente.save()
            
            # También desactivamos su usuario para que no pueda entrar al sistema
            if docente.usuario:
                docente.usuario.is_active = False
                docente.usuario.save()
                
        messages.success(request, f"EL DOCENTE {docente.nombre} HA SIDO DESACTIVADO.")
    except Exception as e:
        messages.error(request, f"ERROR AL DESACTIVAR: {str(e).upper()}")
    
    return redirect('gestionar_docentes', colegio_slug=colegio.slug)

@login_required
def reactivar_docente(request, colegio_slug, persona_id):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    docente = get_object_or_404(Persona, id=persona_id, colegio=colegio, tipo='DOCENTE')

    try:
        with transaction.atomic():
            docente.activo = True
            docente.save()
            if docente.usuario:
                docente.usuario.is_active = True
                docente.usuario.save()
        messages.success(request, f"EL DOCENTE {docente.nombre} HA SIDO REACTIVADO.")
    except Exception as e:
        messages.error(request, f"ERROR AL REACTIVAR: {str(e).upper()}")
    
    return redirect('gestionar_docentes', colegio_slug=colegio.slug)

def descargar_plantilla_docentes(request):
    import pandas as pd
    from django.http import HttpResponse
    from io import BytesIO

    # Definimos las columnas necesarias para docentes
    columnas = ['CEDULA', 'NOMBRE', 'APELLIDO', 'EMAIL', 'TELEFONO', 'PROFESION', 'ESPECIALIDAD', 'DIRECCION', 'FECHA_NACIMIENTO', 'FECHA_INGRESO']
    
    # Creamos un DataFrame vacío
    df = pd.DataFrame(columns=columnas)
    
    # Configuramos la respuesta HTTP para descargar el archivo
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Plantilla')
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_docentes.xlsx'
    return response

@login_required
def cargar_docentes_excel(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    # Seguridad básica
    if request.user.rol.nombre not in ['Admin', 'Super']:
        raise PermissionDenied

    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        archivo = request.FILES['archivo_excel']
        
        try:
            df = pd.read_excel(archivo)
            # Normalizar nombres de columnas
            df.columns = [c.strip().upper() for c in df.columns]
            
            creados = 0
            errores = []
            
            # Obtener el Rol de Docente para este colegio
            rol_docente = Rol.objects.filter(colegio=colegio, nombre__iexact='Docente').first()
            
            with transaction.atomic():
                for index, row in df.iterrows():
                    try:
                        email = str(row['EMAIL']).strip().lower()
                        cedula = str(row['CEDULA']).strip().upper()

                        if Usuario.objects.filter(email=email).exists():
                            errores.append(f"Fila {index+2}: El correo {email} ya existe.")
                            continue

                        # 1. Crear Usuario (clave por defecto es la cédula)
                        user = Usuario.objects.create_user(
                            email=email,
                            password=cedula,
                            colegio=colegio,
                            rol=rol_docente
                        )

                        # 2. Crear Persona
                        persona = Persona.objects.create(
                            usuario=user,
                            colegio=colegio,
                            cedula=cedula,
                            nombre=str(row['NOMBRE']).strip().upper(),
                            apellido=str(row['APELLIDO']).strip().upper(),
                            telefono=str(row.get('TELEFONO', '')).strip(),
                            direccion=str(row.get('DIRECCION', '')).strip().upper(),
                            fecha_nacimiento=pd.to_datetime(row['FECHA_NACIMIENTO']).date() if pd.notnull(row.get('FECHA_NACIMIENTO')) else None,
                            profesion=str(row.get('PROFESION', '')).strip().upper(),
                            es_docente=True,
                            tipo='DOCENTE'
                        )

                        # 3. Crear Detalle Docente
                        fecha_ing = None
                        if pd.notnull(row.get('FECHA_INGRESO')):
                            fecha_ing = pd.to_datetime(row['FECHA_INGRESO']).date()

                        DocenteDetalle.objects.create(
                            docente=persona,
                            especialidad=str(row.get('ESPECIALIDAD', '')).strip().upper(),
                            fecha_inicio=fecha_ing
                        )
                        
                        creados += 1

                    except Exception as e:
                        errores.append(f"Fila {index+2}: {str(e)}")

            if creados > 0:
                messages.success(request, f"¡Éxito! Se registraron {creados} nuevos docentes.")
            if errores:
                messages.error(request, f"Errores en {len(errores)} filas. Revise el formato del archivo.")

        except Exception as e:
            messages.error(request, f"Error crítico al leer el Excel: {e}")
            
    return redirect('gestionar_docentes', colegio_slug=colegio.slug)

@login_required
def imprimir_ficha_docente(request, colegio_slug, docente_id):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    docente = get_object_or_404(
            Persona.objects.select_related('detalle_docente', 'usuario'), 
            id=docente_id, 
            colegio=colegio
        )

    detalle = getattr(docente, 'detalle_docente', None)

    context = {
        'docente': docente,
        'detalle': detalle,
        'colegio': colegio,
    }

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="ficha_{docente.cedula}.pdf"'

    template = get_template('users/personas/ficha_docente_pdf.html')
    html = template.render(context)
    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('Ocurrió un error al generar el PDF', status=500)
      
    return response

# --- EXPORTAR DOCENTES A EXCEL ---
@login_required
def exportar_docentes_xls(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    estado = request.GET.get('estado') or 'activo'
    activo_bool = True if estado == 'activo' else False
    query = request.GET.get('q', '').strip()
    orden = request.GET.get('orden') or 'apellido'

    # Queryset base
    docentes = Persona.objects.filter(colegio=colegio, tipo='DOCENTE', activo=activo_bool).select_related('detalle_docente', 'usuario')

    # Aplicar buscador si existe
    if query:
        docentes = docentes.filter(
            Q(nombre__icontains=query) |
            Q(apellido__icontains=query) |
            Q(cedula__icontains=query) |
            Q(detalle_docente__especialidad__icontains=query)
        )

    # Aplicar ordenamiento dinámico
    if orden == 'nombre':
        docentes = docentes.order_by('nombre')
    elif orden == 'cedula':
        docentes = docentes.order_by('cedula')
    elif orden == 'especialidad':
        docentes = docentes.order_by('detalle_docente__especialidad')
    else:
        docentes = docentes.order_by('apellido')

    # Construir la estructura de datos para Pandas
    data = []
    for docente in docentes: 
        detalle = getattr(docente, 'detalle_docente', None)
        
        especialidad = detalle.especialidad if detalle and detalle.especialidad else 'SIN ASIGNAR'
        fecha_inicio = detalle.fecha_inicio.strftime('%d/%m/%Y') if detalle and detalle.fecha_inicio else 'N/A'
        
        data.append({
            'Cédula': docente.cedula,
            'Nombre': docente.nombre,
            'Apellido': docente.apellido,
            'Email': docente.usuario.email if docente.usuario else 'N/A',
            'Especialidad': especialidad,
            'Profesión': docente.profesion or 'N/A',
            'Fecha de Nacimiento': docente.fecha_nacimiento.strftime('%d/%m/%Y') if docente.fecha_nacimiento else 'N/A',
            'Fecha de Inicio': fecha_inicio,
            'Teléfono': docente.telefono or 'N/A',
            'Dirección': docente.direccion or 'N/A'
        })

    # Si el dataframe queda vacío, al menos llevará las columnas estructuradas
    df = pd.DataFrame(data)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=docentes_{colegio.slug}_{estado}.xlsx'
    
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Docentes')
    
    return response


# --- EXPORTAR DOCENTES A PDF ---
@login_required
def exportar_docentes_pdf(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    # CORRECCIÓN: Protegemos también la vista del PDF contra parámetros vacíos
    estado = request.GET.get('estado') or 'activo'
    activo_bool = True if estado == 'activo' else False
    query = request.GET.get('q', '').strip()
    orden = request.GET.get('orden') or 'apellido'

    # Queryset base
    docentes = Persona.objects.filter(colegio=colegio, tipo='DOCENTE', activo=activo_bool).select_related('detalle_docente', 'usuario')

    # Aplicar buscador si existe
    if query:
        docentes = docentes.filter(
            Q(nombre__icontains=query) |
            Q(apellido__icontains=query) |
            Q(cedula__icontains=query) |
            Q(detalle_docente__especialidad__icontains=query)
        )

    # Aplicar ordenamiento dinámico
    if orden == 'nombre':
        docentes = docentes.order_by('nombre')
    elif orden == 'cedula':
        docentes = docentes.order_by('cedula')
    elif orden == 'especialidad':
        docentes = docentes.order_by('detalle_docente__especialidad')
    else:
        docentes = docentes.order_by('apellido')

    context = {
        'colegio': colegio,
        'docentes': docentes,
        'estado_actual': estado.upper(),
        'fecha': datetime.now(),
    }
    
    html_string = render_to_string('users/personas/pdf_docentes.html', context)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="lista_docentes_{colegio.slug}_{estado}.pdf"'
    
    pisa_status = pisa.CreatePDF(BytesIO(html_string.encode("UTF-8")), dest=response)
    
    if pisa_status.err:
        return HttpResponse('Ocurrió un error al generar el documento PDF', status=500)
    return response

@login_required
def gestionar_administrativos(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)

    nombre_rol = getattr(request.user.rol, 'nombre', '')
    if nombre_rol not in ['Admin', 'Super', 'ADMINISTRADOR']:
        raise PermissionDenied   
    
    # Captura de parámetros de filtrado, orden y paginación
    estado_actual = request.GET.get('estado', 'activo')
    query = request.GET.get('q', '')
    orden = request.GET.get('orden', 'apellido')
    registros_por_pagina = request.GET.get('registros', '10')
    
    base_queryset = Persona.objects.filter(
        colegio=colegio, 
        tipo='ADMIN'
    ).select_related('usuario', 'detalle_laboral')

    # 1. Filtrar según la pestaña (Activo / Inactivo)
    if estado_actual == 'inactivo':
        administrativos = base_queryset.filter(activo=False)
    else:
        administrativos = base_queryset.filter(activo=True)
   
    # 2. Filtrar por barra de búsqueda
    if query:
        administrativos = administrativos.filter(
            Q(nombre__icontains=query) |
            Q(apellido__icontains=query) |
            Q(cedula__icontains=query) |
            Q(detalle_laboral__cargo__icontains=query)
        )

    # 3. Aplicar Criterio de Ordenamiento
    if orden == 'nombre':
        administrativos = administrativos.order_by('nombre', 'apellido')
    elif orden == 'cedula':
        administrativos = administrativos.order_by('cedula')
    elif orden == 'cargo':
        administrativos = administrativos.order_by('detalle_laboral__cargo')
    else:  # Por defecto: 'apellido'
        administrativos = administrativos.order_by('apellido', 'nombre')

    # 4. Procesar la Paginación
    try:
        limite = int(registros_por_pagina)
    except ValueError:
        limite = 10  # Fallback seguro

    paginator = Paginator(administrativos, limite)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
 
    return render(request, 'users/personas/gestionar_administrativos.html', {
        'administrativos': page_obj,  # El objeto de página es iterable directo en el template
        'estado_actual': estado_actual,
        'colegio': colegio,
        'query': query,
        'orden_actual': orden,
        'registros_actual': registros_por_pagina
    })

@login_required
def crear_administrativo(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)

    # Obtenemos los roles para el select (exceptuando quizás el de SuperUser)
    roles = Rol.objects.all()

    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()
        cedula = request.POST.get('cedula', '').strip().upper()
        nombre = request.POST.get('nombre', '').strip().upper()
        apellido = request.POST.get('apellido', '').strip().upper()
        telefono = request.POST.get('telefono', '').strip().upper()
        cargo = request.POST.get('cargo', '').strip().upper()
        profesion = request.POST.get('profesion', '').strip().upper()
        direccion = request.POST.get('direccion', '').strip().upper()
        fecha_inicio = request.POST.get('fecha_inicio', '').strip()
        fecha_nacimiento = request.POST.get('fecha_nacimiento', '').strip()
        rol_id = request.POST.get('rol') # Capturamos el ID del rol seleccionado

        if Persona.objects.filter(cedula=cedula, colegio=colegio).exists():
            messages.error(request, "Ya existe una persona con esta cédula en este colegio.")
            return render(request, 'users/personas/crear_administrativo.html', {'colegio': colegio})

        try:
            # Buscamos el rol para el personal de oficina
            rol_admin = Rol.objects.filter(colegio=colegio, nombre__iexact='Administrativo').first()
            
            if not rol_admin:
                messages.error(request, "DEBE CREAR EL ROL 'ADMINISTRATIVO' PRIMERO.")
                return redirect('gestionar_administrativos', colegio_slug=colegio.slug)

            with transaction.atomic():
                # 1. Crear el usuario
                nuevo_usuario = Usuario.objects.create_user(
                    email=email, 
                    password=cedula, 
                    colegio=colegio, 
                    rol_id=rol_id  # Asignamos el rol seleccionado
                )
                
                # 2. Crear la persona y GUARDARLA en una variable
                nuevo_administrativo = Persona.objects.create(
                    usuario=nuevo_usuario, 
                    colegio=colegio, 
                    cedula=cedula,
                    nombre=nombre, 
                    apellido=apellido, 
                    telefono=telefono,    
                    direccion=direccion,                 
                    profesion=profesion,
                    fecha_nacimiento=fecha_nacimiento if fecha_nacimiento else None,
                    es_admin=True, 
                    tipo='ADMIN',
                    activo=True
                )
                
                # 3. Crear el detalle usando la variable anterior
                # Nota: Asegúrate de que el campo en el modelo se llame 'administrativo' o 'persona'
                AdministrativoDetalle.objects.create(
                    administrativo=nuevo_administrativo, # Cambié 'persona' por 'administrativo' según tu tabla
                    cargo=cargo,
                    fecha_inicio=fecha_inicio if fecha_inicio else None
                )
            messages.success(request, f"ADMINISTRATIVO {nombre} CREADO CON ÉXITO.")
            return redirect('gestionar_administrativos', colegio_slug=colegio.slug)
        except Exception as e:
            messages.error(request, f"ERROR: {str(e).upper()}")
            
    return render(request, 'users/personas/crear_administrativo.html', {
        'colegio': colegio,
        'roles': roles # Pasamos los roles al template para que el usuario pueda seleccionar el rol al crear un administrativo
    })

@login_required
def editar_administrativo(request, colegio_slug, persona_id):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    # Obtenemos la persona asegurándonos que pertenece al colegio
    administrativo = get_object_or_404(Persona, id=persona_id, colegio=colegio, tipo='ADMIN')

    if request.method == 'POST':
        email = request.POST.get('email')
        nombre = request.POST.get('nombre').upper()
        apellido = request.POST.get('apellido').upper()
        telefono = request.POST.get('telefono')
        cargo = request.POST.get('cargo').upper()
        profesion = request.POST.get('profesion').upper()
        direccion = request.POST.get('direccion').upper()
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_nacimiento = request.POST.get('fecha_nacimiento')

        # Capturar el estado del toggle 'activo'
        # Si el checkbox está marcado llega como 'on', si no, llega None
        esta_activo = request.POST.get('activo') == 'on'

        try:
            with transaction.atomic():
                # 1. Actualizar datos del Usuario (Auth)
                usuario = administrativo.usuario
                if usuario:
                    usuario.email = email.lower()
                    usuario.is_active = esta_activo # Sincroniza el acceso con el estado del registro
                    usuario.save()

                # 2. Actualizar datos de la Persona
                administrativo.nombre = nombre
                administrativo.apellido = apellido
                administrativo.telefono = telefono
                administrativo.fecha_nacimiento = fecha_nacimiento if fecha_nacimiento else None
                administrativo.direccion = direccion
                administrativo.profesion = profesion
                administrativo.activo = esta_activo
                administrativo.save()

                # 3. Actualizar o Crear el Detalle Laboral
                # Usamos el related_name 'detalle_laboral' que definimos en el modelo
                AdministrativoDetalle.objects.update_or_create(
                    administrativo=administrativo,
                    defaults={
                        'cargo': cargo,
                        'fecha_inicio': fecha_inicio if fecha_inicio else None
                    }
                )

                messages.success(request, f"El perfil de {nombre} ha sido actualizado correctamente.")
                return redirect('gestionar_administrativos', colegio_slug=colegio.slug)

        except Exception as e:
            messages.error(request, f"Error al actualizar: {str(e)}")
    
    return render(request, 'users/personas/editar_administrativo.html', {
        'administrativo': administrativo,
        'colegio': colegio,
        'detalle': getattr(administrativo, 'detalle_laboral', None)
    })

@login_required
def eliminar_administrativo(request, colegio_slug, persona_id):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    administrativo = get_object_or_404(Persona, id=persona_id, colegio=colegio, tipo='ADMIN')

    try:
        with transaction.atomic():
            administrativo.activo = False # Marcamos como inactivo
            administrativo.save()
            
            # También desactivamos su usuario para que no pueda entrar al sistema
            if administrativo.usuario:
                administrativo.usuario.is_active = False
                administrativo.usuario.save()
                
        messages.success(request, f"EL ADMINISTRATIVO {administrativo.nombre} HA SIDO DESACTIVADO.")
    except Exception as e:
        messages.error(request, f"ERROR AL DESACTIVAR: {str(e).upper()}")
    
    return redirect('gestionar_administrativos', colegio_slug=colegio.slug)

@login_required
def reactivar_administrativo(request, colegio_slug, persona_id):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    administrativo = get_object_or_404(Persona, id=persona_id, colegio=colegio, tipo='ADMIN')

    try:
        with transaction.atomic():
            administrativo.activo = True
            administrativo.save()
            if administrativo.usuario:
                administrativo.usuario.is_active = True
                administrativo.usuario.save()
        messages.success(request, f"EL ADMINISTRATIVO {administrativo.nombre} HA SIDO REACTIVADO.")
    except Exception as e:
        messages.error(request, f"ERROR AL REACTIVAR: {str(e).upper()}")
    
    return redirect(f"{reverse('gestionar_administrativos', args=[colegio.slug])}?estado=inactivo")

@login_required
def imprimir_ficha_administrativo(request, colegio_slug, administrativo_id):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    administrativo = get_object_or_404(Persona, id=administrativo_id, colegio=colegio)
    
    context = {
        'administrativo': administrativo,
        'detalle': administrativo.detalle_laboral,
        'colegio': colegio,
    }

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="ficha_{administrativo.cedula}.pdf"'
    
    template = get_template('users/personas/ficha_administrativo_pdf.html')
    html = template.render(context)
    pisa_status = pisa.CreatePDF(html, dest=response, link_callback=link_callback)

    if pisa_status.err:
        return HttpResponse('Ocurrió un error al generar el PDF', status=500)
      
    return response

def descargar_plantilla_admin(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)

    import openpyxl
    from openpyxl.styles import Font, PatternFill
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Administrativos"

    # Encabezados exactos para el proceso
    headers = ['CEDULA', 'NOMBRE', 'APELLIDO', 'EMAIL', 'TELEFONO', 'DIRECCION', 'CARGO', 'PROFESION', 'FECHA_NACIMIENTO', 'FECHA_INGRESO']
    ws.append(headers)

    # Estilo para el encabezado
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=plantilla_administrativos.xlsx'
    wb.save(response)
    return response

@login_required
def cargar_administrativos_excel(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    if request.method == 'POST':
        archivo = request.FILES.get('archivo_excel')
        if not archivo:
            messages.error(request, "Debe seleccionar un archivo Excel.")
            return redirect('cargar_administrativos_excel', colegio_slug=colegio.slug)

        try:
            df = pd.read_excel(archivo)
            df.columns = [c.strip().upper() for c in df.columns] # Estandarizar columnas
            
            # Buscamos el rol para asignarlo a los nuevos usuarios
            rol_secretaria = Rol.objects.filter(nombre__icontains='Secretaria').first()

            # Validación de seguridad: si no existe 'Secretaria', busca 'Administrativo' o el primero disponible
            if not rol_secretaria:
                rol_secretaria = Rol.objects.first()
            
            administrativos_creados = 0
            errores = []

            with transaction.atomic():
                # Reemplaza la sección del for dentro de tu vista por esto:
                for index, row in df.iterrows():
                    try:
                        # Función auxiliar rápida para limpiar textos de celdas vacías de Pandas
                        def limpiar_texto(valor, mayuscula=True):
                            if pd.isnull(valor):
                                return ""
                            texto = str(valor).strip()
                            return texto.upper() if mayuscula else texto.lower()

                        email = limpiar_texto(row.get('EMAIL'), mayuscula=False)
                        cedula = str(row['CEDULA']).split('.')[0].strip() # Evita que cédulas numéricas se lean como "12345.0"
                        
                        if not email or not cedula:
                            id_fila = row.get('NOMBRE') or f"Línea {index+2}"
                            raise ValueError(f"El administrativo {id_fila} no tiene CÉDULA o EMAIL válido.")

                        # 1. Crear Usuario
                        usuario, u_created = Usuario.objects.get_or_create(
                            email=email,
                            defaults={
                                'password': make_password(cedula),
                                'colegio': colegio,
                                'rol': rol_secretaria,
                                'is_active': True
                            }
                        )

                        # 2. Crear Persona utilizando la limpieza segura
                        persona, p_created = Persona.objects.update_or_create(
                            cedula=cedula,
                            colegio=colegio,
                            defaults={
                                'usuario': usuario,
                                'nombre': limpiar_texto(row.get('NOMBRE')),
                                'apellido': limpiar_texto(row.get('APELLIDO')),
                                'telefono': limpiar_texto(row.get('TELEFONO'), mayuscula=False), # Teléfonos sin mayúsculas
                                'direccion': limpiar_texto(row.get('DIRECCION')),
                                'profesion': limpiar_texto(row.get('PROFESION')),
                                'fecha_nacimiento': pd.to_datetime(row['FECHA_NACIMIENTO']).date() if pd.notnull(row.get('FECHA_NACIMIENTO')) else None,
                                'tipo': 'ADMIN',
                                'es_admin': True,   
                                'activo': True
                            }
                        )

                        # 3. Detalle Laboral
                        fecha_ingreso = None
                        if pd.notnull(row.get('FECHA_INGRESO')):
                            fecha_ingreso = pd.to_datetime(row['FECHA_INGRESO']).date()

                        AdministrativoDetalle.objects.update_or_create(
                            administrativo=persona,
                            defaults={
                                'cargo': limpiar_texto(row.get('CARGO')),
                                'fecha_inicio': fecha_ingreso
                            }
                        )
                        administrativos_creados += 1

                    except Exception as e:
                        errores.append(f"Fila {index+2}: {str(e)}")

            if administrativos_creados > 0:
                messages.success(request, f"Se cargaron {administrativos_creados} administrativos correctamente.")
            
            if errores:
                msg_error = "Errores encontrados: <br>" + "<br>".join(errores[:5])
                messages.error(request, msg_error, extra_tags='safe')

            return redirect('gestionar_administrativos', colegio_slug=colegio.slug)

        except Exception as e:
            messages.error(request, f"Error al procesar el Excel: {str(e)}")
            
    return render(request, 'users/personas/cargar_masiva_admin.html', {'colegio': colegio})

# --- EXPORTAR ADMINISTRATIVOS A EXCEL ---
@login_required
def exportar_administrativos_xls(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    # Capturar los mismos filtros del frontend
    estado = request.GET.get('estado', 'activo')
    activo_bool = True if estado == 'activo' else False
    query = request.GET.get('q', '')
    orden = request.GET.get('orden', 'apellido')

    # Queryset base
    administrativos = Persona.objects.filter(colegio=colegio, tipo='ADMIN', activo=activo_bool)

    # Aplicar buscador si existe
    if query:
        administrativos = administrativos.filter(
            Q(nombre__icontains=query) |
            Q(apellido__icontains=query) |
            Q(cedula__icontains=query) |
            Q(detalle_laboral__cargo__icontains=query)
        )

    # Aplicar ordenamiento dinámico
    if orden == 'nombre':
        administrativos = administrativos.order_by('nombre')
    elif orden == 'cedula':
        administrativos = administrativos.order_by('cedula')
    elif orden == 'cargo':
        administrativos = administrativos.order_by('detalle_laboral__cargo')
    else:
        administrativos = administrativos.order_by('apellido')

    # Construir la estructura de datos para Pandas
    data = []
    for admin in administrativos:
        # Extraer datos de la relación OneToOne de manera segura
        cargo = admin.detalle_laboral.cargo if hasattr(admin, 'detalle_laboral') else 'SIN ASIGNAR'
        fecha_inicio = admin.detalle_laboral.fecha_inicio.strftime('%d/%m/%Y') if hasattr(admin, 'detalle_laboral') and admin.detalle_laboral.fecha_inicio else 'N/A'
        
        data.append({
            'Cédula': admin.cedula,
            'Nombre': admin.nombre,
            'Apellido': admin.apellido,
            'Email': admin.usuario.email if admin.usuario else 'N/A',
            'Cargo': cargo,
            'Profesión': admin.profesion or 'N/A',
            'fecha_nacimiento': admin.fecha_nacimiento.strftime('%d/%m/%Y') if admin.fecha_nacimiento else 'N/A',
            'Fecha de Inicio': fecha_inicio,
            'Teléfono': admin.telefono or 'N/A',
            'Dirección': admin.direccion or 'N/A'
        })

    df = pd.DataFrame(data)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=administrativos_{colegio.slug}_{estado}.xlsx'
    
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Administrativos')
    
    return response


# --- EXPORTAR ADMINISTRATIVOS A PDF ---
@login_required
def exportar_administrativos_pdf(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    # Capturar los mismos filtros del frontend
    estado = request.GET.get('estado', 'activo')
    activo_bool = True if estado == 'activo' else False
    query = request.GET.get('q', '')
    orden = request.GET.get('orden', 'apellido')

    # Queryset base
    administrativos = Persona.objects.filter(colegio=colegio, tipo='ADMIN', activo=activo_bool)

    # Aplicar buscador si existe
    if query:
        administrativos = administrativos.filter(
            Q(nombre__icontains=query) |
            Q(apellido__icontains=query) |
            Q(cedula__icontains=query) |
            Q(detalle_laboral__cargo__icontains=query)
        )

    # Aplicar ordenamiento dinámico
    if orden == 'nombre':
        administrativos = administrativos.order_by('nombre')
    elif orden == 'cedula':
        administrativos = administrativos.order_by('cedula')
    elif orden == 'cargo':
        administrativos = administrativos.order_by('detalle_laboral__cargo')
    else:
        administrativos = administrativos.order_by('apellido')

    context = {
        'colegio': colegio,
        'administrativos': administrativos,
        'estado_actual': estado.upper(),
        'fecha': datetime.now(),
    }
    
    # Renderizamos el template HTML estructurado para xhtml2pdf
    html_string = render_to_string('users/personas/pdf_administrativos.html', context)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="lista_administrativos_{colegio.slug}_{estado}.pdf"'
    
    # Crear el PDF usando BytesIO
    pisa_status = pisa.CreatePDF(BytesIO(html_string.encode("UTF-8")), dest=response)
    
    if pisa_status.err:
        return HttpResponse('Ocurrió un error al generar el documento PDF', status=500)
    return response

@login_required
def crear_rol(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    # Verificación de seguridad
    if getattr(request.user.rol, 'nombre', '') not in ['Admin', 'Super']:
        raise PermissionDenied

    if request.method == 'POST':
        nombre = request.POST.get('nombre').strip()
        
        if nombre:
            # Creamos el Rol usando los nombres exactos de tu modelo
            Rol.objects.create(
                colegio=colegio,
                nombre=nombre,
                # Módulo Contenido
                can_manage_news='can_manage_news' in request.POST,
                can_manage_gallery='can_manage_gallery' in request.POST,
                # Módulo Personas
                can_manage_staff='can_manage_staff' in request.POST,
                can_manage_students='can_manage_students' in request.POST,
                # Módulo Académico
                can_manage_grades='can_manage_grades' in request.POST,
                # Módulo Finanzas
                can_manage_finances='can_manage_finances' in request.POST,
                # Módulo Cantina
                can_manage_canteen='can_manage_canteen' in request.POST,
            )
            messages.success(request, f"EL ROL '{nombre.upper()}' HA SIDO CONFIGURADO CORRECTAMENTE.")
            return redirect('gestionar_roles', colegio_slug=colegio.slug)

    return render(request, 'users/seguridad/crear_rol.html', {'colegio': colegio})

@login_required
def alternar_estado_usuario(request, usuario_id):
    if getattr(request.user.rol, 'nombre', '') not in ['Admin', 'Super']:
        raise PermissionDenied
    
    # Buscamos al usuario que queremos bloquear/desbloquear
    usuario = get_object_or_404(Usuario, id=usuario_id)
    
    # No permitir que el Super se bloquee a sí mismo
    if usuario == request.user:
        messages.error(request, "No puedes suspender tu propia cuenta.")
    else:
        usuario.is_active = not usuario.is_active
        usuario.save()
        estado = "ACTIVADO" if usuario.is_active else "SUSPENDIDO"
        messages.success(request, f"El usuario {usuario.email} ahora está {estado}.")
    
    # Volvemos a la lista usando el slug del colegio del usuario actual
    return redirect('gestionar_usuarios', colegio_slug=request.user.colegio.slug)

@login_required
def resetear_password_usuario(request, usuario_id):
    # 1. Seguridad básica: solo Admin o Super
    if getattr(request.user.rol, 'nombre', '') not in ['Admin', 'Super']:
        raise PermissionDenied
    
    # 2. Definimos las variables base (esto evita el error de TemplateDoesNotExist/NoReverseMatch)
    usuario = get_object_or_404(Usuario, id=usuario_id)
    colegio = request.user.colegio  # <--- Definimos 'colegio' aquí para que esté disponible siempre
    
    # 3. Validar que el Admin no intente cambiar claves de otro colegio
    if usuario.colegio != colegio:
        raise PermissionDenied

    if request.method == 'POST':
        nueva_clave = request.POST.get('nueva_password')
        if nueva_clave:
            # Django se encarga de la encriptación
            usuario.set_password(nueva_clave)
            usuario.save()
            messages.success(request, f"CONTRASEÑA ACTUALIZADA PARA: {usuario.email}")
            return redirect('gestionar_usuarios', colegio_slug=colegio.slug)
    
    # 4. Enviamos ambas variables al template
    return render(request, 'users/seguridad/resetear_password.html', {
        'usuario_target': usuario, 
        'colegio': colegio
    })

def gestionar_estudiantes(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    # 1. Capturar los parámetros de filtrado, orden y paginación desde el GET
    pestana = request.GET.get('tab', 'activos')
    query = request.GET.get('q', '').strip()
    grado_actual = request.GET.get('grado', '')
    seccion_actual = request.GET.get('seccion', '')
    nivel_actual = request.GET.get('nivel', '')          
    anio_actual = request.GET.get('anio_escolar', '')    
    
    order_by = request.GET.get('order', 'apellido')
    per_page = request.GET.get('per_page', '10')
    page = request.GET.get('page', 1)
    
    anio_activo = AnioEscolar.objects.filter(colegio=colegio, activo=True).order_by('-id').first()
    
    # Si es la primera carga de la página, usamos el año activo por defecto
    if not anio_actual and anio_activo:
        anio_actual = str(anio_activo.id)
    
    # 2. Queryset base segmentando la lógica según la pestaña seleccionada
    if pestana == 'inactivos':
        estudiantes_list = Persona.objects.filter(colegio=colegio, tipo='ESTUDIANTE', activo=False)
        if anio_actual:
            estudiantes_list = estudiantes_list.filter(inscripciones__anio_escolar_id=anio_actual)
            
    elif pestana == 'sin_inscripcion':
        # Flotantes: Están activos en el sistema pero NO tienen inscripción en el año escolar seleccionado
        estudiantes_list = Persona.objects.filter(colegio=colegio, tipo='ESTUDIANTE', activo=True)
        if anio_actual:
            estudiantes_list = estudiantes_list.exclude(inscripciones__anio_escolar_id=anio_actual)
            
    else:  # 'activos'
        # Inscritos: Están activos y SÍ tienen inscripción en el año escolar seleccionado
        estudiantes_list = Persona.objects.filter(colegio=colegio, tipo='ESTUDIANTE', activo=True)
        if anio_actual:
            estudiantes_list = estudiantes_list.filter(inscripciones__anio_escolar_id=anio_actual)
        
    # 4. Filtro de Búsqueda de texto (Cédula, Nombre o Apellido)
    if query:
        estudiantes_list = estudiantes_list.filter(
            Q(cedula__icontains=query) |
            Q(nombre__icontains=query) |
            Q(apellido__icontains=query)
        )
        
    # 5. Filtros Académicos (Nivel, Grado, Sección)
    # Nota: En 'sin_inscripcion' estos filtros buscarán en el historial de inscripciones previas del alumno
    if nivel_actual:
        estudiantes_list = estudiantes_list.filter(inscripciones__seccion__nivel=nivel_actual)
        
    if grado_actual:
        estudiantes_list = estudiantes_list.filter(inscripciones__seccion__grado=grado_actual)
        
    if seccion_actual:
        estudiantes_list = estudiantes_list.filter(inscripciones__seccion_id=seccion_actual)
        
    # Evitamos duplicados causados por los JOINs de las inscripciones
    estudiantes_list = estudiantes_list.distinct()
    
    # 7. Ordenamiento dinámico
    if order_by == 'nombre':
        estudiantes_list = estudiantes_list.order_by('nombre', 'apellido')
    elif order_by == 'cedula':
        estudiantes_list = estudiantes_list.order_by('cedula')
    else:
        estudiantes_list = estudiantes_list.order_by('apellido', 'nombre')
        
    # 8. Obtener datos dinámicos para popular los selectores del Formulario
    secciones = Seccion.objects.filter(colegio=colegio).order_by('nombre')
    if nivel_actual:
        secciones = secciones.filter(nivel=nivel_actual)
        
    grados = secciones.values_list('grado', flat=True).distinct().order_by('grado')
    anios_escolares = AnioEscolar.objects.filter(colegio=colegio).order_by('-id')
    niveles = Seccion.objects.filter(colegio=colegio).values_list('nivel', flat=True).distinct().order_by('nivel')
    
    # 9. Paginación Dinámica
    try:
        limit = int(per_page)
    except ValueError:
        limit = 10
        
    paginator = Paginator(estudiantes_list, limit)
    try:
        estudiantes = paginator.page(page)
    except PageNotAnInteger:
        estudiantes = paginator.page(1)
    except EmptyPage:
        estudiantes = paginator.page(paginator.num_pages)
        
    # 10. Construir el contexto para la plantilla
    context = {
        'colegio': colegio,
        'pestana': pestana,
        'query': query,
        'grado_actual': grado_actual,
        'seccion_actual': seccion_actual,
        'nivel_actual': nivel_actual,            
        'anio_actual': anio_actual,              
        'order_by': order_by,
        'per_page': str(limit),
        'grados': grados,       
        'secciones': secciones,
        'anios_escolares': anios_escolares,      
        'niveles': niveles,                                             
        'estudiantes': estudiantes,
    }
    
    return render(request, 'users/personas/gestionar_estudiantes.html', context)

@login_required
def crear_estudiante_completo(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    if request.method == 'POST':
        try:
            # Ponemos el bloque atómico únicamente en la zona donde se guardan los datos
            with transaction.atomic():
                # 1. PROCESAR REPRESENTANTE
                cedula_rep = request.POST.get('cedula_rep').strip().upper()
                telefono_rep = request.POST.get('telefono_rep') 
                direccion_comun = request.POST.get('direccion_rep').strip().upper()
                fecha_nac_rep = request.POST.get('fecha_nacimiento_rep')

                representante, created = Persona.objects.update_or_create(
                    cedula=cedula_rep,
                    colegio=colegio,
                    defaults={
                        'nombre': request.POST.get('nombre_rep').upper(),
                        'apellido': request.POST.get('apellido_rep').upper(),
                        'telefono': telefono_rep, 
                        'es_representante': True,                    
                        'direccion': direccion_comun,
                        'fecha_nacimiento': fecha_nac_rep,
                        'tipo': 'REPRESENTANTE'
                    }
                )
                
                # Crear usuario para representante si no tiene
                if not representante.usuario:
                    rol_rep = Rol.objects.filter(nombre__iexact='Representante', colegio=colegio).first()
                    email_rep = request.POST.get('email_rep').strip().lower()
                    user_rep = Usuario.objects.filter(email=email_rep).first()

                    if not user_rep:
                        user_rep = Usuario.objects.create_user(
                            email=email_rep,
                            password=cedula_rep, 
                            colegio=colegio,
                            rol=rol_rep
                        )
                    else:
                        user_rep.colegio = colegio
                        user_rep.rol = rol_rep
                        user_rep.save()

                    representante.usuario = user_rep
                    representante.save()

                # 2. PROCESAR ESTUDIANTE
                cedula_est = request.POST.get('cedula_est').strip().upper()
                fecha_nac_est = request.POST.get('fecha_nacimiento_est')
                
                # Validación manual extra para evitar el quiebre de la base de datos
                if Persona.objects.filter(cedula=cedula_est, es_estudiante=True, colegio=colegio).exists():
                    raise ValueError(f"La cédula {cedula_est} ya pertenece a un estudiante registrado en este colegio.")

                estudiante = Persona.objects.create(
                    colegio=colegio,
                    cedula=cedula_est,
                    nombre=request.POST.get('nombre_est').upper(),
                    apellido=request.POST.get('apellido_est').upper(),
                    telefono=telefono_rep,                
                    direccion=direccion_comun,   
                    fecha_nacimiento=fecha_nac_est,
                    es_estudiante=True,
                    tipo='ESTUDIANTE'
                )

                # 3. USUARIO PARA EL ESTUDIANTE
                rol_est = Rol.objects.filter(colegio=colegio, nombre__iexact='Estudiante').first()
                user_est = Usuario.objects.create_user(
                    email=f"{cedula_est}@colegio.com",
                    password=cedula_est,
                    colegio=colegio,
                    rol=rol_est
                )
                estudiante.usuario = user_est
                estudiante.save()

                # 4. VINCULACIÓN FAMILIAR
                RelacionFamiliar.objects.create(
                    representante=representante, 
                    estudiante=estudiante, 
                    parentesco=request.POST.get('parentesco')
                )

            # Si el bloque 'with' termina sin errores, guardamos el mensaje y redirigimos
            messages.success(request, f"Estudiante {estudiante.nombre} {estudiante.apellido} registrado exitosamente.")
            return redirect('gestionar_estudiantes', colegio_slug=colegio.slug)
            
        except Exception as e:
            # Ahora que salimos del bloque with, la transacción ya hizo Rollback limpio
            # y Django nos permitirá volver a pintar el render e interrogar a la BD sin problemas.
            print(f"Error real oculto en base de datos: {e}")
            messages.error(request, f"Error al registrar: {str(e)}")

    return render(request, 'users/personas/crear_estudiante.html', {
        'colegio': colegio
    })

def buscar_representante_ajx(request, colegio_slug): # Vista AJAX para buscar una persona por cédula y verificar si ya es representante o tiene vinculaciones
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    cedula = request.GET.get('cedula')
    data = {'existe': False, 'mensaje': 'NUEVO REPRESENTANTE'}
    persona = Persona.objects.filter(cedula=cedula, colegio=colegio).first()
    
    if persona:
        # Verificamos si ya tiene vinculaciones
        relaciones = RelacionFamiliar.objects.filter(representante=persona)
        lista_hijos = [
            {"cedula": r.estudiante.cedula, "nombre": f"{r.estudiante.nombre} {r.estudiante.apellido}".upper()} 
            for r in relaciones
        ]

        ultima_relacion = relaciones.last()
        parentesco_previo = ultima_relacion.parentesco if ultima_relacion else ""

        data = {
            'existe': True,
            'mensaje': 'REPRESENTANTE YA EXISTE' if persona.es_representante else 'PERSONA ENCONTRADA (NUEVO ROL: REPRESENTANTE)',
            'nombre': persona.nombre.upper(),
            'apellido': persona.apellido.upper(),
            'email': persona.usuario.email if persona.usuario else "",
            'telefono': persona.telefono if persona.telefono else "", 
            'direccion': persona.direccion if persona.direccion else "",
            'fecha_nacimiento': persona.fecha_nacimiento.strftime('%Y-%m-%d') if persona.fecha_nacimiento else '',
            'representados': lista_hijos,
            'parentesco': parentesco_previo  
        }
            
    return JsonResponse(data)

def buscar_estudiante_ajx(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    cedula = request.GET.get('cedula')
    estudiante = Persona.objects.filter(cedula=cedula, es_estudiante=True, colegio=colegio).first()

    if estudiante:
        return JsonResponse({
            'existe': True,
            'mensaje': f"EL ESTUDIANTE YA EXISTE: {estudiante.nombre} {estudiante.apellido}."
        })
    return JsonResponse({'existe': False})


@login_required
@require_POST
def subir_foto_persona(request, colegio_slug, persona_id):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    # Buscamos a la persona (sirve para Administrativos, Estudiantes, Docentes, etc.)
    persona = get_object_or_404(Persona, id=persona_id, colegio=colegio)

    if request.FILES.get('foto'):
        persona.foto = request.FILES['foto']
        persona.save()
        messages.success(request, f"Fotografía de {persona.nombre} actualizada.")
    else:
        messages.error(request, "No se recibió ninguna imagen.")

    # REDIRECCIÓN INTELIGENTE:
    if persona.es_admin:
        return redirect('gestionar_administrativos', colegio_slug=colegio_slug)
    
    if persona.es_docente:
        return redirect('gestionar_docentes', colegio_slug=colegio_slug)
    
    return redirect('gestionar_estudiantes', colegio_slug=colegio_slug)

@login_required
def ficha_estudiante_modal(request, colegio_slug, estudiante_id):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    estudiante = get_object_or_404(Persona, id=estudiante_id, colegio=colegio, tipo='ESTUDIANTE')
    
    ano_activo = AnioEscolar.objects.filter(colegio=colegio, activo=True).first()
    
    # Obtenemos los detalles a través de related_name
    estudiante_detalle = getattr(estudiante, 'detalle_estudiante', None)
    
    relacion = RelacionFamiliar.objects.filter(estudiante=estudiante).first()
    representante = relacion.representante if relacion else None
    parentesco = relacion.parentesco if relacion else None
    
    representante_detalle = getattr(representante, 'detalle_representante', None) if representante else None
    
    context = {
        'colegio': colegio,
        'estudiante': estudiante,
        'estudiante_detalle': estudiante_detalle,
        'representante': representante,
        'parentesco': parentesco,
        'representante_detalle': representante_detalle,
        'ano_activo': ano_activo,
    }
    return render(request, 'users/personas/ficha_estudiante_modal.html', context)

@login_required
def imprimir_ficha_estudiante_pdf(request, colegio_slug, estudiante_id):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    estudiante = get_object_or_404(Persona, id=estudiante_id, colegio=colegio, tipo='ESTUDIANTE')
    
    ano_activo = AnioEscolar.objects.filter(colegio=colegio, activo=True).first()
    estudiante_detalle = getattr(estudiante, 'detalle_estudiante', None)
    
    relacion = RelacionFamiliar.objects.filter(estudiante=estudiante).first()
    representante = relacion.representante if relacion else None
    parentesco = relacion.parentesco if relacion else None
    representante_detalle = getattr(representante, 'detalle_representante', None) if representante else None

    # --- EL TRUCO ESTÁ AQUÍ ---
    # Extraemos la ruta absoluta del sistema operativo si la foto existe
    foto_path = None
    if estudiante.foto and hasattr(estudiante.foto, 'path'):
        foto_path = estudiante.foto.path

    context = {
        'colegio': colegio,
        'estudiante': estudiante,
        'estudiante_detalle': estudiante_detalle,
        'representante': representante,
        'parentesco': parentesco,
        'representante_detalle': representante_detalle,
        'ano_activo': ano_activo,
        'foto_path': foto_path, # Pasamos la ruta física al template
    }

    template = get_template('users/personas/ficha_estudiante_pdf.html')
    html = template.render(context)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Ficha_Estudiante_{estudiante.cedula}.pdf"'
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    
    if pisa_status.err:
        return HttpResponse('Tuvimos errores generando el PDF', status=500)
    
    return response

@login_required
def editar_estudiante(request, colegio_slug, estudiante_id):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    estudiante = get_object_or_404(Persona, id=estudiante_id, colegio=colegio, tipo='ESTUDIANTE')

    if not request.user.rol.nombre == 'ADMINISTRADOR' and not request.user.is_superuser:
        messages.error(request, "No tienes permiso para editar registros.")
        return redirect('dashboard_colegio', colegio_slug=colegio.slug)
    
    relacion = RelacionFamiliar.objects.filter(estudiante=estudiante).first()
    representante = relacion.representante if relacion else None

    # Obtener Año Activo e Inscripción Actual
    anio_activo = AnioEscolar.objects.filter(colegio=colegio, activo=True).first()
    inscripcion_activa = None
    if anio_activo:
        inscripcion_activa = Inscripcion.objects.filter(estudiante=estudiante, anio_escolar=anio_activo).first()

    if request.method == 'POST':
        nuevo_email = request.POST.get('email').lower().strip()
        representante_id = request.POST.get('representante_id')
        parentesco = request.POST.get('parentesco', '').upper().strip()
        
        # --- CAPTURA DE LOS MANDOS ACADÉMICOS ---
        desinscribir_solicitado = 'desinscribir' in request.POST
        promover_solicitado = 'promover' in request.POST  # <-- NUEVO CONTROL
        
        try:
            with transaction.atomic():
                # 1. Actualizar Usuario vinculado al estudiante
                if estudiante.usuario:
                    user = estudiante.usuario
                    if user.email != nuevo_email:
                        if Usuario.objects.filter(email=nuevo_email).exclude(id=user.id).exists():
                            messages.error(request, f"El correo {nuevo_email} ya está en uso.")
                            return render(request, 'users/personas/editar_estudiante.html', {
                                'colegio': colegio, 'estudiante': estudiante, 'representante': representative, 
                                'relacion': relacion, 'anio_activo': anio_activo, 'inscripcion_activa': inscripcion_activa
                            })
                        
                        user.email = nuevo_email
                        user.username = nuevo_email
                        user.save()
                
                # 2. Obtener el representante asignado desde el formulario
                nuevo_rep = None
                if representante_id:
                    nuevo_rep = get_object_or_404(Persona, id=representante_id, colegio=colegio, tipo='REPRESENTANTE')

                # 3. Actualizar Datos Propios del Estudiante
                estudiante.nombre = request.POST.get('nombre').upper().strip()
                estudiante.apellido = request.POST.get('apellido').upper().strip()
                estudiante.cedula = request.POST.get('cedula').strip()
                estudiante.telefono = request.POST.get('telefono').strip()
                estudiante.fecha_nacimiento = request.POST.get('fecha_nacimiento')
                estudiante.activo = 'activo' in request.POST
                
                if nuevo_rep:
                    estudiante.direccion = nuevo_rep.direccion
                else:
                    estudiante.direccion = request.POST.get('direccion', '').upper().strip()
                
                estudiante.save()

                # 4. Reasignar o Crear la Relación Familiar
                if nuevo_rep:
                    if relacion:
                        relacion.representante = nuevo_rep
                        relacion.parentesco = parentesco
                        relacion.save()
                    else:
                        relacion = RelacionFamiliar.objects.create(
                            estudiante=estudiante,
                            representante=nuevo_rep,
                            parentesco=parentesco
                        )
                else:
                    if relacion:
                        relacion.delete()
                
                # 5. SECCIÓN DE PROCESAMIENTO ACADÉMICO
                if desinscribir_solicitado and inscripcion_activa:
                    # CASO A: Anulación total (Borra el historial)
                    inscripcion_activa.delete()
                    if hasattr(estudiante, 'detalle_academico') and estudiante.detalle_academico:
                        estudiante.detalle_academico.seccion = None
                        estudiante.detalle_academico.save()
                    messages.warning(request, f"La inscripción de {estudiante.nombre} para el año {anio_activo.nombre} fue anulada por completo.")
                
                elif promover_solicitado and inscripcion_activa:
                    # CASO B: Promoción / Pasar a Flotante (PRESERVA el historial de Inscripcion)
                    if hasattr(estudiante, 'detalle_academico') and estudiante.detalle_academico:
                        estudiante.detalle_academico.seccion = None  # Lo removemos del aula actual
                        estudiante.detalle_academico.save()
                    messages.success(request, f"¡{estudiante.nombre} ha sido promovido con éxito! El estudiante quedó en estado 'Flotante' y su registro de inscripción histórico permanece resguardado.")
                
                else:
                    # Mensaje estándar si solo modificó datos personales
                    messages.success(request, f"El registro de {estudiante.nombre} se actualizó correctamente.")
                
                return redirect('gestionar_estudiantes', colegio_slug=colegio.slug)
        
        except Exception as e:
            messages.error(request, f"Error al guardar: {str(e)}")

    return render(request, 'users/personas/editar_estudiante.html', {
        'colegio': colegio, 'estudiante': estudiante, 'representante': representante,
        'relacion': relacion, 'anio_activo': anio_activo, 'inscripcion_activa': inscripcion_activa
    })

# NUEVA VISTA: Buscador rápido via AJAX para el formulario de edición
@login_required
def buscar_representante_ajax_editar_estudiante(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    cedula = request.GET.get('cedula', '').strip()
    
    if not cedula:
        return JsonResponse({'success': False, 'message': 'Debe ingresar un número de cédula.'})
        
    try:
        rep = Persona.objects.get(cedula=cedula, colegio=colegio, tipo='REPRESENTANTE')
        return JsonResponse({
            'success': True,
            'id': rep.id,
            'nombre': rep.nombre,
            'apellido': rep.apellido,
            'cedula': rep.cedula,
            'telefono': rep.telefono or 'No registrado',
            'direccion': rep.direccion or 'No registrada'
        })
    except Persona.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'No se encontró ningún representante registrado con esa cédula.'})

@login_required
@require_POST
def desactivar_estudiante(request, colegio_slug, estudiante_id):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    if not request.user.rol.nombre == 'ADMINISTRADOR' and not request.user.is_superuser:
        return JsonResponse({'status': 'error', 'message': 'No autorizado'}, status=403)

    estudiante = get_object_or_404(Persona, id=estudiante_id, colegio=colegio)
    hoy = timezone.now().date()

    with transaction.atomic():
        # 1. Desactivar al estudiante
        estudiante.activo = False
        estudiante.save()

        # 2. Retirar de la sección
        Inscripcion.objects.filter(estudiante=estudiante, estado='ACTIVO').update(estado='RETIRADO')

        # 3. ¡LA MAGIA DE LOS PAGOS!: Exonerar cuotas futuras pendientes
        Pago.objects.filter(
            estudiante=estudiante,
            pagado=False,          # Solo los que no ha pagado
            activo=True,           # Solo los que estén vigentes
            fecha_vencimiento__gt=hoy  # Que venzan después de hoy (Julio, Agosto, etc.)
        ).update(
            activo=False,
            usuario=request.user,
            motivo="EXONERADO POR RETIRO DEL ESTUDIANTE"
        )

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'status': 'success', 'message': f'{estudiante.nombre} ha sido retirado y sus cuotas futuras fueron exoneradas.'})
    
    messages.warning(request, f"El estudiante {estudiante.nombre} ha sido desactivado y sus cuotas futuras fueron exoneradas.")
    return redirect('gestionar_estudiantes', colegio_slug=colegio.slug)


@login_required
def activar_estudiante(request, colegio_slug, persona_id):
    persona = get_object_or_404(Persona, id=persona_id, colegio__slug=colegio_slug, tipo='ESTUDIANTE')
    
    with transaction.atomic():
        # 1. Reincorporar la ficha
        persona.activo = True
        persona.save()
        
        # 2. Volver a activar su inscripción
        Inscripcion.objects.filter(estudiante=persona, estado='RETIRADO').update(estado='ACTIVO')
        
        # 3. Restaurar los pagos que fueron exonerados exclusivamente por el retiro
        Pago.objects.filter(
            estudiante=persona,
            activo=False,
            motivo="EXONERADO POR RETIRO DEL ESTUDIANTE"
        ).update(
            activo=True,
            usuario=request.user,
            motivo="RESTAURADO POR REINCORPORACIÓN"  # O dejarlo en None, según prefieras
        )
    
    messages.success(request, f"El estudiante {persona.nombre} ha sido reincorporado y sus cuentas reactivadas.")
    return redirect(f"{reverse('gestionar_estudiantes', kwargs={'colegio_slug': colegio_slug})}?tab=inactivos")

# --- EXPORTAR ESTUDIANTES A EXCEL ---
@login_required
def exportar_estudiantes_excel(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    # Filtramos las personas que son estudiantes en este colegio
    estudiantes = Persona.objects.filter(colegio=colegio, tipo='ESTUDIANTE', activo=True)

    data = []
    for est in estudiantes:
        # 1. Intentamos buscar la sección desde el detalle académico o su última inscripción activa
        seccion_obj = None
        if hasattr(est, 'detalle_academico') and est.detalle_academico.seccion:
            seccion_obj = est.detalle_academico.seccion
        else:
            ultima_ins = est.inscripciones.filter(estado='ACTIVO').select_related('seccion__grado').last()
            if ultima_ins:
                seccion_obj = ultima_ins.seccion

        # 2. Evaluación de solvencia simulada/segura 
        # (Si en el futuro agregas un método o campo solvente en Persona o Inscripción, el código lo usará)
        es_solvente = True
        if hasattr(est, 'solvente'):
            es_solvente = est.solvente() if callable(est.solvente) else est.solvente
        elif hasattr(est, 'detalle_academico') and hasattr(est.detalle_academico, 'solvente'):
            es_solvente = est.detalle_academico.solvente() if callable(est.detalle_academico.solvente) else est.detalle_academico.solvente

        data.append({
            'Cédula': est.cedula,
            'Nombre': est.nombre,
            'Apellido': est.apellido,
            'Grado': seccion_obj.grado.nombre if (seccion_obj and seccion_obj.grado) else 'SIN ASIGNAR',
            'Sección': seccion_obj.nombre if seccion_obj else 'SIN ASIGNAR',
            'Beca': f"{est.detalle_academico.porcentaje_beca}%" if (hasattr(est, 'detalle_academico') and est.detalle_academico.becado) else 'NINGUNA',
            'Estado Pago': 'SOLVENTE' if es_solvente else 'DEUDOR',
            'Teléfono': est.telefono or 'N/A',
            'Fecha de Nacimiento': est.fecha_nacimiento.strftime('%d/%m/%Y') if est.fecha_nacimiento else 'N/A'
        })

    df = pd.DataFrame(data)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=estudiantes_{colegio.slug}.xlsx'
    
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Estudiantes')
    
    return response


# --- EXPORTAR ESTUDIANTES A PDF  ---
@login_required
def exportar_estudiantes_pdf(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    # Traemos los estudiantes ordenados por apellido
    estudiantes = Persona.objects.filter(colegio=colegio, tipo='ESTUDIANTE', activo=True).order_by('apellido')
    
    context = {
        'colegio': colegio,
        'estudiantes': estudiantes,
        'fecha': datetime.now(),
    }
    
    # Renderizamos la plantilla HTML destinada al PDF
    html_string = render_to_string('users/personas/pdf_estudiantes.html', context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="lista_estudiantes_{colegio.slug}.pdf"'
    
    # Compilar usando xhtml2pdf de forma segura
    pisa_status = pisa.CreatePDF(BytesIO(html_string.encode("UTF-8")), dest=response)
    if pisa_status.err:
        return HttpResponse('Ocurrió un error al generar el PDF', status=500)
    return response

@login_required
def descargar_plantilla_estudiantes(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    # Verificación de seguridad básica (mismo permiso que gestionar estudiantes)
    # Suponiendo que request.user.rol tiene permisos

    # Crear el libro y la hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Carga Estudiantes"

    # Definir los encabezados (Nombres de columna exactos para procesar después)
    encabezados = [
        "CEDULA_REPRESENTANTE", "NOMBRE_REPRESENTANTE", "APELLIDO_REPRESENTANTE", "EMAIL_REPRESENTANTE", "TELEFONO_REPRESENTANTE",
        "CEDULA_ESTUDIANTE", "NOMBRE_ESTUDIANTE", "APELLIDO_ESTUDIANTE", "FECHA_NACIMIENTO_ESTUDIANTE (AAAA-MM-DD)",
        "PARENTESCO (con el est.)", "GRADO (nombre exacto)", "SECCION (letra)"
    ]

    # Estilo para los encabezados (negrita y fondo gris)
    fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    font = Font(bold=True)

    # Escribir encabezados y aplicar estilo
    for col_num, header in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = font
        cell.fill = fill
        # Ajuste básico de ancho
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = len(header) + 5

    # --- DATOS DE REFERENCIA (Opcional, para ayudar al usuario) ---
    # Podríamos agregar hojas extra con la lista de grados y secciones válidos,
    # o usar validación de datos en el Excel, pero para simplificar,
    # solo definimos los encabezados.

    # Preparar la respuesta HTTP para descarga
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="plantilla_carga_estudiantes_{colegio.slug}.xlsx"'
    
    wb.save(response)
    return response

@login_required
@require_POST
@transaction.atomic # Toda la operación debe ser atómica
def procesar_carga_masiva_estudiantes(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    # Seguridad (mismo permiso que crear_estudiante_completo)

    archivo_excel = request.FILES.get('archivo_excel')
    if not archivo_excel:
        messages.error(request, "No se seleccionó ningún archivo.")
        return redirect('gestionar_estudiantes', colegio_slug=colegio.slug)

    # Usaremos pandas porque facilita mucho la validación de columnas y tipos de datos
    try:
        df = pd.read_excel(archivo_excel, dtype={'CEDULA_REPRESENTANTE': str, 'CEDULA_ESTUDIANTE': str})
    except Exception as e:
        messages.error(request, f"Error al leer el archivo Excel: {e}")
        return redirect('gestionar_estudiantes', colegio_slug=colegio.slug)

    # 1. Validar Encabezados de Columna Obligatorios
    encabezados_obligatorios = [
        "CEDULA_REPRESENTANTE", "NOMBRE_REPRESENTANTE", "APELLIDO_REPRESENTANTE",
        "CEDULA_ESTUDIANTE", "NOMBRE_ESTUDIANTE", "APELLIDO_ESTUDIANTE",
        "PARENTESCO (con el est.)", "GRADO (nombre exacto)", "SECCION (letra)"
    ]
    
    if not all(col in df.columns for col in encabezados_obligatorios):
        messages.error(request, "La plantilla de Excel no es válida. Faltan columnas obligatorias o los encabezados fueron modificados.")
        return redirect('gestionar_estudiantes', colegio_slug=colegio.slug)

    # 2. Configuración Base para la Carga
    anio_activo = AnioEscolar.objects.filter(activo=True).first()
    if not anio_activo:
        messages.error(request, "No hay un Año Escolar activo configurado. No se puede proceder con la inscripción masiva.")
        return redirect('gestionar_estudiantes', colegio_slug=colegio.slug)

    # Roles necesarios (asumimos que existen, sino, lanzaremos error)
    rol_rep = Rol.objects.filter(colegio=colegio, nombre__iexact='Representante').first()
    rol_est = Rol.objects.filter(colegio=colegio, nombre__iexact='Estudiante').first()
    
    if not rol_rep or not rol_est:
        messages.error(request, "Los roles 'Representante' y 'Estudiante' deben estar configurados en este colegio antes de la carga masiva.")
        return redirect('gestionar_roles', colegio_slug=colegio.slug)

    estudiantes_creados_count = 0
    errores = []

    # 3. Iterar fila por fila (fila es una tupla, index es el número de fila)
    for index, row in df.iterrows():
        fila_excel = index + 2 # Para reportar errores, la primera fila de datos es la 2

        try:
            # 3.1. CAPTURAR Y LIMPIAR DATOS DE LA FILA
            # Usamos .get() y .strip() para evitar errores si la celda está vacía o tiene espacios extra.
            
            # Datos Representante
            cedula_rep = str(row['CEDULA_REPRESENTANTE']).strip().upper()
            nombre_rep = str(row['NOMBRE_REPRESENTANTE']).strip().upper()
            apellido_rep = str(row['APELLIDO_REPRESENTANTE']).strip().upper()
            email_rep = str(row['EMAIL_REPRESENTANTE']).strip().lower() if pd.notnull(row['EMAIL_REPRESENTANTE']) else None
            telefono_rep = str(row['TELEFONO_REPRESENTANTE']).strip() if pd.notnull(row['TELEFONO_REPRESENTANTE']) else ""

            # Datos Estudiante
            cedula_est = str(row['CEDULA_ESTUDIANTE']).strip().upper()
            nombre_est = str(row['NOMBRE_ESTUDIANTE']).strip().upper()
            apellido_est = str(row['APELLIDO_ESTUDIANTE']).strip().upper()
            fecha_nac_str = str(row['FECHA_NACIMIENTO_ESTUDIANTE (AAAA-MM-DD)']).strip() if pd.notnull(row['FECHA_NACIMIENTO_ESTUDIANTE (AAAA-MM-DD)']) else None
            
            # Datos Académicos
            parentesco = str(row['PARENTESCO (con el est.)']).strip().upper()
            grado_nombre = str(row['GRADO (nombre exacto)']).strip()
            seccion_letra = str(row['SECCION (letra)']).strip().upper()

            # Validación de campos obligatorios en la fila
            if not all([cedula_rep, nombre_rep, apellido_rep, cedula_est, nombre_est, apellido_est, parentesco, grado_nombre, seccion_letra]):
                raise ValidationError(f"Fila {fila_excel}: Faltan datos obligatorios.")

            # 3.2. BUSCAR GRADO Y SECCION
            # El Excel pide '1er Grado' y 'A'
            seccion = Seccion.objects.filter(
                grado__nombre=grado_nombre, 
                grado__nivel__colegio=colegio, 
                nombre=seccion_letra
            ).first()

            if not seccion:
                raise ValidationError(f"Fila {fila_excel}: No se encontró el Grado '{grado_nombre}' o la Sección '{seccion_letra}' válidos para este colegio.")

            # 3.3. PROCESAR REPRESENTANTE (get_or_create)
            # Buscamos por cédula y colegio
            representante, created_rep = Persona.objects.get_or_create(
                cedula=cedula_rep,
                colegio=colegio,
                defaults={
                    'nombre': nombre_rep,
                    'apellido': apellido_rep,
                    'telefono': telefono_rep,
                    'es_representante': True,
                    'tipo': 'REPRESENTANTE',
                }
            )

            # Si el representante ya existía pero no estaba marcado como representante, lo actualizamos
            if not created_rep:
                # Opcional: Actualizar teléfono o email si cambiaron, pero lo más seguro
                # es mantener los datos existentes para no sobrescribir información crítica.
                if not representante.es_representante:
                    representante.es_representante = True
                    representante.tipo = 'REPRESENTANTE'
                    representante.save()

            # 3.4. USUARIO PARA EL REPRESENTANTE (Si no tiene)
            # Solo si el Excel proporcionó un email válido
            if email_rep and not representante.usuario:
                # Verificar si el email ya está en uso por otra persona
                if not Usuario.objects.filter(email=email_rep).exists():
                    user_rep = Usuario.objects.create_user(
                        email=email_rep,
                        password=cedula_rep, # Contraseña por defecto: cédula
                        colegio=colegio,
                        rol=rol_rep
                    )
                    representante.usuario = user_rep
                    representante.save()
                else:
                    # Opcional: ¿Qué hacer si el email existe pero no está vinculado a este representante?
                    # Por simplicidad, registramos el error en la fila
                    #errores.append(f"Fila {fila_excel}: El email de representante '{email_rep}' ya está en uso por otro usuario.")
                    #raise ValidationError(f"Email en uso.")
                    pass # Seguimos adelante sin crear usuario

            # 3.5. PROCESAR ESTUDIANTE (create)
            # El estudiante siempre se crea de cero en la carga masiva
            # Verificamos si la cédula ya existe para este colegio
            if Persona.objects.filter(cedula=cedula_est, colegio=colegio).exists():
                 raise ValidationError(f"Fila {fila_excel}: Ya existe un estudiante con la cédula '{cedula_est}' en este colegio.")

            estudiante = Persona.objects.create(
                colegio=colegio,
                cedula=cedula_est,
                nombre=nombre_est,
                apellido=apellido_est,
                telefono=telefono_rep, # Hereda el teléfono del representante si no se especificó otro
                es_estudiante=True,
                tipo='ESTUDIANTE'
            )

            # 3.6. USUARIO PARA EL ESTUDIANTE (Automático)
            # Formato: cedula@colegio.com
            user_est_email = f"{cedula_est}@{colegio.slug}.com"
            if not Usuario.objects.filter(email=user_est_email).exists():
                user_est = Usuario.objects.create_user(
                    email=user_est_email,
                    password=cedula_est,
                    colegio=colegio,
                    rol=rol_est
                )
                estudiante.usuario = user_est
                estudiante.save()

            # 3.7. VINCULACIÓN FAMILIAR
            RelacionFamiliar.objects.create(
                representante=representante,
                estudiante=estudiante,
                parentesco=parentesco
                # Opcional: Podríamos validar que el parentesco sea uno de los predefinidos (Padre, Madre, etc.)
            )

            # 3.8. INSCRIPCIÓN ACADÉMICA
            Inscripcion.objects.create(
                estudiante=estudiante,
                seccion=seccion,
                anio_escolar=anio_activo,
                costo_mensualidad=0, # Monto por defecto, se ajusta manualmente después si es necesario
                estado='ACTIVO'
            )

            estudiantes_creados_count += 1

        except ValidationError as ve:
            errores.append(str(ve))
            # No hacemos nada más, la transacción del rollback asegurará que no se guarde nada parcial
            # para esta fila, pero las filas anteriores si se guardaron.
        except Exception as e:
            # Captura errores inesperados de base de datos o lógica
            # Esto debería ser raro dentro de una transacción atómica y get_or_create
            # Pero es bueno para el debugging
            errores.append(f"Fila {fila_excel}: Error inesperado: {str(e)}")
            # En caso de error inesperado, forzamos un rollback completo de TODA la carga
            # para evitar que el Excel se cargue a medias, excepto las filas ya procesadas.
            # Por simplicidad, como ya estamos reportando errores fila por fila,
            # solo seguimos adelante.

    # 4. Reporte de Resultados
    if errores:
        # Mostramos los primeros 5 errores para no saturar
        errores_limitados = errores[:5]
        errores_count = len(errores)
        if errores_count > 5:
            errores_limitados.append(f"... y {errores_count - 5} errores más.")
        
        # Un mensaje de error grande
        mensaje_error = "Se encontraron problemas al procesar el archivo Excel. No se guardaron los registros de las filas con error:<br><br>" + "<br>".join(errores_limitados)
        messages.error(request, mensaje_error, extra_tags='safe') # safe para que el HTML se renderice

    # Mensaje de éxito si se creó al menos uno
    if estudiantes_creados_count > 0:
        messages.success(request, f"Se han inscrito exitosamente {estudiantes_creados_count} estudiantes desde el archivo Excel.")

    return redirect('gestionar_estudiantes', colegio_slug=colegio.slug)

login_required
def gestionar_representantes(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    ver_inactivos = request.GET.get('estado') == 'inactivo'

     # Seguridad: Solo Admin y Super pueden gestionar personas
    if request.user.rol.nombre not in ['Admin', 'Super']:
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied

    # 1. OBTENER PARÁMETROS DE FILTRO Y ORDEN
    query = request.GET.get('q', '')
    ver_inactivos = request.GET.get('estado') == 'inactivo'
    order_by = request.GET.get('order', 'apellido') # Orden por defecto
    per_page = request.GET.get('per_page', '10')    # Registros por defecto

    # 2. FILTRADO BASE CON PREFETCH EFICIENTE
    # Traemos las relaciones familiares cruzando de una vez con el estudiante (select_related)
    prefetch_estudiantes = Prefetch(
        'representados', # <-- Este debe ser el related_name de tu FK en RelacionFamiliar
        queryset=RelacionFamiliar.objects.select_related('estudiante'),
        to_attr='mis_estudiantes' # <-- Nombre de la lista interna que usaremos en el template
    )
    
    # 2. FILTRADO BASE
    representantes_list = Persona.objects.filter(
        colegio=colegio, 
        tipo='REPRESENTANTE', 
        activo=not ver_inactivos
    ).prefetch_related(prefetch_estudiantes).order_by(order_by)

    if query:
        representantes_list = representantes_list.filter(
            Q(nombre__icontains=query) | 
            Q(apellido__icontains=query) | 
            Q(cedula__icontains=query)
        )

    # 3. PAGINACIÓN
    paginator = Paginator(representantes_list, per_page)
    page_number = request.GET.get('page')
    representantes = paginator.get_page(page_number)

    return render(request, 'users/personas/gestionar_representantes.html', {
        'colegio': colegio,
        'representantes': representantes,
        'query': query,
        'order_by': order_by,
        'per_page': per_page,
    })

@login_required
def crear_representante(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)

    if request.user.rol.nombre != 'Admin' and request.user.rol.nombre != 'Super':
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied
    
    if request.method == 'POST':

        email = request.POST.get('email', '').strip().lower()
        cedula = request.POST.get('cedula', '').strip().upper()
        nombre = request.POST.get('nombre', '').strip().upper()
        apellido = request.POST.get('apellido', '').strip().upper()
        telefono = request.POST.get('telefono', '').strip().upper()
        profesion = request.POST.get('profesion', '').strip().upper()
        direccion = request.POST.get('direccion', '').strip().upper()
        fecha_nac = request.POST.get('fecha_nacimiento')

        # Validar si el correo ya está registrado en este colegio
        if Usuario.objects.filter(email=email, colegio=colegio).exists():
            messages.error(request, "EL CORREO ELECTRÓNICO YA SE ENCUENTRA REGISTRADO.")
            return render(request, 'users/personas/crear_representante.html', {'colegio': colegio})

        # Validar si la cédula ya está registrada en este colegio
        if Persona.objects.filter(cedula=cedula, colegio=colegio).exists():
            messages.error(request, "LA CÉDULA DE IDENTIDAD YA SE ENCUENTRA REGISTRADA.")
            return render(request, 'users/personas/crear_representante.html', {'colegio': colegio})

        try:
            # Importante: Asegúrate de que el Rol 'Representante' existe en este colegio
            rol_representante = Rol.objects.filter(colegio=colegio, nombre__iexact='Representante').first()
            if not rol_representante:
                print("ERROR: Debes crear primero el Rol 'Representante' en este colegio.")
                # Podrías crear el rol aquí mismo si no existe

            # Usamos una transacción para asegurar que se creen ambos o ninguno
            from django.db import transaction
            with transaction.atomic():
                nuevo_usuario = Usuario.objects.create_user(
                    email=email,
                    password=cedula,
                    colegio=colegio,
                    rol=rol_representante
                )

                persona = Persona.objects.create(
                    usuario=nuevo_usuario,
                    colegio=colegio,
                    cedula=cedula,
                    nombre=nombre,
                    apellido=apellido,
                    fecha_nacimiento=fecha_nac,
                    telefono=telefono,
                    direccion=direccion,
                    profesion=profesion,
                    es_representante=True,
                    tipo='REPRESENTANTE'
                )
           
            print("¡ÉXITO! Usuario y Persona creados.")
            return redirect('gestionar_representantes', colegio_slug=colegio.slug)
            
        except Rol.DoesNotExist:
            messages.error(request, "El rol 'Representante' no existe. Créalo en la sección de Roles antes de continuar.")
        except Exception as e:
            messages.error(request, f"Error al crear el registro: {e}")
            
    return render(request, 'users/personas/crear_representante.html', {'colegio': colegio})

@login_required
def verificar_cedula_representante(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    cedula = request.GET.get('cedula', '').strip().upper()
    
    if not cedula:
        return JsonResponse({'error': 'No se proporcionó una cédula'}, status=400)
    
    # Buscamos si ya existe una persona con esa cédula en el colegio
    persona_existente = Persona.objects.filter(cedula=cedula, colegio=colegio).first()
    
    if persona_existente:
        return JsonResponse({
            'existe': True,
            'tipo': persona_existente.tipo,
            'nombre': persona_existente.nombre,
            'apellido': persona_existente.apellido,
            'email': persona_existente.usuario.email if persona_existente.usuario else '',
            'telefono': persona_existente.telefono,
            'profesion': persona_existente.profesion,
            'direccion': persona_existente.direccion,
            'fecha_nacimiento': persona_existente.fecha_nacimiento.strftime('%Y-%m-%d') if persona_existente.fecha_nacimiento else ''
        })
        
    return JsonResponse({'existe': False})

@login_required
def editar_representante(request, colegio_slug, persona_id):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    # Buscamos la persona asegurándonos que sea un REPRESENTANTE de este colegio
    representante = get_object_or_404(Persona, id=persona_id, colegio=colegio, tipo='REPRESENTANTE')

    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()
        cedula = request.POST.get('cedula', '').strip().upper()
        nombre = request.POST.get('nombre', '').strip().upper()
        apellido = request.POST.get('apellido', '').strip().upper()
        telefono = request.POST.get('telefono', '').strip().upper()
        profesion = request.POST.get('profesion', '').strip().upper()
        fecha_nac = request.POST.get('fecha_nacimiento')
        direccion = request.POST.get('direccion', '').strip().upper()

        try:
            with transaction.atomic():
                # 1. Actualizar datos del Usuario vinculado
                usuario = representante.usuario
                if usuario:
                    usuario.email = email
                    # Si quieres que la cédula nueva sea la clave si cambió, podrías usar set_password
                    # pero normalmente solo actualizamos el email aquí.
                    usuario.save()

                # 2. Actualizar datos de la Persona
                representante.cedula = cedula
                representante.nombre = nombre
                representante.apellido = apellido
                representante.telefono = telefono
                representante.direccion = direccion
                representante.profesion = profesion
                representante.fecha_nacimiento = fecha_nac
                representante.save()

            messages.success(request, f"REPRESENTANTE {nombre} ACTUALIZADO CORRECTAMENTE.")
            return redirect('gestionar_representantes', colegio_slug=colegio.slug)
        except Exception as e:
            messages.error(request, f"ERROR AL ACTUALIZAR: {str(e).upper()}")

    return render(request, 'users/personas/editar_representante.html', {
        'colegio': colegio,
        'representante': representante
    })

@login_required
def ficha_representante_modal(request, colegio_slug, pk):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    # Buscamos la persona y nos aseguramos que sea representante de ese colegio
    representante = get_object_or_404(Persona, pk=pk, colegio=colegio, tipo='REPRESENTANTE')
    relaciones = RelacionFamiliar.objects.filter(representante=representante).select_related('estudiante')

    # Aquí podrías buscar también a los estudiantes asociados si tienes la relación lista
    # estudiantes = representante.estudiantes.all() 

    context = {
        'representante': representante,
        'colegio': colegio,
        'relaciones': relaciones,
    }
    return render(request, 'users/personas/ficha_representante_content.html', context)

@login_required
def imprimir_ficha_representante_pdf(request, colegio_slug, pk):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    representante = get_object_or_404(Persona, pk=pk, colegio=colegio, tipo='REPRESENTANTE')
    
    # Obtenemos las relaciones familiares
    relaciones = RelacionFamiliar.objects.filter(representante=representante).select_related('estudiante')

    context = {
        'representante': representante,
        'colegio': colegio,
        'relaciones': relaciones,
        'fecha': datetime.now(),
    }
    
    html_string = render_to_string('users/personas/ficha_representante_pdf.html', context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Ficha_{representante.cedula}.pdf"'
    
    pisa_status = pisa.CreatePDF(
        BytesIO(html_string.encode("UTF-8")), 
        dest=response,
        encoding='utf-8'
    )
    
    return response

@login_required
def eliminar_representante(request, colegio_slug, persona_id):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    representante = get_object_or_404(Persona, id=persona_id, colegio=colegio, tipo='REPRESENTANTE')

    try:
        with transaction.atomic():
            representante.activo = False # Marcamos como inactivo
            representante.save()
            
            # También desactivamos su usuario para que no pueda entrar al sistema
            if representante.usuario:
                representante.usuario.is_active = False
                representante.usuario.save()
                
        messages.success(request, f"EL REPRESENTANTE {representante.nombre} HA SIDO DESACTIVADO.")
    except Exception as e:
        messages.error(request, f"ERROR AL DESACTIVAR: {str(e).upper()}")
    
    return redirect('gestionar_representantes', colegio_slug=colegio.slug)

@login_required
def reactivar_representante(request, colegio_slug, persona_id):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    representante = get_object_or_404(Persona, id=persona_id, colegio=colegio, tipo='REPRESENTANTE')

    try:
        with transaction.atomic():
            representante.activo = True
            representante.save()
            if representante.usuario:
                representante.usuario.is_active = True
                representante.usuario.save()
        messages.success(request, f"EL REPRESENTANTE {representante.nombre} HA SIDO REACTIVADO.")
    except Exception as e:
        messages.error(request, f"ERROR AL REACTIVAR: {str(e).upper()}")
    
    return redirect('gestionar_representantes', colegio_slug=colegio.slug)

def descargar_plantilla_representantes(request):
    import pandas as pd
    from django.http import HttpResponse
    from io import BytesIO

    # Definimos las columnas necesarias para representantes
    columnas = ['CEDULA', 'NOMBRE', 'APELLIDO', 'EMAIL', 'TELEFONO', 'DIRECCION', 'FECHA_NACIMIENTO', 'PROFESION']
    
    # Creamos un DataFrame vacío
    df = pd.DataFrame(columns=columnas)
    
    # Configuramos la respuesta HTTP para descargar el archivo
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Plantilla')
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_representantes.xlsx'
    return response

@login_required
def cargar_masiva_representantes(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    # Seguridad básica
    if request.user.rol.nombre not in ['Admin', 'Super']:
        raise PermissionDenied

    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        archivo = request.FILES['archivo_excel']
        
        try:
            df = pd.read_excel(archivo)
            # Normalizar nombres de columnas
            df.columns = [c.strip().upper() for c in df.columns]
            
            creados = 0
            errores = []
            
            # Obtener el Rol de Representante para este colegio
            rol_representante = Rol.objects.filter(colegio=colegio, nombre__iexact='Representante').first()
            
            with transaction.atomic():
                for index, row in df.iterrows():
                    try:
                        email = str(row['EMAIL']).strip().lower()
                        cedula = str(row['CEDULA']).strip().upper()

                        if Usuario.objects.filter(email=email).exists():
                            errores.append(f"Fila {index+2}: El correo {email} ya existe.")
                            continue

                        # 1. Crear Usuario (clave por defecto es la cédula)
                        user = Usuario.objects.create_user(
                            email=email,
                            password=cedula,
                            colegio=colegio,
                            rol=rol_representante
                        )

                        # 2. Crear Persona
                        persona = Persona.objects.create(
                            usuario=user,
                            colegio=colegio,
                            cedula=cedula,
                            nombre=str(row['NOMBRE']).strip().upper(),
                            apellido=str(row['APELLIDO']).strip().upper(),
                            telefono=str(row.get('TELEFONO', '')).strip(),
                            direccion=str(row.get('DIRECCION', '')).strip().upper(),
                            profesion=str(row.get('PROFESION', '')).strip().upper(),
                            fecha_nacimiento=pd.to_datetime(row['FECHA_NACIMIENTO']).date() if pd.notnull(row.get('FECHA_NACIMIENTO')) else None,
                            es_representante=True,
                            tipo='REPRESENTANTE'
                        )
                        
                        creados += 1

                    except Exception as e:
                        errores.append(f"Fila {index+2}: {str(e)}")

            if creados > 0:
                messages.success(request, f"¡Éxito! Se registraron {creados} nuevos representantes.")
            if errores:
                messages.error(request, f"Errores en {len(errores)} filas. Revise el formato del archivo.")

        except Exception as e:
            messages.error(request, f"Error crítico al leer el Excel: {e}")
            
    return redirect('gestionar_representantes', colegio_slug=colegio.slug)

# --- EXPORTAR A EXCEL ---
@login_required
def exportar_representantes_excel(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    representantes = Persona.objects.filter(colegio=colegio, tipo='REPRESENTANTE', activo=True)

    data = []
    for rep in representantes:
        data.append({
            'Cédula': rep.cedula,
            'Nombre': rep.nombre,
            'Apellido': rep.apellido,
            'Profesión': rep.profesion or 'N/A',
            'Edad': rep.edad,
            'Dirección': rep.direccion or 'N/A',
            'Teléfono': rep.telefono or 'N/A',
            'Email': rep.usuario.email if rep.usuario else 'N/A'
        })

    df = pd.DataFrame(data)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=representantes_{colegio.slug}.xlsx'
    
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Representantes')
    
    return response

# --- EXPORTAR A PDF ---
@login_required
def exportar_representantes_pdf(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    representantes = Persona.objects.filter(colegio=colegio, tipo='REPRESENTANTE', activo=True).order_by('apellido')
    
    context = {
        'colegio': colegio,
        'representantes': representantes,
        'fecha': datetime.now(),
    }
    
    # Renderizamos el template específico para PDF
    html_string = render_to_string('users/personas/pdf_representantes.html', context)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="lista_representantes_{colegio.slug}.pdf"'
    
    # Crear el PDF
    pisa_status = pisa.CreatePDF(BytesIO(html_string.encode("UTF-8")), dest=response)
    
    if pisa_status.err:
        return HttpResponse('Ocurrió un error al generar el PDF', status=500)
    return response

@login_required
def descargar_boleta_representante(request, colegio_slug, estudiante_id):
    # 1. Obtención y validación base de los objetos
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    estudiante = get_object_or_404(Persona, id=estudiante_id, colegio=colegio, es_estudiante=True)
    
    # 2. Control de Acceso: Validar Rol de Representante
    if not request.user.rol or request.user.rol.nombre != 'Representante':
        raise PermissionDenied

    # 3. Control de Acceso: Validar que el estudiante sea su representado legítimo
    representante_persona = getattr(request.user, 'perfil', None)
    if not representante_persona:
        raise PermissionDenied
        
    es_su_representado = RelacionFamiliar.objects.filter(
        representante=representante_persona,
        estudiante=estudiante
    ).exists()
    
    if not es_su_representado:
        raise PermissionDenied

    # 4. Control Administrativo: Validación de Solvencia en el Año Activo
    anio_activo = AnioEscolar.objects.filter(colegio=colegio, activo=True).order_by('-id').first()
    if not anio_activo:
        messages.error(request, "No se encuentra un año escolar activo configurado en el plantel.")
        return redirect('dashboard_colegio', colegio_slug=colegio.slug)
        
    tiene_deudas = Pago.objects.filter(
        estudiante=estudiante,
        anio_escolar=anio_activo,
        pagado=False,
        activo=True
    ).exists()
    
    if tiene_deudas:
        messages.error(request, "Estimado Representante, presenta compromisos administrativos pendientes para este año escolar.")
        return redirect('dashboard_colegio', colegio_slug=colegio.slug)

    # 5. Buscar la inscripción activa del estudiante para conocer su Sección y Grado
    inscripcion = Inscripcion.objects.filter(
        estudiante=estudiante,
        anio_escolar=anio_activo,
        estado='ACTIVO'
    ).select_related('seccion').first()

    if not inscripcion:
        messages.error(request, "El estudiante no posee una inscripción activa para el año escolar corriente.")
        return redirect('dashboard_colegio', colegio_slug=colegio.slug)

    # 6. EXTRACCIÓN DE CALIFICACIONES REALES (Estructuración matricial)
    # Buscamos todas las materias del plan de estudio vinculadas a la sección del alumno
    cargas_academicas = CargaAcademica.objects.filter(
        seccion=inscripcion.seccion
    ).select_related('asignatura')

    # Buscamos todas las notas registradas de este estudiante específico
    notas_alumno = NotaCualitativa.objects.filter(
        estudiante=estudiante,
        carga_academica__in=cargas_academicas
    )

    # Mapeamos las notas en un diccionario rápido: { carga_academica_id: { '1': 'C', '2': 'EP' } }
    mapa_notas = {}
    for nota in notas_alumno:
        if nota.carga_academica_id not in mapa_notas:
            mapa_notas[nota.carga_academica_id] = {}
        mapa_notas[nota.carga_academica_id][nota.lapso] = nota.calificacion

    # Construimos la lista final formateada que la plantilla recorrerá de forma limpia
    calificaciones_finales = []
    for carga in cargas_academicas:
        notas_materia = mapa_notas.get(carga.id, {})
        calificaciones_finales.append({
            'asignatura': carga.asignatura.nombre,
            'lapso1': notas_materia.get('1', ''),  # Mapea LapsoChoices.LAPSO_1
            'lapso2': notas_materia.get('2', ''),  # Mapea LapsoChoices.LAPSO_2
            'lapso3': notas_materia.get('3', ''),  # Mapea LapsoChoices.LAPSO_3
        })

    # 7. Envío de contexto hacia el template preparado para impresión
    context = {
        'colegio': colegio,
        'estudiante': estudiante,
        'anio_escolar': anio_activo,
        'inscripcion': inscripcion,
        'calificaciones': calificaciones_finales,
    }
    
    return render(request, 'users/boleta_imprimible.html', context)

def _verificar_permiso_administrativo(user):
    """Auxiliar para validar si el usuario es Superusuario o Rol Admin/Directivo"""
    es_superuser = user.is_superuser
    es_staff = user.is_staff
    es_rol_admin = user.rol and user.rol.nombre in ['Admin', 'Director', 'Coordinador']
    if not (es_superuser or es_staff or es_rol_admin):
        raise PermissionDenied

@login_required
def ver_calificaciones_seccion(request, colegio_slug, seccion_id):
    # 1. Garantizar Seguridad Administrativa
    _verificar_permiso_administrativo(request.user)

    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    seccion = get_object_or_404(Seccion, id=seccion_id, colegio=colegio)
    anio_activo = seccion.anio_escolar

    # 2. Traer alumnos activos de la sección
    inscripciones = Inscripcion.objects.filter(
        seccion=seccion,
        estado='ACTIVO'
    ).select_related('estudiante').order_by('estudiante__apellido', 'estudiante__nombre')

    # 3. Traer el plan de estudios (materias) de la sección
    cargas_academicas = CargaAcademica.objects.filter(seccion=seccion).select_related('asignatura', 'docente')

    # 4. Auditoría de Solvencia masiva
    estudiantes_ids = [ins.estudiante_id for ins in inscripciones]
    ids_con_deuda = set(
        Pago.objects.filter(
            estudiante_id__in=estudiantes_ids,
            anio_escolar=anio_activo,
            pagado=False,
            activo=True
        ).values_list('estudiante_id', flat=True)
    )

    # 5. Extraer TODAS las notas cualitativas de esta sección de un solo golpe
    notas_seccion = NotaCualitativa.objects.filter(
        carga_academica__seccion=seccion,
        estudiante_id__in=estudiantes_ids
    )

    # Indexar notas en un mapa de acceso rápido: { estudiante_id: { carga_id: { lapso: calificacion } } }
    matriz_notas = {}
    for nota in notas_seccion:
        est_id = nota.estudiante_id
        carga_id = nota.carga_academica_id
        if est_id not in matriz_notas:
            matriz_notas[est_id] = {}
        if carga_id not in matriz_notas[est_id]:
            matriz_notas[est_id][carga_id] = {}
        matriz_notas[est_id][carga_id][nota.lapso] = nota.calificacion

    # 6. Construir el empaquetado estructurado para el template
    alumnos_rendimiento = []
    for inscripcion in inscripciones:
        estudiante = inscripcion.estudiante
        estudiante.esta_solvente = estudiante.id not in ids_con_deuda
        
        # Mapear calificaciones de este estudiante por cada materia
        materias_evaluadas = []
        for carga in cargas_academicas:
            notas_lapso = matriz_notas.get(estudiante.id, {}).get(carga.id, {})
            materias_evaluadas.append({
                'asignatura': carga.asignatura.nombre,
                'docente': f"{carga.docente.apellido}" if carga.docente else "Sin asignar",
                'lapso1': notas_lapso.get('1', '-'),
                'lapso2': notas_lapso.get('2', '-'),
                'lapso3': notas_lapso.get('3', '-'),
            })
            
        alumnos_rendimiento.append({
            'estudiante': estudiante,
            'inscripcion': inscripcion,
            'materias': materias_evaluadas
        })

    context = {
        'colegio': colegio,
        'seccion': seccion,
        'alumnos_rendimiento': alumnos_rendimiento,
        'cargas_academicas': cargas_academicas,
        'total_estudiantes': inscripciones.count(),
        'total_solventes': len(estudiantes_ids) - len(ids_con_deuda)
    }
    return render(request, 'users/ver_calificaciones_seccion.html', context)

@login_required
def imprimir_boleta_admin(request, colegio_slug, estudiante_id):
    """Permite al Admin ver e imprimir la boleta saltándose el bloqueo de deudas"""
    # 1. Validación de seguridad
    _verificar_permiso_administrativo(request.user)
    
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    estudiante = get_object_or_404(Persona, id=estudiante_id, colegio=colegio, es_estudiante=True)
    
    # 2. Obtener la inscripción activa del alumno
    inscripcion = Inscripcion.objects.filter(estudiante=estudiante, estado='ACTIVO').select_related('seccion').first()
    if not inscripcion:
        # Fallback de seguridad por si acaso
        inscripcion = Inscripcion.objects.filter(estudiante=estudiante).select_related('seccion').first()
        
    seccion = inscripcion.seccion if inscripcion else None
    anio_escolar = seccion.anio_escolar if seccion else None
    
    # 3. Cargar las materias de la sección
    cargas = CargaAcademica.objects.filter(seccion=seccion).select_related('asignatura') if seccion else []
    boleta_rendimiento = []
    
    if seccion:
        # === CASO A: EVALUACIÓN CUANTITATIVA ===
        if seccion.es_cuantitativo:
            for c in cargas:
                planes = PlanEvaluacion.objects.filter(carga_academica=c)
                notas = NotaCuantitativa.objects.filter(plan_evaluacion__in=planes, estudiante=estudiante)
                
                lapsos = {'1': 0.0, '2': 0.0, '3': 0.0, 'has': {'1': False, '2': False, '3': False}}
                for n in notas:
                    lapsos[n.plan_evaluacion.lapso] += (n.nota * n.plan_evaluacion.ponderacion) / 100.0
                    lapsos['has'][n.plan_evaluacion.lapso] = True
                    
                l1 = round(lapsos['1']) if lapsos['has']['1'] else '--'
                l2 = round(lapsos['2']) if lapsos['has']['2'] else '--'
                l3 = round(lapsos['3']) if lapsos['has']['3'] else '--'
                
                validos = [v for v in [l1, l2, l3] if isinstance(v, int)]
                def_anual = round(sum(validos) / len(validos)) if validos else '--'
                
                # Diccionario híbrido (Soporta llaves 'l1' y 'lapso1' por si acaso)
                boleta_rendimiento.append({
                    'asignatura': c.asignatura.nombre,
                    'l1': l1, 'l2': l2, 'l3': l3, 'final': def_anual,
                    'lapso1': l1, 'lapso2': l2, 'lapso3': l3,
                })
                
        # === CASO B: EVALUACIÓN CUALITATIVA (Tu caso actual) ===
        else:
            for c in cargas:
                # Buscamos los Informes Cualitativos tal como lo hace tu módulo de docente
                informes = InformeCualitativo.objects.filter(carga_academica=c, estudiante=estudiante)
                inf_map = {i.lapso: i for i in informes}
                
                inf_l1 = inf_map.get('1')
                inf_l2 = inf_map.get('2')
                inf_l3 = inf_map.get('3')
                
                # Enviamos el objeto completo (para l1.descriptor) y texto plano como plan de respaldo
                boleta_rendimiento.append({
                    'asignatura': c.asignatura.nombre,
                    'l1': inf_l1, 
                    'l2': inf_l2, 
                    'l3': inf_l3,
                    # Fallback directo en texto por si el HTML de la boleta pide el string directamente
                    'lapso1': inf_l1.descriptor if inf_l1 else '-',
                    'lapso2': inf_l2.descriptor if inf_l2 else '-',
                    'lapso3': inf_l3.descriptor if inf_l3 else '-',
                })
                
    # 4. Construcción del contexto súper-compatible
    # Duplicamos los nombres de los listados bajo 'rendimiento' y 'calificaciones' 
    # para asegurar compatibilidad total con cualquiera de tus dos plantillas de boletas.
    context = {
        'colegio': colegio,
        'estudiante': estudiante,
        'seccion': seccion, 
        'anio': anio_escolar, 
        'anio_escolar': anio_escolar,
        'rendimiento': boleta_rendimiento, 
        'calificaciones': boleta_rendimiento, 
        'es_cuantitativo': seccion.es_cuantitativo if seccion else False, 
        'es_vista_admin': True
    }
    
    # Renderizamos la plantilla imprimible oficial
    return render(request, 'users/boleta_imprimible.html', context)

@login_required
def dashboard_cantina(request, colegio_slug):
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    # 🔒 Protección estricta: Solo súperuser o el Cantinero asignado a ESTE colegio
    if not request.user.is_superuser:
        if request.user.colegio != colegio or not request.user.rol or request.user.rol.nombre != 'Cantinero':
            raise PermissionDenied

    # 📦 1. Traer los rubros registrados de este colegio
    rubros = RubroCantina.objects.filter(colegio=colegio).order_by('nombre')
    
    # 👥 2. Traer los estudiantes del colegio e incluir su billetera (select_related evita lentitud)
    estudiantes = Persona.objects.filter(
        colegio=colegio,
        es_estudiante=True
    ).select_related('billetera_cantina').order_by('apellido', 'nombre')

    # 📥 3. Inyectar todo al contexto del template
    context = {
        'colegio': colegio,
        'rubros': rubros,
        'estudiantes': estudiantes,
    }

    return render(request, 'users/dashboard_cantina.html', context)

@login_required
@transaction.atomic  
def registrar_consumo_view(request, colegio_slug):
    """
    Procesa el cobro de la cantina descontando de la billetera prepago.
    Protegido contra concurrencia mediante bloqueos de fila (select_for_update).
    """
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    # 🔒 Protección estricta de seguridad identica a tu Dashboard
    if not request.user.is_superuser:
        if request.user.colegio != colegio or not request.user.rol or request.user.rol.nombre != 'Cantinero':
            raise PermissionDenied

    if request.method == 'POST':
        estudiante_id = request.POST.get('estudiante_id')
        
        # 1. Bloquear la billetera en la Base de Datos para esta transacción
        # Nadie más puede leer/modificar este saldo hasta que este bloque termine.
        try:
            billetera = (
                BilleteraCantina.objects
                .select_for_update()
                .get(estudiante_id=estudiante_id, estudiante__colegio=colegio)
            )
        except BilleteraCantina.DoesNotExist:
            messages.error(request, "El estudiante seleccionado no posee una billetera activa.")
            return redirect('dashboard_cantina', colegio_slug=colegio.slug)

        # 2. Capturar las listas de productos enviados por el formulario dinámico
        rubro_ids = request.POST.getlist('rubro_id')
        cantidades = request.POST.getlist('cantidad')
        precios = request.POST.getlist('precio')  # El precio manual que colocó el cantinero

        # Validación estructural básica
        if not rubro_ids or len(rubro_ids) != len(cantidades) or len(rubro_ids) != len(precios):
            messages.error(request, "El ticket de consumo está vacío o faltan datos.")
            return redirect('dashboard_cantina', colegio_slug=colegio.slug)

        # 3. Procesar renglón por renglón y calcular el total en caliente
        detalles_preparados = []
        monto_total_ticket = Decimal('0.00')

        for i in range(len(rubro_ids)):
            try:
                rubro = RubroCantina.objects.get(id=rubro_ids[i], colegio=colegio, activo=True)
                qty = int(cantidades[i])
                precio_manual = Decimal(precios[i])

                if qty <= 0 or precio_manual < 0:
                    raise ValueError

                subtotal = qty * precio_manual
                monto_total_ticket += subtotal

                # Guardamos temporalmente en memoria los datos validados del detalle
                detalles_preparados.append({
                    'rubro': rubro,
                    'cantidad': qty,
                    'precio_unitario': precio_manual
                })

            except (RubroCantina.DoesNotExist, ValueError, InvalidOperation):
                messages.error(request, f"Error de validación en el producto de la línea {i+1}. Verifique precios y cantidades.")
                return redirect('dashboard_cantina', colegio_slug=colegio.slug)

        # 4. Validar disponibilidad de saldo prepago
        if billetera.saldo < monto_total_ticket:
            messages.error(
                request, 
                f"Saldo insuficiente para {billetera.estudiante.nombre}. "
                f"Total Compra: {monto_total_ticket} BS | Saldo Actual: {billetera.saldo} BS."
            )
            return redirect('dashboard_cantina', colegio_slug=colegio.slug)

        # 5. Todo legal: Procedemos a escribir en la Base de Datos
        # Obtenemos el perfil 'Persona' del usuario autenticado (el cantinero)
        cantinero_perfil = getattr(request.user, 'perfil', None)

        # A) Crear la cabecera de la Venta
        venta = VentaCantina.objects.create(
            billetera=billetera,
            monto_total=monto_total_ticket,
            registrado_por=cantinero_perfil
        )

        # B) Crear los renglones del detalle
        for item in detalles_preparados:
            DetalleVentaCantina.objects.create(
                venta=venta,
                rubro=item['rubro'],
                cantidad=item['cantidad'],
                precio_unitario=item['precio_unitario']
            )

        # C) Descontar el saldo de la billetera de forma segura
        billetera.saldo -= monto_total_ticket
        billetera.save()

        messages.success(
            request, 
            f"¡Consumo cobrado con éxito! Total: {monto_total_ticket} BS. "
            f"Nuevo saldo de {billetera.estudiante.nombre}: {billetera.saldo} BS."
        )
        return redirect('dashboard_cantina', colegio_slug=colegio.slug)

    # Si entran por GET, simplemente redirige al panel
    return redirect('dashboard_cantina', colegio_slug=colegio_slug)

@login_required
def guardar_rubro_cantina_view(request, colegio_slug): 
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    
    # Validación de Rol (Solo administradores, superusuarios o cantineros calificados)
    if not request.user.is_superuser and (not request.user.rol or request.user.rol.nombre not in ['Cantinero', 'Admin', 'Super']):
        raise PermissionDenied

    if request.method == 'POST':
        rubro_id = request.POST.get('rubro_id')
        nombre = request.POST.get('nombre').strip()
        activo = request.POST.get('activo') == 'true'

        if rubro_id:
            # Operación: EDICIÓN
            rubro = get_object_or_404(RubroCantina, id=rubro_id, colegio=colegio)
            rubro.nombre = nombre
            rubro.activo = activo
            rubro.save()
            messages.success(request, f"Producto '{nombre}' actualizado con éxito.")
        else:
            # Operación: CREACIÓN
            RubroCantina.objects.create(
                colegio=colegio,
                nombre=nombre,
                activo=activo
            )
            messages.success(request, f"Producto '{nombre}' registrado en el catálogo.")

    url_destino = reverse('dashboard_cantina', kwargs={'colegio_slug': colegio.slug})
    return redirect(f"{url_destino}?tab=catalogo")

@login_required
def cantina_estudiante_view(request, colegio_slug, estudiante_id): 
    colegio = get_object_or_404(Colegio, slug=colegio_slug)
    estudiante = get_object_or_404(Persona, id=estudiante_id, colegio=colegio, es_estudiante=True)
    
    if not request.user.is_superuser:
        if not request.user.perfil.es_representante:
            raise PermissionDenied
        
        es_su_representado = RelacionFamiliar.objects.filter(
            representante=request.user.perfil,
            estudiante=estudiante
        ).exists()
        
        if not es_su_representado:
            raise PermissionDenied

    billetera = get_object_or_404(BilleteraCantina, estudiante=estudiante)
    
    inscripcion_activa = Inscripcion.objects.filter(
        estudiante=estudiante, 
        anio_escolar__activo=True, 
        estado='ACTIVO'
    ).select_related('seccion').first()

    # 📥 PROCESAR FORMULARIO DE RECARGA (POST)
    if request.method == 'POST':
        monto_str = request.POST.get('monto')
        referencia = request.POST.get('referencia')
        metodo_pago = request.POST.get('metodo_pago')
        
        if monto_str and referencia and metodo_pago:
            try:
                # 🛠️ CORRECCIÓN: Uso estricto de Decimal para no perder céntimos
                monto = Decimal(monto_str)
                if monto <= 0:
                    raise ValueError
                
                cantinero_comprobante = Persona.objects.filter(colegio=colegio, es_cantina=True, activo=True).first()
                if not cantinero_comprobante:
                    cantinero_comprobante = request.user.perfil 

                # Protegemos la inserción y la actualización automática del saldo
                with transaction.atomic():
                    # Volvemos a solicitar la billetera con un bloqueo ligero para evitar colisiones
                    billetera_bloqueada = BilleteraCantina.objects.select_for_update().get(id=billetera.id)
                    
                    # Al crearse, el método save() que añadimos en RecargaBilletera sumará el dinero al saldo automáticamente.
                    RecargaBilletera.objects.create(
                        billetera=billetera_bloqueada,
                        representante=request.user.perfil,
                        monto=monto,
                        metodo_pago=metodo_pago,
                        referencia=referencia,
                        registrado_por=cantinero_comprobante
                    )

                    billetera_bloqueada.saldo += monto
                    billetera_bloqueada.save()
                
                messages.success(
                    request, 
                    f"¡Reporte enviado con éxito! Se han abonado Bs. {monto:.2f} al saldo del estudiante."
                )
            except (ValueError, InvalidOperation):
                messages.error(request, "El monto ingresado no cuenta con un formato numérico válido o es menor/igual a cero.")
            
            return redirect('cantina_estudiante', colegio_slug=colegio.slug, estudiante_id=estudiante.id)

    # 📊 HISTORIAL UNIFICADO (Mantiene tu excelente lógica de ordenamiento manual)
    recargas_query = RecargaBilletera.objects.filter(billetera=billetera)
    ventas_query = VentaCantina.objects.filter(billetera=billetera).prefetch_related('detalles__rubro')
    
    movimientos = []
    
    for recarga in recargas_query:
        movimientos.append({
            'fecha': recarga.fecha_registro,
            'tipo': 'RECARGA',
            'concepto': f"Recarga de Saldo ({recarga.get_metodo_pago_display()})",
            'rubro_info': f"Ref: {recarga.referencia or 'N/A'}",
            'monto': recarga.monto
        })
        
    for venta in ventas_query:
        nombres_rubros = [detalle.rubro.nombre for detalle in venta.detalles.all()]
        lista_rubros = ", ".join(nombres_rubros) if nombres_rubros else "Consumo Cantina"
        
        movimientos.append({
            'fecha': venta.fecha_venta,
            'tipo': 'CONSUMO',
            'concepto': f"Consumo de Alimentos (Ticket #{venta.id})",
            'rubro_info': lista_rubros,
            'monto': venta.monto_total
        })
        
    movimientos.sort(key=lambda x: x['fecha'], reverse=True)

    context = {
        'colegio': colegio,
        'estudiante': estudiante,
        'inscripcion': inscripcion_activa,
        'saldo_billetera': billetera.saldo,
        'movimientos': movimientos,
    }
    return render(request, 'users/estado_cuenta_billetera.html', context)